"""
Tests for subcontractor pricing helpers.
Validates load_contract_rates() behavior and SUBCONTRACTOR_SHEET_IDS configuration.
"""

import os
import csv
import importlib
import tempfile
import unittest
from unittest import mock
import generate_weekly_pdfs


class TestLoadContractRates(unittest.TestCase):
    """Tests for the load_contract_rates helper function."""

    def test_loads_valid_csv(self):
        """Test loading a well-formed CSV returns correct rate dictionary."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='') as f:
            writer = csv.writer(f)
            writer.writerow([
                'CU WBS #', 'CU', 'Unit Of Measure', 'Description',
                'Compatible Unit Group', 'Install Hours', 'Removal Hours',
                'Transfer Hours', 'Install Price', 'Removal Price', 'Transfer Price'
            ])
            writer.writerow(['100', 'ABC123', 'EA', 'Test Unit', 'Group1', '1', '0.5', '0.3', '$150.00', '$75.00', '$50.00'])
            writer.writerow(['101', 'DEF456', 'LF', 'Another Unit', 'Group2', '2', '1', '0.5', '200', '100', '60'])
            tmp_path = f.name

        try:
            rates = generate_weekly_pdfs.load_contract_rates(tmp_path)
            self.assertEqual(len(rates), 2)
            self.assertIn('ABC123', rates)
            self.assertIn('DEF456', rates)
            self.assertAlmostEqual(rates['ABC123']['install'], 150.0)
            self.assertAlmostEqual(rates['ABC123']['removal'], 75.0)
            self.assertAlmostEqual(rates['ABC123']['transfer'], 50.0)
            self.assertAlmostEqual(rates['DEF456']['install'], 200.0)
        finally:
            os.unlink(tmp_path)

    def test_missing_file_returns_empty(self):
        """Test that a missing CSV file returns an empty dict gracefully."""
        rates = generate_weekly_pdfs.load_contract_rates('/nonexistent/path.csv')
        self.assertEqual(rates, {})

    def test_missing_file_is_benign_not_error(self):
        """Missing rate CSV must NOT emit ERROR-level log (benign INFO skip, not a Sentry event)."""
        with self.assertNoLogs(level="ERROR"):
            rates = generate_weekly_pdfs.load_contract_rates("/nonexistent/path.csv")
        self.assertEqual(rates, {})

    def test_empty_csv_returns_empty(self):
        """Test that a CSV with only headers returns an empty dict."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='') as f:
            writer = csv.writer(f)
            writer.writerow([
                'CU WBS #', 'CU', 'Unit Of Measure', 'Description',
                'Compatible Unit Group', 'Install Hours', 'Removal Hours',
                'Transfer Hours', 'Install Price', 'Removal Price', 'Transfer Price'
            ])
            tmp_path = f.name

        try:
            rates = generate_weekly_pdfs.load_contract_rates(tmp_path)
            self.assertEqual(len(rates), 0)
        finally:
            os.unlink(tmp_path)

    def test_cu_uppercased(self):
        """Test that CU codes are normalized to uppercase."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='') as f:
            writer = csv.writer(f)
            writer.writerow([
                'CU WBS #', 'CU', 'Unit Of Measure', 'Description',
                'Compatible Unit Group', 'Install Hours', 'Removal Hours',
                'Transfer Hours', 'Install Price', 'Removal Price', 'Transfer Price'
            ])
            writer.writerow(['100', 'abc123', 'EA', 'Test', 'G1', '1', '0', '0', '100', '50', '25'])
            tmp_path = f.name

        try:
            rates = generate_weekly_pdfs.load_contract_rates(tmp_path)
            self.assertIn('ABC123', rates)
            self.assertNotIn('abc123', rates)
        finally:
            os.unlink(tmp_path)

    def test_handles_invalid_price_values(self):
        """Test that non-numeric price values default to 0.0."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='') as f:
            writer = csv.writer(f)
            writer.writerow([
                'CU WBS #', 'CU', 'Unit Of Measure', 'Description',
                'Compatible Unit Group', 'Install Hours', 'Removal Hours',
                'Transfer Hours', 'Install Price', 'Removal Price', 'Transfer Price'
            ])
            writer.writerow(['100', 'BAD1', 'EA', 'Test', 'G1', '1', '0', '0', 'N/A', '', 'error'])
            tmp_path = f.name

        try:
            rates = generate_weekly_pdfs.load_contract_rates(tmp_path)
            self.assertIn('BAD1', rates)
            self.assertAlmostEqual(rates['BAD1']['install'], 0.0)
            self.assertAlmostEqual(rates['BAD1']['removal'], 0.0)
            self.assertAlmostEqual(rates['BAD1']['transfer'], 0.0)
        finally:
            os.unlink(tmp_path)

    def test_skips_blank_cu_rows(self):
        """Test that rows with empty CU field are skipped."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='') as f:
            writer = csv.writer(f)
            writer.writerow([
                'CU WBS #', 'CU', 'Unit Of Measure', 'Description',
                'Compatible Unit Group', 'Install Hours', 'Removal Hours',
                'Transfer Hours', 'Install Price', 'Removal Price', 'Transfer Price'
            ])
            writer.writerow(['100', '', 'EA', 'Blank CU', 'G1', '1', '0', '0', '100', '50', '25'])
            writer.writerow(['101', 'VALID', 'EA', 'Valid CU', 'G1', '1', '0', '0', '200', '100', '50'])
            tmp_path = f.name

        try:
            rates = generate_weekly_pdfs.load_contract_rates(tmp_path)
            self.assertEqual(len(rates), 1)
            self.assertIn('VALID', rates)
        finally:
            os.unlink(tmp_path)

    def test_loads_both_contract_files(self):
        """Test that two separate CSVs load independently with different rates."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='') as f1:
            writer = csv.writer(f1)
            writer.writerow([
                'CU WBS #', 'CU', 'Unit Of Measure', 'Description',
                'Compatible Unit Group', 'Install Hours', 'Removal Hours',
                'Transfer Hours', 'Install Price', 'Removal Price', 'Transfer Price'
            ])
            writer.writerow(['100', 'CU001', 'EA', 'Original', 'G1', '1', '0', '0', '100.00', '50.00', '25.00'])
            path1 = f1.name

        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='') as f2:
            writer = csv.writer(f2)
            writer.writerow([
                'CU WBS #', 'CU', 'Unit Of Measure', 'Description',
                'Compatible Unit Group', 'Install Hours', 'Removal Hours',
                'Transfer Hours', 'Install Price', 'Removal Price', 'Transfer Price'
            ])
            writer.writerow(['100', 'CU001', 'EA', 'Reduced', 'G1', '1', '0', '0', '90.00', '45.00', '22.50'])
            path2 = f2.name

        try:
            original = generate_weekly_pdfs.load_contract_rates(path1)
            contractor = generate_weekly_pdfs.load_contract_rates(path2)
            self.assertAlmostEqual(original['CU001']['install'], 100.0)
            self.assertAlmostEqual(contractor['CU001']['install'], 90.0)
        finally:
            os.unlink(path1)
            os.unlink(path2)


# Canonical 17-column header for the subcontractor rates CSV. Pinned at
# module scope so every TestLoadSubcontractorRates fixture emits the
# same shape and a future header drift fails one test instead of
# silently mis-feeding the loader across every test.
SUBCONTRACTOR_HEADERS = [
    'CU WBS #', 'CU', 'Unit Of Measure', 'Description',
    'Compatible Unit Group', 'Install Hours', 'Removal Hours',
    'Transfer Hours',
    'Install Price (Subcontractor Rates)',
    'Removal Price (Subcontractor Rates)',
    'Transfer Price (Subcontractor Rates)',
    'Install Price (Old Rates)',
    'Removal Price (Old Rates)',
    'Transfer Price (Old Rates)',
    'Install Price (New Rates)',
    'Removal Price (New Rates)',
    'Transfer Price (New Rates)',
]


class TestLoadSubcontractorRates(unittest.TestCase):
    """Regression class for ``load_subcontractor_rates`` (Phase 1
    plan 01-01). Covers decisions D-04..D-07 + D-20."""

    def _write_csv(self, rows, *, write_bom: bool = False) -> str:
        """Write a temp CSV with the canonical 17-column header and the
        supplied data rows. Returns the temp path; the caller is
        responsible for ``os.unlink`` in a ``finally`` block. When
        ``write_bom`` is True a UTF-8 BOM is emitted at file start so
        the ``utf-8-sig`` tolerance can be tested.
        """
        with tempfile.NamedTemporaryFile(
            mode='w', suffix='.csv', delete=False, newline='',
            encoding='utf-8',
        ) as f:
            if write_bom:
                f.write('﻿')
            writer = csv.writer(f)
            writer.writerow(SUBCONTRACTOR_HEADERS)
            for row in rows:
                writer.writerow(row)
            return f.name

    def test_loads_subcontractor_csv_with_currency_strings(self):
        """D-04: ``$150.00`` / ``$1,234.56`` currency cells parse to
        floats via ``parse_price``."""
        tmp_path = self._write_csv([
            [
                '100', 'ABC123', 'EA', 'Test', 'G1',
                '1', '0.5', '0.3',
                '$150.00', '$1,234.56', '$50.00',
                '$0.00', '$0.00', '$0.00',
                '$200.00', '$1,500.00', '$75.00',
            ],
        ])
        try:
            rates = generate_weekly_pdfs.load_subcontractor_rates(tmp_path)
            self.assertIn('ABC123', rates)
            self.assertAlmostEqual(rates['ABC123']['reduced_install_price'], 150.0)
            self.assertAlmostEqual(rates['ABC123']['reduced_remove_price'], 1234.56)
            self.assertAlmostEqual(rates['ABC123']['reduced_transfer_price'], 50.0)
            self.assertAlmostEqual(rates['ABC123']['new_install_price'], 200.0)
            self.assertAlmostEqual(rates['ABC123']['new_remove_price'], 1500.0)
            self.assertAlmostEqual(rates['ABC123']['new_transfer_price'], 75.0)
        finally:
            os.unlink(tmp_path)

    def test_loads_subcontractor_csv_with_utf8_bom(self):
        """D-04: a UTF-8 BOM at file start does not break header
        detection (``encoding='utf-8-sig'`` strips it)."""
        tmp_path = self._write_csv(
            [
                [
                    '100', 'BOM-CU', 'EA', 'Test', 'G1',
                    '1', '0.5', '0.3',
                    '$10.00', '$5.00', '$3.00',
                    '$0.00', '$0.00', '$0.00',
                    '$15.00', '$7.00', '$4.00',
                ],
            ],
            write_bom=True,
        )
        try:
            rates = generate_weekly_pdfs.load_subcontractor_rates(tmp_path)
            self.assertIn('BOM-CU', rates)
            self.assertAlmostEqual(rates['BOM-CU']['reduced_install_price'], 10.0)
            self.assertAlmostEqual(rates['BOM-CU']['new_install_price'], 15.0)
        finally:
            os.unlink(tmp_path)

    def test_skips_all_zero_priced_rows(self):
        """D-04: rows whose all six priced cells are zero are
        excluded from the dict (placeholder CUs)."""
        tmp_path = self._write_csv([
            # Row 1: all-zero priced — must be skipped
            [
                '100', 'ZERO-CU', 'EA', 'Placeholder', 'G1',
                '0', '0', '0',
                '$0.00', '$0.00', '$0.00',
                '$0.00', '$0.00', '$0.00',
                '$0.00', '$0.00', '$0.00',
            ],
            # Row 2: valid priced — must be loaded
            [
                '101', 'VALID-CU', 'EA', 'Real', 'G1',
                '1', '0.5', '0.3',
                '$45.95', '$33.33', '$106.54',
                '$0.00', '$0.00', '$0.00',
                '$52.58', '$38.14', '$121.93',
            ],
        ])
        try:
            rates = generate_weekly_pdfs.load_subcontractor_rates(tmp_path)
            self.assertEqual(len(rates), 1)
            self.assertNotIn('ZERO-CU', rates)
            self.assertIn('VALID-CU', rates)
        finally:
            os.unlink(tmp_path)

    def test_tolerates_na_in_hours_columns(self):
        """D-04: ``'N/A'`` in Hours columns does not break the loader
        (hours are not read at all). Row 2 of the production CSV
        (``ADDITEM-ROW-PURCHASE``) is shaped exactly like this."""
        tmp_path = self._write_csv([
            [
                '100', 'NA-HOURS', 'EA', 'Test', 'G1',
                'N/A', 'N/A', 'N/A',
                '$100.00', '$50.00', '$25.00',
                '$0.00', '$0.00', '$0.00',
                '$120.00', '$60.00', '$30.00',
            ],
        ])
        try:
            rates = generate_weekly_pdfs.load_subcontractor_rates(tmp_path)
            self.assertIn('NA-HOURS', rates)
            self.assertAlmostEqual(rates['NA-HOURS']['reduced_install_price'], 100.0)
            self.assertAlmostEqual(rates['NA-HOURS']['new_install_price'], 120.0)
        finally:
            os.unlink(tmp_path)

    def test_loads_per_cu_rate_variance_literally(self):
        """D-07: per-CU rate variance is real (median ``New/Old =
        1.0300``, min ``1.0244``; median ``Reduced/New = 0.8738``,
        min ``0.4343``). The loader MUST read literal values, never
        compute ``reduced = old × 0.87`` or ``new = old × 1.03``
        shortcuts."""
        # Outlier CU: New/Old = 2.0725 (max), Reduced/New = 0.4343 (min).
        # If the loader computed shortcuts, reduced_install would be
        # 0.87 × 20.73 ≈ 18.03 (not the literal 9.00), and new_install
        # would be 1.03 × 10.00 = 10.30 (not the literal 20.73).
        tmp_path = self._write_csv([
            [
                '100', 'OUTLIER', 'EA', 'Test', 'G1',
                '1', '0', '0',
                '$9.00', '$5.00', '$3.00',     # reduced
                '$10.00', '$5.00', '$3.00',    # old
                '$20.73', '$10.30', '$6.18',   # new
            ],
        ])
        try:
            rates = generate_weekly_pdfs.load_subcontractor_rates(tmp_path)
            self.assertIn('OUTLIER', rates)
            # Reduced must be the literal $9.00, not 0.87 × $10.00
            self.assertAlmostEqual(rates['OUTLIER']['reduced_install_price'], 9.0)
            # New must be the literal $20.73, not 1.03 × $10.00
            self.assertAlmostEqual(rates['OUTLIER']['new_install_price'], 20.73)
            # Compute ratios to confirm the outliers landed verbatim
            ratio_reduced_over_new = (
                rates['OUTLIER']['reduced_install_price']
                / rates['OUTLIER']['new_install_price']
            )
            self.assertAlmostEqual(ratio_reduced_over_new, 0.4341, places=3)
        finally:
            os.unlink(tmp_path)

    def test_old_rates_columns_not_loaded(self):
        """D-06: Old-Rates columns (12-14) are NOT loaded into the
        per-CU value dict. Carrying them would create a 3rd source of
        truth for pricing — explicitly forbidden by the design."""
        tmp_path = self._write_csv([
            [
                '100', 'HAS-OLD', 'EA', 'Test', 'G1',
                '1', '0', '0',
                '$45.95', '$33.33', '$106.54',
                # Old-Rates: deliberately distinct values so the test
                # would catch a key like ``'old_install_price'`` if
                # the loader regressed and included them
                '$999.00', '$888.00', '$777.00',
                '$52.58', '$38.14', '$121.93',
            ],
        ])
        try:
            rates = generate_weekly_pdfs.load_subcontractor_rates(tmp_path)
            self.assertIn('HAS-OLD', rates)
            value = rates['HAS-OLD']
            # No legacy or alternate-name keys for the old-rates columns
            self.assertNotIn('install_price_old', value)
            self.assertNotIn('old_install_price', value)
            self.assertNotIn('removal_price_old', value)
            self.assertNotIn('old_removal_price', value)
            self.assertNotIn('transfer_price_old', value)
            self.assertNotIn('old_transfer_price', value)
            # Defensive: no key in the value dict contains 'old'
            old_keys = [k for k in value.keys() if 'old' in k.lower()]
            self.assertEqual(old_keys, [], f"Found Old-Rates keys: {old_keys}")
            # Old-Rates values 999/888/777 must not appear anywhere
            for v in value.values():
                if isinstance(v, (int, float)):
                    self.assertNotAlmostEqual(v, 999.0)
                    self.assertNotAlmostEqual(v, 888.0)
                    self.assertNotAlmostEqual(v, 777.0)
        finally:
            os.unlink(tmp_path)

    def test_subcontractor_rates_fingerprint_deterministic(self):
        """D-20: two byte-identical inputs (different dict insertion
        order included) MUST produce the same 16-char fingerprint."""
        d1 = {
            'CU-A': {
                'cu_code': 'CU-A', 'cu_wbs': '', 'compatible_unit_group': '',
                'reduced_install_price': 10.0, 'reduced_remove_price': 5.0,
                'reduced_transfer_price': 3.0,
                'new_install_price': 12.0, 'new_remove_price': 6.0,
                'new_transfer_price': 4.0,
            },
            'CU-B': {
                'cu_code': 'CU-B', 'cu_wbs': '', 'compatible_unit_group': '',
                'reduced_install_price': 20.0, 'reduced_remove_price': 8.0,
                'reduced_transfer_price': 5.0,
                'new_install_price': 22.0, 'new_remove_price': 9.0,
                'new_transfer_price': 6.0,
            },
        }
        # Reverse insertion order — fingerprint must be identical
        # because the helper sorts keys before hashing.
        d2 = {
            'CU-B': dict(d1['CU-B']),
            'CU-A': dict(d1['CU-A']),
        }
        fp1 = generate_weekly_pdfs._compute_subcontractor_rates_fingerprint(d1)
        fp2 = generate_weekly_pdfs._compute_subcontractor_rates_fingerprint(d2)
        self.assertEqual(fp1, fp2)
        self.assertEqual(len(fp1), 16)
        # Output charset is hex
        self.assertRegex(fp1, r'^[0-9a-f]{16}$')

    def test_subcontractor_rates_fingerprint_changes_on_edit(self):
        """D-20: editing one CU's priced field (any of the six) MUST
        change the fingerprint."""
        base = {
            'CU-EDIT': {
                'cu_code': 'CU-EDIT', 'cu_wbs': '', 'compatible_unit_group': '',
                'reduced_install_price': 10.0, 'reduced_remove_price': 5.0,
                'reduced_transfer_price': 3.0,
                'new_install_price': 12.0, 'new_remove_price': 6.0,
                'new_transfer_price': 4.0,
            },
        }
        fp_base = generate_weekly_pdfs._compute_subcontractor_rates_fingerprint(base)

        mutated = {
            'CU-EDIT': dict(base['CU-EDIT'], new_install_price=12.01),
        }
        fp_mutated = generate_weekly_pdfs._compute_subcontractor_rates_fingerprint(mutated)
        self.assertNotEqual(fp_base, fp_mutated)
        self.assertEqual(len(fp_mutated), 16)


class TestSubcontractorSheetIdsConfig(unittest.TestCase):
    """Test SUBCONTRACTOR_SHEET_IDS configuration parsing."""

    def test_default_is_empty_set(self):
        """Verify that SUBCONTRACTOR_SHEET_IDS is a set attribute on the module."""
        self.assertIsInstance(generate_weekly_pdfs.SUBCONTRACTOR_SHEET_IDS, set)

    def test_parse_sheet_ids_skips_invalid(self):
        """Verify _parse_sheet_ids gracefully skips non-integer tokens."""
        result = generate_weekly_pdfs._parse_sheet_ids('123,abc,456,,  789  ')
        self.assertEqual(result, [123, 456, 789])

    def test_parse_sheet_ids_empty_string(self):
        """Verify _parse_sheet_ids returns empty list for empty string."""
        result = generate_weekly_pdfs._parse_sheet_ids('')
        self.assertEqual(result, [])


class TestRevertSubcontractorPrice(unittest.TestCase):
    """Tests for the revert_subcontractor_price helper function."""

    def setUp(self):
        self.rates = {
            'CU-INSTALL': {'install': 100.0, 'removal': 50.0, 'transfer': 30.0},
            'CU-MULTI': {'install': 200.0, 'removal': 80.0, 'transfer': 60.0},
        }

    def test_basic_install_reversion(self):
        """Test that install price is recalculated from original rate × quantity."""
        row = {'CU': 'CU-INSTALL', 'Work Type': 'Install', 'Quantity': '3', 'Units Total Price': '$270.00'}
        result = generate_weekly_pdfs.revert_subcontractor_price(row, self.rates)
        self.assertAlmostEqual(result, 300.0)
        self.assertAlmostEqual(row['Units Total Price'], 300.0)

    def test_removal_work_type(self):
        """Test that 'Removal' work type maps to removal rates."""
        row = {'CU': 'CU-INSTALL', 'Work Type': 'Removal', 'Quantity': '2', 'Units Total Price': '$90.00'}
        result = generate_weekly_pdfs.revert_subcontractor_price(row, self.rates)
        self.assertAlmostEqual(result, 100.0)

    def test_transfer_work_type(self):
        """Test that 'Transfer' work type maps to transfer rates."""
        row = {'CU': 'CU-INSTALL', 'Work Type': 'Transfer', 'Quantity': '4', 'Units Total Price': '$108.00'}
        result = generate_weekly_pdfs.revert_subcontractor_price(row, self.rates)
        self.assertAlmostEqual(result, 120.0)

    def test_xfr_work_type(self):
        """Test that 'xfr' in work type maps to transfer rates."""
        row = {'CU': 'CU-INSTALL', 'Work Type': 'XFR', 'Quantity': '1', 'Units Total Price': '$27.00'}
        result = generate_weekly_pdfs.revert_subcontractor_price(row, self.rates)
        self.assertAlmostEqual(result, 30.0)

    def test_cu_helper_preferred_over_cu(self):
        """Test that CU Helper field is preferred over CU field."""
        row = {'CU Helper': 'CU-MULTI', 'CU': 'CU-INSTALL', 'Work Type': 'Install', 'Quantity': '1', 'Units Total Price': '$90.00'}
        result = generate_weekly_pdfs.revert_subcontractor_price(row, self.rates)
        self.assertAlmostEqual(result, 200.0)

    def test_nan_cu_helper_falls_back_to_cu(self):
        """Test that NaN CU Helper falls back to CU field."""
        row = {'CU Helper': 'nan', 'CU': 'CU-INSTALL', 'Work Type': 'Install', 'Quantity': '2', 'Units Total Price': '$180.00'}
        result = generate_weekly_pdfs.revert_subcontractor_price(row, self.rates)
        self.assertAlmostEqual(result, 200.0)

    def test_unknown_cu_returns_parsed_price(self):
        """Test that unknown CU code returns the original parsed price unchanged."""
        row = {'CU': 'UNKNOWN-CU', 'Work Type': 'Install', 'Quantity': '1', 'Units Total Price': '$55.00'}
        result = generate_weekly_pdfs.revert_subcontractor_price(row, self.rates)
        self.assertAlmostEqual(result, 55.0)
        self.assertEqual(row['Units Total Price'], '$55.00')

    def test_zero_quantity(self):
        """Test that zero quantity yields zero price."""
        row = {'CU': 'CU-INSTALL', 'Work Type': 'Install', 'Quantity': '0', 'Units Total Price': '$0.00'}
        result = generate_weekly_pdfs.revert_subcontractor_price(row, self.rates)
        self.assertAlmostEqual(result, 0.0)


def _make_children_page(sheet_ids=(), subfolder_ids=(), last_key=None):
    """Build a MagicMock paginated children result containing real Sheet/Folder instances."""
    from unittest.mock import MagicMock
    from smartsheet.models.sheet import Sheet
    from smartsheet.models.folder import Folder
    data = [Sheet({'id': sid, 'name': f'sheet-{sid}'}) for sid in sheet_ids]
    data += [Folder({'id': fid, 'name': f'folder-{fid}'}) for fid in subfolder_ids]
    page = MagicMock()
    page.data = data
    page.last_key = last_key
    return page


class TestDiscoverFolderSheets(unittest.TestCase):
    """Tests for folder-based sheet discovery."""

    def test_discover_folder_sheets_returns_set(self):
        """Test discover_folder_sheets returns a set of ints."""
        from unittest.mock import MagicMock
        mock_client = MagicMock()
        mock_client.Folders.get_folder_children.return_value = _make_children_page(sheet_ids=[111, 222])

        result = generate_weekly_pdfs.discover_folder_sheets(mock_client, [9999], 'test')
        self.assertEqual(result, {111, 222})
        mock_client.Folders.get_folder_children.assert_called_once()
        call_args = mock_client.Folders.get_folder_children.call_args
        self.assertEqual(call_args.args[0], 9999)

    def test_discover_folder_sheets_handles_api_error(self):
        """Test graceful handling when folder API call fails."""
        from unittest.mock import MagicMock
        mock_client = MagicMock()
        mock_client.Folders.get_folder_children.side_effect = Exception("API error")

        result = generate_weekly_pdfs.discover_folder_sheets(mock_client, [9999], 'test')
        self.assertEqual(result, set())

    def test_discover_folder_sheets_multiple_folders(self):
        """Test discovery across multiple folder IDs with deduplication."""
        from unittest.mock import MagicMock
        mock_client = MagicMock()
        mock_client.Folders.get_folder_children.side_effect = [
            _make_children_page(sheet_ids=[100, 200]),
            _make_children_page(sheet_ids=[100]),  # duplicate 100
        ]

        result = generate_weekly_pdfs.discover_folder_sheets(mock_client, [1, 2], 'test')
        self.assertEqual(result, {100, 200})

    def test_discover_folder_sheets_empty_list(self):
        """Test with no folder IDs returns empty set."""
        from unittest.mock import MagicMock
        mock_client = MagicMock()
        result = generate_weekly_pdfs.discover_folder_sheets(mock_client, [], 'test')
        self.assertEqual(result, set())

    def test_discover_folder_sheets_paginates_last_key(self):
        """Multi-page last_key pagination is followed until last_key is falsy."""
        from unittest.mock import MagicMock
        mock_client = MagicMock()
        # Page 1 returns two sheets + a continuation token; page 2 returns one more and terminates.
        mock_client.Folders.get_folder_children.side_effect = [
            _make_children_page(sheet_ids=[301, 302], last_key='k1'),
            _make_children_page(sheet_ids=[303], last_key=None),
        ]

        result = generate_weekly_pdfs.discover_folder_sheets(mock_client, [4242], 'test')
        self.assertEqual(result, {301, 302, 303})
        self.assertEqual(mock_client.Folders.get_folder_children.call_count, 2)
        # The second call should forward the last_key from the first response.
        second_kwargs = mock_client.Folders.get_folder_children.call_args_list[1].kwargs
        self.assertEqual(second_kwargs.get('last_key'), 'k1')

    def test_discover_folder_sheets_recurses_into_subfolders(self):
        """Folder children returned as subfolders trigger recursive discovery."""
        from unittest.mock import MagicMock
        mock_client = MagicMock()

        def _children(fid, **kwargs):
            if fid == 10:
                # Top folder has one sheet and one subfolder child
                return _make_children_page(sheet_ids=[401], subfolder_ids=[11])
            if fid == 11:
                # Subfolder has two sheets and no further nesting
                return _make_children_page(sheet_ids=[402, 403])
            return _make_children_page()

        mock_client.Folders.get_folder_children.side_effect = _children

        result = generate_weekly_pdfs.discover_folder_sheets(mock_client, [10], 'test')
        self.assertEqual(result, {401, 402, 403})
        called_ids = [c.args[0] for c in mock_client.Folders.get_folder_children.call_args_list]
        self.assertIn(10, called_ids)
        self.assertIn(11, called_ids)

    def test_discover_folder_sheets_stops_on_repeated_last_key(self):
        """A repeated last_key must short-circuit pagination to avoid an API burst."""
        from unittest.mock import MagicMock
        mock_client = MagicMock()
        # A misbehaving API keeps returning the same continuation token forever.
        # The discovery loop should stop after detecting the repeat rather than
        # calling through to max_pages.
        mock_client.Folders.get_folder_children.side_effect = [
            _make_children_page(sheet_ids=[501], last_key='stuck'),
            _make_children_page(sheet_ids=[502], last_key='stuck'),
            _make_children_page(sheet_ids=[503], last_key='stuck'),
        ]

        result = generate_weekly_pdfs.discover_folder_sheets(mock_client, [7777], 'test')
        # Sheets from pages fetched before the repeat-stop are preserved.
        self.assertEqual(result, {501, 502})
        # Exactly 2 calls: page 1 (token 'stuck' recorded), page 2 (token repeats → stop).
        self.assertEqual(mock_client.Folders.get_folder_children.call_count, 2)

    def test_discover_folder_sheets_stops_at_max_pages(self):
        """Pagination must terminate at the 100-page safety cap."""
        from unittest.mock import MagicMock
        mock_client = MagicMock()
        # Generate a unique last_key per call so the repeated-token guard never trips —
        # only the max_pages ceiling can terminate the loop.
        counter = {'n': 0}

        def _children(fid, **kwargs):
            counter['n'] += 1
            return _make_children_page(
                sheet_ids=[1000 + counter['n']],
                last_key=f"token-{counter['n']}",
            )

        mock_client.Folders.get_folder_children.side_effect = _children

        result = generate_weekly_pdfs.discover_folder_sheets(mock_client, [8888], 'test')
        # Exactly max_pages (100) calls — not unbounded.
        self.assertEqual(mock_client.Folders.get_folder_children.call_count, 100)
        self.assertEqual(len(result), 100)


class TestIdentityNormalization(unittest.TestCase):
    """Tests for the None vs '' identity comparison fix."""

    def test_none_equals_empty_string_after_normalization(self):
        """Verify that (None or '') == ('' or '') is True."""
        ident_identifier = None
        identifier = ''
        self.assertEqual((ident_identifier or ''), (identifier or ''))

    def test_non_empty_identifiers_still_match(self):
        """Verify that real identifiers still match correctly."""
        ident_identifier = 'John|Dept1|Job1'
        identifier = 'John|Dept1|Job1'
        self.assertEqual((ident_identifier or ''), (identifier or ''))

    def test_different_identifiers_do_not_match(self):
        """Verify that different identifiers don't match."""
        ident_identifier = 'John|Dept1|Job1'
        identifier = 'Jane|Dept2|Job2'
        self.assertNotEqual((ident_identifier or ''), (identifier or ''))


class TestLoadNewContractRates(unittest.TestCase):
    """Tests for loading the 2026-format new contract rates CSV."""

    def test_loads_new_format_csv(self):
        """Test loading a CSV with 3 metadata rows and positional columns."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='') as f:
            writer = csv.writer(f)
            # 3 metadata rows
            writer.writerow(['', '', '', '', '', '', '2026 Update', '', '0.03'])
            writer.writerow(['', '', '', '', '', '', 'Revised Pricing', '', ''])
            writer.writerow(['', '', '', '', '', '', 'Install', 'Remove ', 'Transfer'])
            # Data rows
            writer.writerow(['ANC-H', 'Anchor Assembly (Hand)', 'EA', 'Overhead', 'AEP TX', '01-18-26', '814.28', '29.45', '0'])
            writer.writerow(['ANC-M', 'Anchor Assembly (Machine)', 'EA', 'Overhead', 'AEP TX', '01-18-26', '224.06', '29.46', '0'])
            writer.writerow(['ARM-DW', 'Double Wood Crossarm', 'EA', 'Overhead', 'AEP TX', '01-18-26', '330.24', '75.94', '183.98'])
            tmp_path = f.name

        try:
            rates = generate_weekly_pdfs.load_new_contract_rates(tmp_path)
            self.assertEqual(len(rates), 3)
            self.assertIn('ANC-H', rates)
            self.assertIn('ANC-M', rates)
            self.assertIn('ARM-DW', rates)
            self.assertAlmostEqual(rates['ANC-H']['install'], 814.28)
            self.assertAlmostEqual(rates['ANC-H']['removal'], 29.45)
            self.assertAlmostEqual(rates['ANC-H']['transfer'], 0.0)
            self.assertAlmostEqual(rates['ANC-M']['install'], 224.06)
            self.assertAlmostEqual(rates['ARM-DW']['transfer'], 183.98)
        finally:
            os.unlink(tmp_path)

    def test_missing_file_returns_empty(self):
        """Test that a missing file returns empty dict."""
        rates = generate_weekly_pdfs.load_new_contract_rates('/nonexistent/new_rates.csv')
        self.assertEqual(rates, {})

    def test_skips_short_rows(self):
        """Test that rows with fewer than 9 columns are skipped."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['', '', '', '', '', '', 'Header', '', ''])
            writer.writerow(['', '', '', '', '', '', '', '', ''])
            writer.writerow(['', '', '', '', '', '', '', '', ''])
            writer.writerow(['SHORT', 'Only 5 cols', 'EA', 'Cat', 'Region'])  # Too short
            writer.writerow(['VALID', 'Full row', 'EA', 'Cat', 'Region', '01-18-26', '100', '50', '25'])
            tmp_path = f.name

        try:
            rates = generate_weekly_pdfs.load_new_contract_rates(tmp_path)
            self.assertEqual(len(rates), 1)
            self.assertIn('VALID', rates)
        finally:
            os.unlink(tmp_path)

    def test_group_code_uppercased(self):
        """Test that group codes are normalized to uppercase."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='') as f:
            writer = csv.writer(f)
            for _ in range(3):
                writer.writerow([''] * 9)
            writer.writerow(['anc-h', 'Anchor', 'EA', 'OH', 'AEP', '01-18-26', '100', '50', '25'])
            tmp_path = f.name

        try:
            rates = generate_weekly_pdfs.load_new_contract_rates(tmp_path)
            self.assertIn('ANC-H', rates)
            self.assertNotIn('anc-h', rates)
        finally:
            os.unlink(tmp_path)


class TestBuildCuToGroupMapping(unittest.TestCase):
    """Tests for building the CU-to-group code mapping."""

    def test_builds_mapping_from_old_csv(self):
        """Test building CU -> Compatible Unit Group mapping."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='') as f:
            writer = csv.writer(f)
            writer.writerow([
                'CU WBS #', 'CU', 'Unit Of Measure', 'Description',
                'Compatible Unit Group', 'Install Hours', 'Removal Hours',
                'Transfer Hours', 'Install Price', 'Removal Price', 'Transfer Price'
            ])
            writer.writerow(['100', 'ANC-DHM-10-84-D1', 'EA', 'Anchor', 'ANC-M', '0.24', '0.14', '0', '217.53', '28.60', '0'])
            writer.writerow(['101', 'ANC-DSC-16-96-D1', 'EA', 'Anchor Disc', 'ANC-H', '0.90', '0.14', '0', '790.56', '28.59', '0'])
            writer.writerow(['102', 'ARM-10D-60HS', 'EA', 'Crossarm', 'ARM-DW', '0.66', '0.38', '0.93', '320.62', '73.73', '178.62'])
            tmp_path = f.name

        try:
            mapping = generate_weekly_pdfs.build_cu_to_group_mapping(tmp_path)
            self.assertEqual(len(mapping), 3)
            self.assertEqual(mapping['ANC-DHM-10-84-D1'], 'ANC-M')
            self.assertEqual(mapping['ANC-DSC-16-96-D1'], 'ANC-H')
            self.assertEqual(mapping['ARM-10D-60HS'], 'ARM-DW')
        finally:
            os.unlink(tmp_path)

    def test_missing_file_returns_empty(self):
        """Test that missing file returns empty mapping."""
        mapping = generate_weekly_pdfs.build_cu_to_group_mapping('/nonexistent/old.csv')
        self.assertEqual(mapping, {})

    def test_missing_file_is_benign_not_error(self):
        """Missing old-rates CSV must NOT emit ERROR-level log (benign INFO skip, not a Sentry event)."""
        with self.assertNoLogs(level="ERROR"):
            mapping = generate_weekly_pdfs.build_cu_to_group_mapping("/nonexistent/old.csv")
        self.assertEqual(mapping, {})

    def test_cu_codes_uppercased(self):
        """Test that CU codes and group codes are uppercased."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='') as f:
            writer = csv.writer(f)
            writer.writerow([
                'CU WBS #', 'CU', 'Unit Of Measure', 'Description',
                'Compatible Unit Group', 'Install Hours', 'Removal Hours',
                'Transfer Hours', 'Install Price', 'Removal Price', 'Transfer Price'
            ])
            writer.writerow(['100', 'lower-cu', 'EA', 'Test', 'lower-group', '1', '0', '0', '100', '50', '25'])
            tmp_path = f.name

        try:
            mapping = generate_weekly_pdfs.build_cu_to_group_mapping(tmp_path)
            self.assertIn('LOWER-CU', mapping)
            self.assertEqual(mapping['LOWER-CU'], 'LOWER-GROUP')
        finally:
            os.unlink(tmp_path)


class TestRecalculateRowPrice(unittest.TestCase):
    """Tests for the date-based rate recalculation function."""

    def setUp(self):
        self.cu_to_group = {
            'ANC-DHM-10-84-D1': 'ANC-M',
            'ANC-DSC-16-96-D1': 'ANC-H',
            'ARM-10D-60HS': 'ARM-DW',
        }
        self.rates_primary = {
            'ANC-M': {'install': 224.06, 'removal': 29.46, 'transfer': 0.0},
            'ANC-H': {'install': 814.28, 'removal': 29.45, 'transfer': 0.0},
            'ARM-DW': {'install': 330.24, 'removal': 75.94, 'transfer': 183.98},
        }

    def test_basic_install_recalculation(self):
        """Test basic install price recalculation via CU-to-group mapping."""
        row = {'CU': 'ANC-DHM-10-84-D1', 'Work Type': 'Install', 'Quantity': '3', 'Units Total Price': '$650.00'}
        result = generate_weekly_pdfs.recalculate_row_price(row, self.cu_to_group, self.rates_primary)
        expected = round(224.06 * 3, 2)  # 672.18
        self.assertAlmostEqual(result, expected)
        self.assertAlmostEqual(row['Units Total Price'], expected)

    def test_removal_work_type(self):
        """Test removal work type mapping."""
        row = {'CU': 'ARM-10D-60HS', 'Work Type': 'Removal', 'Quantity': '2', 'Units Total Price': '$100.00'}
        result = generate_weekly_pdfs.recalculate_row_price(row, self.cu_to_group, self.rates_primary)
        expected = round(75.94 * 2, 2)  # 151.88
        self.assertAlmostEqual(result, expected)

    def test_transfer_work_type(self):
        """Test transfer work type mapping."""
        row = {'CU': 'ARM-10D-60HS', 'Work Type': 'Transfer', 'Quantity': '1', 'Units Total Price': '$150.00'}
        result = generate_weekly_pdfs.recalculate_row_price(row, self.cu_to_group, self.rates_primary)
        self.assertAlmostEqual(result, 183.98)

    def test_xfr_work_type(self):
        """Test that 'xfr' maps to transfer rates."""
        row = {'CU': 'ARM-10D-60HS', 'Work Type': 'XFR', 'Quantity': '1', 'Units Total Price': '$150.00'}
        result = generate_weekly_pdfs.recalculate_row_price(row, self.cu_to_group, self.rates_primary)
        self.assertAlmostEqual(result, 183.98)

    def test_unknown_cu_keeps_smartsheet_price(self):
        """Test that unknown CU codes keep the original SmartSheet price."""
        row = {'CU': 'UNKNOWN-CU-999', 'Work Type': 'Install', 'Quantity': '1', 'Units Total Price': '$55.00'}
        result = generate_weekly_pdfs.recalculate_row_price(row, self.cu_to_group, self.rates_primary)
        self.assertAlmostEqual(result, 55.0)
        # Original string should be unchanged
        self.assertEqual(row['Units Total Price'], '$55.00')

    def test_direct_group_code_lookup(self):
        """Test that if SmartSheet row uses a group code directly, it still works."""
        row = {'CU': 'ANC-M', 'Work Type': 'Install', 'Quantity': '2', 'Units Total Price': '$400.00'}
        result = generate_weekly_pdfs.recalculate_row_price(row, self.cu_to_group, self.rates_primary)
        expected = round(224.06 * 2, 2)  # 448.12
        self.assertAlmostEqual(result, expected)

    def test_cu_helper_preferred(self):
        """Test that CU Helper field is preferred over CU field."""
        row = {'CU Helper': 'ANC-DSC-16-96-D1', 'CU': 'ARM-10D-60HS', 'Work Type': 'Install', 'Quantity': '1', 'Units Total Price': '$300.00'}
        result = generate_weekly_pdfs.recalculate_row_price(row, self.cu_to_group, self.rates_primary)
        self.assertAlmostEqual(result, 814.28)  # ANC-H install rate

    def test_arrowhead_discount_rates(self):
        """Test that Arrowhead (subcontractor) rates are 90% of primary."""
        arrowhead_rates = {
            group: {
                'install': round(r['install'] * 0.90, 2),
                'removal': round(r['removal'] * 0.90, 2),
                'transfer': round(r['transfer'] * 0.90, 2),
            }
            for group, r in self.rates_primary.items()
        }
        row = {'CU': 'ANC-DHM-10-84-D1', 'Work Type': 'Install', 'Quantity': '1', 'Units Total Price': '$200.00'}
        result = generate_weekly_pdfs.recalculate_row_price(row, self.cu_to_group, arrowhead_rates)
        expected = round(224.06 * 0.90, 2)  # 201.65
        self.assertAlmostEqual(result, expected)

    def test_zero_quantity_keeps_smartsheet_price(self):
        """Test that zero quantity keeps original SmartSheet price instead of zeroing it out."""
        row = {'CU': 'ANC-DHM-10-84-D1', 'Work Type': 'Install', 'Quantity': '0', 'Units Total Price': '$55.00'}
        result = generate_weekly_pdfs.recalculate_row_price(row, self.cu_to_group, self.rates_primary)
        self.assertAlmostEqual(result, 55.0)
        # Original string should be unchanged
        self.assertEqual(row['Units Total Price'], '$55.00')

    def test_missing_quantity_keeps_smartsheet_price(self):
        """Test that missing/empty quantity keeps original SmartSheet price."""
        row = {'CU': 'ANC-DHM-10-84-D1', 'Work Type': 'Install', 'Units Total Price': '$100.00'}
        result = generate_weekly_pdfs.recalculate_row_price(row, self.cu_to_group, self.rates_primary)
        self.assertAlmostEqual(result, 100.0)
        self.assertEqual(row['Units Total Price'], '$100.00')

    def test_zero_rate_keeps_smartsheet_price(self):
        """Test that a zero rate for a work type keeps the original SmartSheet price."""
        # ANC-M has transfer rate of 0.0 in the test data
        row = {'CU': 'ANC-DHM-10-84-D1', 'Work Type': 'Transfer', 'Quantity': '2', 'Units Total Price': '$75.00'}
        result = generate_weekly_pdfs.recalculate_row_price(row, self.cu_to_group, self.rates_primary)
        self.assertAlmostEqual(result, 75.0)
        self.assertEqual(row['Units Total Price'], '$75.00')

    def test_cu_direct_fallback_when_mapped_group_absent_from_new_rates(self):
        """Regression for VAC crew pricing lag: when the old CSV maps a CU to
        a verbose group name that is NOT a key in the new rates table, the
        recalc should fall back to looking up the CU code directly before
        giving up. Prevents silent old-price retention on specialized work
        items (e.g. vacuum switches) whose CU codes are themselves the key
        in the new contract rates.
        """
        cu_to_group = {'CPD-VS-15-20': 'VACUUM SWITCH'}  # verbose group name
        rates = {
            'CPD-VS-15-20': {'install': 500.00, 'removal': 100.00, 'transfer': 0.0},
        }
        row = {
            'CU': 'CPD-VS-15-20',
            'Work Type': 'Install',
            'Quantity': '2',
            'Units Total Price': '$250.00',
        }
        result = generate_weekly_pdfs.recalculate_row_price(row, cu_to_group, rates)
        self.assertAlmostEqual(result, 1000.00)
        self.assertAlmostEqual(row['Units Total Price'], 1000.00)

    def test_retains_smartsheet_price_when_neither_group_nor_cu_in_new_rates(self):
        """When the mapped group is absent AND the CU is also absent from the
        new rates, the row must retain its SmartSheet price unchanged rather
        than inventing a rate. This guards against the CU-direct fallback
        being too aggressive."""
        cu_to_group = {'CPD-VS-15-20': 'VACUUM SWITCH'}
        rates = {'ANC-M': {'install': 224.06, 'removal': 29.46, 'transfer': 0.0}}
        row = {
            'CU': 'CPD-VS-15-20',
            'Work Type': 'Install',
            'Quantity': '2',
            'Units Total Price': '$250.00',
        }
        result = generate_weekly_pdfs.recalculate_row_price(row, cu_to_group, rates)
        self.assertAlmostEqual(result, 250.00)
        self.assertEqual(row['Units Total Price'], '$250.00')

    def test_out_status_recalculated_on_successful_lookup(self):
        """recalculate_row_price writes outcome='recalculated' when a rate
        was successfully applied (even if the new price equals the existing
        SmartSheet price)."""
        row = {'CU': 'ANC-DHM-10-84-D1', 'Work Type': 'Install', 'Quantity': '3', 'Units Total Price': '$672.18'}
        status = {}
        generate_weekly_pdfs.recalculate_row_price(row, self.cu_to_group, self.rates_primary, out_status=status)
        self.assertEqual(status.get('outcome'), 'recalculated')

    def test_out_status_missing_rate_when_cu_unmapped_and_absent(self):
        """When the CU isn't in cu_to_group and also isn't a direct key in
        the rates dict, out_status['outcome'] must be 'missing_rate' — this
        is the only outcome the per-sheet 'skipped' summary should count."""
        row = {'CU': 'UNKNOWN-999', 'Work Type': 'Install', 'Quantity': '2', 'Units Total Price': '$100.00'}
        status = {}
        generate_weekly_pdfs.recalculate_row_price(row, self.cu_to_group, self.rates_primary, out_status=status)
        self.assertEqual(status.get('outcome'), 'missing_rate')

    def test_out_status_missing_rate_when_group_absent_and_no_cu_fallback(self):
        """When CU maps to a verbose group name that isn't in the new rates
        table AND the CU itself also isn't a direct key, out_status reports
        'missing_rate'."""
        cu_to_group = {'CPD-VS-15-20': 'VACUUM SWITCH'}
        rates = {'ANC-M': {'install': 224.06, 'removal': 29.46, 'transfer': 0.0}}
        row = {'CU': 'CPD-VS-15-20', 'Work Type': 'Install', 'Quantity': '2', 'Units Total Price': '$250.00'}
        status = {}
        generate_weekly_pdfs.recalculate_row_price(row, cu_to_group, rates, out_status=status)
        self.assertEqual(status.get('outcome'), 'missing_rate')

    def test_out_status_invalid_quantity(self):
        """Zero/missing quantity short-circuits with outcome='invalid_quantity',
        not 'missing_rate' — the per-sheet 'skipped' summary must not
        attribute this to CSV coverage gaps."""
        row = {'CU': 'ANC-DHM-10-84-D1', 'Work Type': 'Install', 'Quantity': '0', 'Units Total Price': '$55.00'}
        status = {}
        generate_weekly_pdfs.recalculate_row_price(row, self.cu_to_group, self.rates_primary, out_status=status)
        self.assertEqual(status.get('outcome'), 'invalid_quantity')

    def test_out_status_zero_rate(self):
        """Zero rate for the resolved work type yields outcome='zero_rate'."""
        # ANC-M has transfer rate = 0.0 in the fixture
        row = {'CU': 'ANC-DHM-10-84-D1', 'Work Type': 'Transfer', 'Quantity': '2', 'Units Total Price': '$75.00'}
        status = {}
        generate_weekly_pdfs.recalculate_row_price(row, self.cu_to_group, self.rates_primary, out_status=status)
        self.assertEqual(status.get('outcome'), 'zero_rate')

    def test_out_status_optional_preserves_backward_compat(self):
        """Callers that omit out_status must continue to get a float price."""
        row = {'CU': 'ANC-DHM-10-84-D1', 'Work Type': 'Install', 'Quantity': '3', 'Units Total Price': '$650.00'}
        result = generate_weekly_pdfs.recalculate_row_price(row, self.cu_to_group, self.rates_primary)
        self.assertIsInstance(result, float)
        self.assertAlmostEqual(result, round(224.06 * 3, 2))


class TestResolveCuCode(unittest.TestCase):
    """Tests for the _resolve_cu_code helper used by both recalc and the
    per-sheet skipped-CU summary counter, so they agree on which CU a row
    is attributed to."""

    def test_prefers_cu_helper_over_cu(self):
        row = {'CU Helper': 'ANC-DSC-16-96-D1', 'CU': 'ARM-10D-60HS', 'Billable Unit Code': 'BLT-12'}
        self.assertEqual(generate_weekly_pdfs._resolve_cu_code(row), 'ANC-DSC-16-96-D1')

    def test_nan_helper_falls_back_to_cu(self):
        row = {'CU Helper': 'nan', 'CU': 'ARM-10D-60HS'}
        self.assertEqual(generate_weekly_pdfs._resolve_cu_code(row), 'ARM-10D-60HS')

    def test_falls_back_to_billable_unit_code(self):
        row = {'Billable Unit Code': 'Something-Mixed'}
        self.assertEqual(generate_weekly_pdfs._resolve_cu_code(row), 'SOMETHING-MIXED')

    def test_returns_empty_when_all_blank(self):
        self.assertEqual(generate_weekly_pdfs._resolve_cu_code({}), '')
        self.assertEqual(generate_weekly_pdfs._resolve_cu_code({'CU': None, 'Billable Unit Code': ''}), '')


class TestRateCutoffConfig(unittest.TestCase):
    """Tests for rate cutoff configuration."""

    def test_rate_cutoff_attribute_exists(self):
        """Test that RATE_CUTOFF_DATE attribute exists on the module."""
        self.assertTrue(hasattr(generate_weekly_pdfs, 'RATE_CUTOFF_DATE'))

    def test_arrowhead_discount_value(self):
        """Test that ARROWHEAD_DISCOUNT is 0.90 (10% reduction)."""
        self.assertAlmostEqual(generate_weekly_pdfs.ARROWHEAD_DISCOUNT, 0.90)

    def test_new_rates_csv_attribute(self):
        """Test that NEW_RATES_CSV attribute exists and is a non-empty path."""
        self.assertTrue(hasattr(generate_weekly_pdfs, 'NEW_RATES_CSV'))
        self.assertTrue(len(generate_weekly_pdfs.NEW_RATES_CSV) > 0)

    def test_rates_fingerprint_attribute_exists(self):
        """Test that _RATES_FINGERPRINT attribute exists on the module."""
        self.assertTrue(hasattr(generate_weekly_pdfs, '_RATES_FINGERPRINT'))


class TestRatesFingerprint(unittest.TestCase):
    """Tests for rate table fingerprint computation."""

    def test_fingerprint_deterministic(self):
        """Test that same rates produce same fingerprint."""
        rates = {'ANC-H': {'install': 100.0, 'removal': 50.0, 'transfer': 25.0}}
        fp1 = generate_weekly_pdfs._compute_rates_fingerprint(rates)
        fp2 = generate_weekly_pdfs._compute_rates_fingerprint(rates)
        self.assertEqual(fp1, fp2)

    def test_fingerprint_changes_with_rates(self):
        """Test that different rates produce different fingerprints."""
        rates1 = {'ANC-H': {'install': 100.0, 'removal': 50.0, 'transfer': 25.0}}
        rates2 = {'ANC-H': {'install': 103.0, 'removal': 51.5, 'transfer': 25.75}}
        fp1 = generate_weekly_pdfs._compute_rates_fingerprint(rates1)
        fp2 = generate_weekly_pdfs._compute_rates_fingerprint(rates2)
        self.assertNotEqual(fp1, fp2)

    def test_fingerprint_is_12_chars(self):
        """Test that fingerprint is 12 hex characters."""
        rates = {'X': {'install': 1.0, 'removal': 2.0, 'transfer': 3.0}}
        fp = generate_weekly_pdfs._compute_rates_fingerprint(rates)
        self.assertEqual(len(fp), 12)


class TestCutoffDateRecalculationIntegration(unittest.TestCase):
    """Integration tests for date-based rate recalculation logic."""

    def setUp(self):
        self.cu_to_group = {
            'ANC-DHM-10-84-D1': 'ANC-M',
        }
        self.rates_primary = {
            'ANC-M': {'install': 224.06, 'removal': 29.46, 'transfer': 0.0},
        }

    def test_pre_cutoff_row_keeps_smartsheet_price(self):
        """Verify a row with Snapshot Date before cutoff is not recalculated."""
        import datetime as dt
        cutoff = dt.date(2026, 4, 19)
        row = {
            'CU': 'ANC-DHM-10-84-D1', 'Work Type': 'Install',
            'Quantity': '1', 'Units Total Price': '$200.00',
            'Snapshot Date': '2026-04-18',
        }
        snap = generate_weekly_pdfs.excel_serial_to_date(row['Snapshot Date'])
        snap_date = snap.date() if hasattr(snap, 'date') else snap
        # Pre-cutoff: should NOT recalculate
        self.assertLess(snap_date, cutoff)
        # Price unchanged
        self.assertEqual(row['Units Total Price'], '$200.00')

    def test_post_cutoff_row_gets_recalculated(self):
        """Verify a row with Snapshot Date on/after cutoff gets new rates."""
        import datetime as dt
        cutoff = dt.date(2026, 4, 19)
        row = {
            'CU': 'ANC-DHM-10-84-D1', 'Work Type': 'Install',
            'Quantity': '2', 'Units Total Price': '$400.00',
            'Snapshot Date': '2026-04-19',
        }
        snap = generate_weekly_pdfs.excel_serial_to_date(row['Snapshot Date'])
        snap_date = snap.date() if hasattr(snap, 'date') else snap
        self.assertGreaterEqual(snap_date, cutoff)
        # Recalculate
        new_price = generate_weekly_pdfs.recalculate_row_price(
            row, self.cu_to_group, self.rates_primary)
        self.assertAlmostEqual(new_price, 448.12)  # 224.06 * 2

    def test_discounted_rate_table_math(self):
        """Verify recalculate_row_price correctly applies a 90% discounted rate table (for future Arrowhead use)."""
        arrowhead_rates = {
            'ANC-M': {
                'install': round(224.06 * 0.90, 2),
                'removal': round(29.46 * 0.90, 2),
                'transfer': 0.0,
            }
        }
        row = {
            'CU': 'ANC-DHM-10-84-D1', 'Work Type': 'Install',
            'Quantity': '1', 'Units Total Price': '$200.00',
        }
        new_price = generate_weekly_pdfs.recalculate_row_price(
            row, self.cu_to_group, arrowhead_rates)
        expected = round(224.06 * 0.90, 2)  # 201.65
        self.assertAlmostEqual(new_price, expected)

    def test_snapshot_date_parsing_iso_format(self):
        """Test that ISO format snapshot dates are parsed correctly."""
        dt = generate_weekly_pdfs.excel_serial_to_date('2026-04-19')
        self.assertIsNotNone(dt)

    def test_snapshot_date_parsing_returns_none_for_empty(self):
        """Test that empty/None snapshot dates return None."""
        self.assertIsNone(generate_weekly_pdfs.excel_serial_to_date(None))
        self.assertIsNone(generate_weekly_pdfs.excel_serial_to_date(''))


class TestWeeklyRefDateFallbackCutoff(unittest.TestCase):
    """Regression tests for the Weekly-Ref-Date rate-recalc fallback.

    Production incident context: VAC crew Excel files were being
    generated for week ending 04/12/26 but NOT for 04/19/26. Root
    cause was that the pre-acceptance rate recalc only fired when
    ``Snapshot Date >= RATE_CUTOFF_DATE``. For current-week rows the
    Smartsheet snapshot automation had not yet populated Snapshot
    Date, so recalc was silently skipped, ``Units Total Price`` stayed
    at 0 for VAC crew specialty CUs, ``has_price`` evaluated False,
    and the row was dropped before VAC crew detection could even run.

    The fix is a narrowly-scoped fallback inside
    ``_resolve_rate_recalc_cutoff_date``: when Snapshot Date is blank
    AND Weekly Reference Logged Date parses to a date >= cutoff, use
    the weekly date as the effective gate. Rows that DO have a
    Snapshot Date are unaffected — the snapshot-keyed business rule
    remains primary.
    """

    def setUp(self):
        import datetime as dt
        self.cutoff = dt.date(2026, 4, 19)

    def test_env_constant_exists_and_is_bool(self):
        """``RATE_RECALC_WEEKLY_FALLBACK`` is wired into the module."""
        self.assertTrue(hasattr(generate_weekly_pdfs, 'RATE_RECALC_WEEKLY_FALLBACK'))
        self.assertIsInstance(generate_weekly_pdfs.RATE_RECALC_WEEKLY_FALLBACK, bool)

    def test_snapshot_post_cutoff_returns_snapshot_no_fallback(self):
        """Row with Snapshot Date >= cutoff: primary rule wins, no fallback."""
        import datetime as dt
        row = {
            'Snapshot Date': '2026-04-22',
            'Weekly Reference Logged Date': '2026-04-19',
        }
        effective, used_fallback = generate_weekly_pdfs._resolve_rate_recalc_cutoff_date(
            row, self.cutoff,
        )
        self.assertEqual(effective, dt.date(2026, 4, 22))
        self.assertFalse(used_fallback)

    def test_snapshot_pre_cutoff_returns_none_even_if_weekly_post_cutoff(self):
        """Genuinely pre-cutoff row: fallback does NOT override snapshot rule.

        The snapshot-keyed business rule is authoritative when Snapshot
        Date IS populated — even if the weekly date would say
        otherwise. This preserves the ledger guardrail: "Do NOT change
        the cutoff column from Snapshot Date to Weekly Reference
        Logged Date."
        """
        row = {
            'Snapshot Date': '2026-04-10',
            'Weekly Reference Logged Date': '2026-04-19',
        }
        effective, used_fallback = generate_weekly_pdfs._resolve_rate_recalc_cutoff_date(
            row, self.cutoff,
        )
        self.assertIsNone(effective)
        self.assertFalse(used_fallback)

    def test_blank_snapshot_post_cutoff_weekly_triggers_fallback(self):
        """The incident case: blank Snapshot Date, current-week weekly date.

        This is what caused WE 04/19 VAC crew rows to silently drop.
        With the fallback enabled, recalc now runs using the weekly
        date as the effective gate.
        """
        import datetime as dt
        row = {
            'Snapshot Date': None,
            'Weekly Reference Logged Date': '2026-04-19',
        }
        effective, used_fallback = generate_weekly_pdfs._resolve_rate_recalc_cutoff_date(
            row, self.cutoff,
        )
        self.assertEqual(effective, dt.date(2026, 4, 19))
        self.assertTrue(used_fallback)

    def test_blank_snapshot_blank_weekly_returns_none(self):
        """No usable date on either column → no recalc (unchanged)."""
        row = {'Snapshot Date': None, 'Weekly Reference Logged Date': ''}
        effective, used_fallback = generate_weekly_pdfs._resolve_rate_recalc_cutoff_date(
            row, self.cutoff,
        )
        self.assertIsNone(effective)
        self.assertFalse(used_fallback)

    def test_blank_snapshot_pre_cutoff_weekly_returns_none(self):
        """Historical row with blank Snapshot and pre-cutoff Weekly: no recalc.

        Ensures the fallback is not a universal override — it still
        requires the weekly date to be >= cutoff, preserving contract
        versioning semantics.
        """
        row = {
            'Snapshot Date': None,
            'Weekly Reference Logged Date': '2026-04-12',
        }
        effective, used_fallback = generate_weekly_pdfs._resolve_rate_recalc_cutoff_date(
            row, self.cutoff,
        )
        self.assertIsNone(effective)
        self.assertFalse(used_fallback)

    def test_fallback_disabled_preserves_legacy_behaviour(self):
        """With the fallback disabled, blank Snapshot Date → skip recalc.

        This matches the pre-fix behaviour and proves the env gate is
        respected.
        """
        row = {
            'Snapshot Date': None,
            'Weekly Reference Logged Date': '2026-04-19',
        }
        effective, used_fallback = generate_weekly_pdfs._resolve_rate_recalc_cutoff_date(
            row, self.cutoff, weekly_fallback_enabled=False,
        )
        self.assertIsNone(effective)
        self.assertFalse(used_fallback)

    def test_unparseable_snapshot_falls_through_to_fallback(self):
        """A garbage Snapshot Date behaves like a blank one for fallback purposes.

        ``excel_serial_to_date`` returns ``None`` on unparseable input,
        which the helper treats the same as a blank value. Without
        this, a corrupted Snapshot Date cell would silently suppress
        recalc even when Weekly Reference Logged Date is valid.
        """
        import datetime as dt
        row = {
            'Snapshot Date': 'not-a-real-date',
            'Weekly Reference Logged Date': '2026-04-19',
        }
        effective, used_fallback = generate_weekly_pdfs._resolve_rate_recalc_cutoff_date(
            row, self.cutoff,
        )
        self.assertEqual(effective, dt.date(2026, 4, 19))
        self.assertTrue(used_fallback)

    def test_none_cutoff_always_returns_none(self):
        """No configured cutoff → helper must never authorise recalc.

        Matches the outer production guard ``if RATE_CUTOFF_DATE and
        _rate_new_primary and not is_subcontractor_sheet:`` but is
        also checked inside the helper as a defensive measure so
        callers/tests cannot accidentally enable recalc on a
        cutoff-disabled deployment.
        """
        row = {
            'Snapshot Date': None,
            'Weekly Reference Logged Date': '2026-04-19',
        }
        effective, used_fallback = generate_weekly_pdfs._resolve_rate_recalc_cutoff_date(
            row, None,
        )
        self.assertIsNone(effective)
        self.assertFalse(used_fallback)

    def test_fallback_end_to_end_produces_recalculated_price(self):
        """End-to-end: fallback → recalculate_row_price → new price applied.

        Proves the combined effect is what operators need: a VAC crew
        row with blank Snapshot Date, blank SmartSheet price, CU
        present in the new rates table, and a current-week Weekly
        Reference Logged Date comes out with a recalculated non-zero
        price — so the downstream ``has_price`` gate accepts it
        instead of dropping it silently.
        """
        cu_to_group = {'ANC-DHM-10-84-D1': 'ANC-M'}
        rates = {'ANC-M': {'install': 224.06, 'removal': 29.46, 'transfer': 0.0}}
        row = {
            'Snapshot Date': None,
            'Weekly Reference Logged Date': '2026-04-19',
            'CU': 'ANC-DHM-10-84-D1',
            'Work Type': 'Install',
            'Quantity': '2',
            'Units Total Price': 0,
        }
        effective, used_fallback = generate_weekly_pdfs._resolve_rate_recalc_cutoff_date(
            row, self.cutoff,
        )
        self.assertIsNotNone(effective)
        self.assertTrue(used_fallback)
        new_price = generate_weekly_pdfs.recalculate_row_price(row, cu_to_group, rates)
        self.assertAlmostEqual(new_price, 448.12)
        # Confirm row was updated in-place so the downstream has_price
        # check will pass for this row.
        self.assertEqual(row['Units Total Price'], 448.12)


class TestOriginalContractFolderSkipsRateRecalc(unittest.TestCase):
    """Regression tests for the Smartsheet-native pricing guard.

    Production context: Smartsheet now emits the correct post-cutoff
    ``Units Total Price`` natively for sheets discovered via the two
    folders in ``ORIGINAL_CONTRACT_FOLDER_IDS``. Running Python-side
    rate recalc on top of Smartsheet's authoritative price risked
    overwriting it with a CSV-derived ``rate × qty`` value that did
    not always agree — producing over/under-billed rows. The guard
    introduced in this PR short-circuits the recalc gate for sheets
    whose IDs are in ``_FOLDER_DISCOVERED_ORIG_IDS`` (populated by
    ``discover_folder_sheets`` at every run start), behind a
    default-ON env var so the behaviour is reversible by operators.
    """

    def setUp(self):
        # Snapshot module state so individual tests can mutate
        # _FOLDER_DISCOVERED_ORIG_IDS / SUBCONTRACTOR_SHEET_IDS /
        # RATE_CUTOFF_DATE without leaking into other suites.
        self._orig_folder_ids = set(generate_weekly_pdfs._FOLDER_DISCOVERED_ORIG_IDS)
        self._orig_sub_ids = set(generate_weekly_pdfs.SUBCONTRACTOR_SHEET_IDS)
        self._orig_cutoff = generate_weekly_pdfs.RATE_CUTOFF_DATE
        self._orig_skip_flag = generate_weekly_pdfs.RATE_RECALC_SKIP_ORIGINAL_CONTRACT

    def tearDown(self):
        generate_weekly_pdfs._FOLDER_DISCOVERED_ORIG_IDS.clear()
        generate_weekly_pdfs._FOLDER_DISCOVERED_ORIG_IDS.update(self._orig_folder_ids)
        generate_weekly_pdfs.SUBCONTRACTOR_SHEET_IDS.clear()
        generate_weekly_pdfs.SUBCONTRACTOR_SHEET_IDS.update(self._orig_sub_ids)
        generate_weekly_pdfs.RATE_CUTOFF_DATE = self._orig_cutoff
        generate_weekly_pdfs.RATE_RECALC_SKIP_ORIGINAL_CONTRACT = self._orig_skip_flag

    def _evaluate_gate(self, sheet_id):
        """Mirror the original-contract-skip portion of the production gate.

        The full row-level recalc gate in ``_fetch_and_process_sheet``
        also requires ``_rate_new_primary`` to be populated (the new
        rates dict, loaded by ``load_rate_versions()`` only when
        ``RATE_CUTOFF_DATE`` is set). That branch is exercised by the
        existing recalc-integration tests in
        ``TestCutoffDateRecalculationIntegration`` and
        ``TestWeeklyRefDateFallbackCutoff``. This helper deliberately
        narrows the surface to the **original-contract skip
        composite** so the truth-table tests below stay fast and don't
        require seeding a CSV-loaded rates dict for what is purely a
        boolean-gating concern.

        Keeping the boolean inline (vs. importing a production helper)
        is intentional — if the production ``_skip_recalc_original_contract``
        expression drifts, these tests must be updated in the same PR
        so the invariant stays locked.
        """
        is_subcontractor_sheet = sheet_id in generate_weekly_pdfs.SUBCONTRACTOR_SHEET_IDS
        is_original_contract_sheet = (
            sheet_id in generate_weekly_pdfs._FOLDER_DISCOVERED_ORIG_IDS
        )
        _skip_recalc_original_contract = (
            generate_weekly_pdfs.RATE_CUTOFF_DATE is not None
            and generate_weekly_pdfs.RATE_RECALC_SKIP_ORIGINAL_CONTRACT
            and is_original_contract_sheet
            and not is_subcontractor_sheet
        )
        recalc_would_run = (
            generate_weekly_pdfs.RATE_CUTOFF_DATE is not None
            and not is_subcontractor_sheet
            and not _skip_recalc_original_contract
        )
        return recalc_would_run, _skip_recalc_original_contract

    def test_env_var_exists_and_is_bool(self):
        """``RATE_RECALC_SKIP_ORIGINAL_CONTRACT`` is wired into the module."""
        self.assertTrue(
            hasattr(generate_weekly_pdfs, 'RATE_RECALC_SKIP_ORIGINAL_CONTRACT')
        )
        self.assertIsInstance(
            generate_weekly_pdfs.RATE_RECALC_SKIP_ORIGINAL_CONTRACT, bool
        )

    def test_default_folder_ids_include_smartsheet_priced_folders(self):
        """Default ORIGINAL_CONTRACT_FOLDER_IDS covers the two Smartsheet-priced folders.

        Incident: user reported Smartsheet natively prices rows in
        folders 7644752003786628 and 8815193070299012 for post-cutoff
        ``Units Completed?`` rows. If these IDs are ever removed from
        the default list without also updating the env var wiring in
        ``.github/workflows/weekly-excel-generation.yml``, the guard
        becomes a no-op on CI runs that rely on the default.
        """
        defaults = generate_weekly_pdfs.ORIGINAL_CONTRACT_FOLDER_IDS
        self.assertIn(7644752003786628, defaults)
        self.assertIn(8815193070299012, defaults)

    def test_guard_fires_for_original_contract_sheet(self):
        """Sheet in ORIG folder + cutoff set + env on → recalc skipped."""
        import datetime as dt
        generate_weekly_pdfs.RATE_CUTOFF_DATE = dt.date(2026, 4, 12)
        generate_weekly_pdfs.RATE_RECALC_SKIP_ORIGINAL_CONTRACT = True
        generate_weekly_pdfs._FOLDER_DISCOVERED_ORIG_IDS.clear()
        generate_weekly_pdfs._FOLDER_DISCOVERED_ORIG_IDS.add(111111)
        generate_weekly_pdfs.SUBCONTRACTOR_SHEET_IDS.clear()

        recalc_would_run, skip_fired = self._evaluate_gate(111111)
        self.assertFalse(recalc_would_run)
        self.assertTrue(skip_fired)

    def test_guard_does_not_fire_for_non_original_contract_sheet(self):
        """Sheet NOT in ORIG folder → recalc still runs (pre-fix behaviour)."""
        import datetime as dt
        generate_weekly_pdfs.RATE_CUTOFF_DATE = dt.date(2026, 4, 12)
        generate_weekly_pdfs.RATE_RECALC_SKIP_ORIGINAL_CONTRACT = True
        generate_weekly_pdfs._FOLDER_DISCOVERED_ORIG_IDS.clear()
        generate_weekly_pdfs._FOLDER_DISCOVERED_ORIG_IDS.add(111111)
        generate_weekly_pdfs.SUBCONTRACTOR_SHEET_IDS.clear()

        recalc_would_run, skip_fired = self._evaluate_gate(222222)
        self.assertTrue(recalc_would_run)
        self.assertFalse(skip_fired)

    def test_env_var_off_restores_legacy_behaviour(self):
        """``RATE_RECALC_SKIP_ORIGINAL_CONTRACT=False`` → recalc runs on ORIG sheet too.

        Proves the env-var kill switch works — operators can flip off
        the guard if Smartsheet-native pricing ever breaks or needs
        to be bypassed.
        """
        import datetime as dt
        generate_weekly_pdfs.RATE_CUTOFF_DATE = dt.date(2026, 4, 12)
        generate_weekly_pdfs.RATE_RECALC_SKIP_ORIGINAL_CONTRACT = False
        generate_weekly_pdfs._FOLDER_DISCOVERED_ORIG_IDS.clear()
        generate_weekly_pdfs._FOLDER_DISCOVERED_ORIG_IDS.add(111111)
        generate_weekly_pdfs.SUBCONTRACTOR_SHEET_IDS.clear()

        recalc_would_run, skip_fired = self._evaluate_gate(111111)
        self.assertTrue(recalc_would_run)
        self.assertFalse(skip_fired)

    def test_no_cutoff_no_recalc_regardless_of_folder(self):
        """Without ``RATE_CUTOFF_DATE``, recalc is disabled globally.

        Confirms the original outer guard is preserved — the new
        folder skip is additive, not a replacement.
        """
        generate_weekly_pdfs.RATE_CUTOFF_DATE = None
        generate_weekly_pdfs.RATE_RECALC_SKIP_ORIGINAL_CONTRACT = True
        generate_weekly_pdfs._FOLDER_DISCOVERED_ORIG_IDS.clear()
        generate_weekly_pdfs._FOLDER_DISCOVERED_ORIG_IDS.add(111111)
        generate_weekly_pdfs.SUBCONTRACTOR_SHEET_IDS.clear()

        recalc_would_run, skip_fired = self._evaluate_gate(111111)
        self.assertFalse(recalc_would_run)
        # skip_fired must be False when cutoff is disabled: the skip
        # flag only matters when recalc was otherwise eligible, and
        # operators should not see the "🛡️ Skipping..." log on a
        # cutoff-disabled deployment.
        self.assertFalse(skip_fired)

    def test_subcontractor_sheet_wins_over_original_contract(self):
        """Sheet in BOTH sub and orig sets: subcontractor exclusion wins.

        Pathological but possible (misconfiguration). The subcontractor
        exclusion at the recalc gate is primary and unconditional, so
        the sheet skips recalc via the subcontractor path and the
        original-contract skip log never fires (avoiding duplicate
        "skipping" messages for the same sheet).
        """
        import datetime as dt
        generate_weekly_pdfs.RATE_CUTOFF_DATE = dt.date(2026, 4, 12)
        generate_weekly_pdfs.RATE_RECALC_SKIP_ORIGINAL_CONTRACT = True
        generate_weekly_pdfs._FOLDER_DISCOVERED_ORIG_IDS.clear()
        generate_weekly_pdfs._FOLDER_DISCOVERED_ORIG_IDS.add(111111)
        generate_weekly_pdfs.SUBCONTRACTOR_SHEET_IDS.clear()
        generate_weekly_pdfs.SUBCONTRACTOR_SHEET_IDS.add(111111)

        recalc_would_run, skip_fired = self._evaluate_gate(111111)
        self.assertFalse(recalc_would_run)
        # Subcontractor-exclusion path short-circuits first, so the
        # original-contract skip must NOT fire for the same sheet.
        self.assertFalse(skip_fired)

    def test_guard_does_not_mutate_recalculate_row_price(self):
        """``recalculate_row_price`` itself is unchanged by the guard.

        The guard is a sheet-level gate; it does NOT modify
        ``recalculate_row_price``'s behaviour. A caller that invokes
        the function directly (e.g., a future one-off reprice script)
        must still get the full recalc behaviour regardless of env
        vars.
        """
        cu_to_group = {'ANC-DHM-10-84-D1': 'ANC-M'}
        rates = {'ANC-M': {'install': 224.06, 'removal': 29.46, 'transfer': 0.0}}
        row = {
            'CU': 'ANC-DHM-10-84-D1',
            'Work Type': 'Install',
            'Quantity': '2',
            'Units Total Price': 0,
        }
        # Flip the env flag on — helper should still recalc, proving
        # the guard is at the caller (sheet-level gate), not here.
        generate_weekly_pdfs.RATE_RECALC_SKIP_ORIGINAL_CONTRACT = True
        new_price = generate_weekly_pdfs.recalculate_row_price(row, cu_to_group, rates)
        self.assertAlmostEqual(new_price, 448.12)
        self.assertEqual(row['Units Total Price'], 448.12)


class TestExpandedHashCoverage(unittest.TestCase):
    """Tests for the expanded calculate_data_hash field coverage."""

    def test_hash_changes_when_customer_name_changes(self):
        """Verify the hash changes when Customer Name is modified."""
        base_row = {
            'Work Request #': '12345', 'Snapshot Date': '2025-01-01',
            'CU': 'ABC', 'Quantity': '1', 'Units Total Price': '100.00',
            'Work Type': 'install', 'Dept #': '10',
            'Customer Name': 'OriginalCustomer', '__variant': 'primary',
        }
        import copy
        modified_row = copy.deepcopy(base_row)
        modified_row['Customer Name'] = 'DifferentCustomer'

        hash1 = generate_weekly_pdfs.calculate_data_hash([base_row])
        hash2 = generate_weekly_pdfs.calculate_data_hash([modified_row])
        self.assertNotEqual(hash1, hash2)

    def test_hash_changes_when_job_number_changes(self):
        """Verify the hash changes when Job # is modified."""
        base_row = {
            'Work Request #': '12345', 'Snapshot Date': '2025-01-01',
            'CU': 'ABC', 'Quantity': '1', 'Units Total Price': '100.00',
            'Work Type': 'install', 'Job #': 'JOB001', '__variant': 'primary',
        }
        import copy
        modified_row = copy.deepcopy(base_row)
        modified_row['Job #'] = 'JOB002'

        hash1 = generate_weekly_pdfs.calculate_data_hash([base_row])
        hash2 = generate_weekly_pdfs.calculate_data_hash([modified_row])
        self.assertNotEqual(hash1, hash2)

    def test_hash_stable_when_no_changes(self):
        """Verify the hash is deterministic for identical input."""
        row = {
            'Work Request #': '12345', 'Snapshot Date': '2025-01-01',
            'CU': 'ABC', 'Quantity': '1', 'Units Total Price': '100.00',
            'Work Type': 'install', '__variant': 'primary',
        }
        hash1 = generate_weekly_pdfs.calculate_data_hash([row])
        hash2 = generate_weekly_pdfs.calculate_data_hash([row])
        self.assertEqual(hash1, hash2)


class TestSubcontractorVariantGrouping(unittest.TestCase):
    """Plan 01-03 Task 1: subcontractor variant tagging in group_source_rows().

    Per the plan's committed plumbing decision (Blocker 3), the gate is
    PER-ROW via ``r.get('__source_sheet_id') in _FOLDER_DISCOVERED_SUB_IDS``.
    Each test snapshots & restores the module's folder-id sets and the
    kill-switch env flag so a test's mutation cannot leak.
    """

    _SUB_SHEET_ID = 8162920222379908
    _ORIG_SHEET_ID = 7644752003786628

    def setUp(self):
        self._orig_sub_ids = set(generate_weekly_pdfs._FOLDER_DISCOVERED_SUB_IDS)
        self._orig_orig_ids = set(generate_weekly_pdfs._FOLDER_DISCOVERED_ORIG_IDS)
        self._orig_kill = generate_weekly_pdfs.SUBCONTRACTOR_RATE_VARIANTS_ENABLED
        # Seed the SUB folder set with our test sheet id so the per-row
        # gate trips.
        generate_weekly_pdfs._FOLDER_DISCOVERED_SUB_IDS.clear()
        generate_weekly_pdfs._FOLDER_DISCOVERED_SUB_IDS.add(self._SUB_SHEET_ID)
        generate_weekly_pdfs._FOLDER_DISCOVERED_ORIG_IDS.clear()
        generate_weekly_pdfs._FOLDER_DISCOVERED_ORIG_IDS.add(self._ORIG_SHEET_ID)
        generate_weekly_pdfs.SUBCONTRACTOR_RATE_VARIANTS_ENABLED = True

    def tearDown(self):
        generate_weekly_pdfs._FOLDER_DISCOVERED_SUB_IDS.clear()
        generate_weekly_pdfs._FOLDER_DISCOVERED_SUB_IDS.update(self._orig_sub_ids)
        generate_weekly_pdfs._FOLDER_DISCOVERED_ORIG_IDS.clear()
        generate_weekly_pdfs._FOLDER_DISCOVERED_ORIG_IDS.update(self._orig_orig_ids)
        generate_weekly_pdfs.SUBCONTRACTOR_RATE_VARIANTS_ENABLED = self._orig_kill

    def _make_row(
        self,
        wr,
        date_str,
        price,
        snapshot=None,
        source_sheet_id=None,
        is_helper=False,
        helper_foreman='',
        helper_dept='',
        helper_job='',
    ):
        """Minimal valid source row for group_source_rows().

        ``source_sheet_id`` populates ``row['__source_sheet_id']`` —
        the field the plan's per-row gate reads. Defaults to the
        test class's seeded subcontractor sheet id.
        """
        if source_sheet_id is None:
            source_sheet_id = self._SUB_SHEET_ID
        row = {
            'Work Request #': wr,
            'Weekly Reference Logged Date': date_str,
            'Units Completed?': True,
            'Units Total Price': price,
            'Snapshot Date': snapshot if snapshot is not None else date_str,
            '__effective_user': 'TestForeman',
            '__assignment_method': 'FOREMAN_COLUMN',
            '__is_helper_row': is_helper,
            '__helper_foreman': helper_foreman,
            '__helper_dept': helper_dept,
            '__helper_job': helper_job,
            '__is_vac_crew': False,
            '__source_sheet_id': source_sheet_id,
        }
        return row

    def test_post_cutoff_subcontractor_row_emits_aep_billable_and_reduced_sub(self):
        """Test 1: post-cutoff snapshot, SUB sheet, kill-switch on → both new variant group keys appear."""
        row = self._make_row(
            wr='WR_X',
            date_str='2026-04-19',
            price='$100.00',
            snapshot='2026-04-19',  # post-cutoff (>= 2026-04-12)
        )
        groups = generate_weekly_pdfs.group_source_rows([row])
        keys = list(groups.keys())
        self.assertTrue(
            any('_AEPBILLABLE' in k and '_HELPER_' not in k for k in keys),
            f"Expected an _AEPBILLABLE group key for post-cutoff sub row; got: {keys}",
        )
        self.assertTrue(
            any('_REDUCEDSUB' in k and '_HELPER_' not in k for k in keys),
            f"Expected a _REDUCEDSUB group key for sub row; got: {keys}",
        )

    def test_pre_cutoff_subcontractor_row_emits_reduced_sub_only(self):
        """Test 2: pre-cutoff snapshot → ReducedSub yes, AEPBillable no (D-08)."""
        row = self._make_row(
            wr='WR_Y',
            date_str='2026-04-05',
            price='$100.00',
            snapshot='2026-04-05',  # pre-cutoff (< 2026-04-12)
        )
        groups = generate_weekly_pdfs.group_source_rows([row])
        keys = list(groups.keys())
        self.assertTrue(
            any('_REDUCEDSUB' in k for k in keys),
            f"Expected a _REDUCEDSUB group key for pre-cutoff sub row; got: {keys}",
        )
        self.assertFalse(
            any('_AEPBILLABLE' in k for k in keys),
            f"Expected NO _AEPBILLABLE group key for pre-cutoff sub row; got: {keys}",
        )

    def test_helper_event_emits_shadow_variants_when_post_cutoff(self):
        """Test 3: helper-foreman event on sub WR post-cutoff → both shadow variants appear."""
        row = self._make_row(
            wr='WR_Z',
            date_str='2026-04-19',
            price='$100.00',
            snapshot='2026-04-19',
            is_helper=True,
            helper_foreman='Jane Smith',
            helper_dept='123',
            helper_job='J-1',
        )
        groups = generate_weekly_pdfs.group_source_rows([row])
        keys = list(groups.keys())
        self.assertTrue(
            any('_AEPBILLABLE_HELPER_Jane_Smith' in k for k in keys),
            f"Expected _AEPBILLABLE_HELPER_Jane_Smith key; got: {keys}",
        )
        self.assertTrue(
            any('_REDUCEDSUB_HELPER_Jane_Smith' in k for k in keys),
            f"Expected _REDUCEDSUB_HELPER_Jane_Smith key; got: {keys}",
        )

    def test_non_subcontractor_sheet_emits_no_new_variants(self):
        """Test 4: row from a non-SUB sheet → no new variant keys (per-row gate proof)."""
        row = self._make_row(
            wr='WR_A',
            date_str='2026-04-19',
            price='$100.00',
            snapshot='2026-04-19',
            source_sheet_id=self._ORIG_SHEET_ID,  # in ORIG, NOT in SUB
        )
        groups = generate_weekly_pdfs.group_source_rows([row])
        keys = list(groups.keys())
        self.assertFalse(
            any('_AEPBILLABLE' in k or '_REDUCEDSUB' in k for k in keys),
            f"Expected NO new variant keys for non-sub row; got: {keys}",
        )

    def test_kill_switch_off_emits_no_new_variants(self):
        """Test 5: SUBCONTRACTOR_RATE_VARIANTS_ENABLED=False → no new variant keys (D-13)."""
        generate_weekly_pdfs.SUBCONTRACTOR_RATE_VARIANTS_ENABLED = False
        row = self._make_row(
            wr='WR_B',
            date_str='2026-04-19',
            price='$100.00',
            snapshot='2026-04-19',
        )
        groups = generate_weekly_pdfs.group_source_rows([row])
        keys = list(groups.keys())
        self.assertFalse(
            any('_AEPBILLABLE' in k or '_REDUCEDSUB' in k for k in keys),
            f"Expected NO new variant keys with kill switch off; got: {keys}",
        )

    def test_variant_string_tagging_uses_canonical_lowercase(self):
        """Test 6: r_copy['__variant'] in the new variants uses the exact lowercase strings."""
        row = self._make_row(
            wr='WR_C',
            date_str='2026-04-19',
            price='$100.00',
            snapshot='2026-04-19',
        )
        groups = generate_weekly_pdfs.group_source_rows([row])
        variants_seen = set()
        for key, rows in groups.items():
            if '_AEPBILLABLE' in key or '_REDUCEDSUB' in key:
                variants_seen.add(rows[0].get('__variant'))
        # At minimum reduced_sub and aep_billable must be present.
        self.assertIn('reduced_sub', variants_seen, f"expected 'reduced_sub' tagging; saw {variants_seen}")
        self.assertIn('aep_billable', variants_seen, f"expected 'aep_billable' tagging; saw {variants_seen}")
        self.assertTrue(
            variants_seen.issubset(
                {'reduced_sub', 'aep_billable', 'reduced_sub_helper', 'aep_billable_helper'}
            ),
            f"new-variant group rows must tag __variant only with the four lowercase strings; got {variants_seen}",
        )

    def test_helper_name_with_apostrophe_sanitized_in_key(self):
        """Test 7: helper name with non-word chars is sanitized via _RE_SANITIZE_HELPER_NAME before key embedding."""
        row = self._make_row(
            wr='WR_D',
            date_str='2026-04-19',
            price='$100.00',
            snapshot='2026-04-19',
            is_helper=True,
            helper_foreman="Jane O'Brien",
            helper_dept='456',
            helper_job='J-2',
        )
        groups = generate_weekly_pdfs.group_source_rows([row])
        keys = list(groups.keys())
        # ``_RE_SANITIZE_HELPER_NAME`` replaces ``'`` with ``_``.
        sanitized_expected = 'Jane_O_Brien'
        self.assertTrue(
            any(f'_REDUCEDSUB_HELPER_{sanitized_expected}' in k for k in keys),
            f"Expected sanitized helper name {sanitized_expected!r} in a REDUCEDSUB_HELPER key; got: {keys}",
        )
        self.assertTrue(
            any(f'_AEPBILLABLE_HELPER_{sanitized_expected}' in k for k in keys),
            f"Expected sanitized helper name {sanitized_expected!r} in an AEPBILLABLE_HELPER key; got: {keys}",
        )

    def test_per_row_gate_does_not_bleed_across_rows_in_same_call(self):
        """Test 8: a single call with a sub row + a non-sub row → only the sub row produces new variants.

        Regression guard against accidental per-CALL gating that would
        emit variant keys for every row in the call once any row was
        on a SUB sheet.
        """
        row_sub = self._make_row(
            wr='WR_E',
            date_str='2026-04-19',
            price='$100.00',
            snapshot='2026-04-19',
            source_sheet_id=self._SUB_SHEET_ID,
        )
        row_orig = self._make_row(
            wr='WR_F',
            date_str='2026-04-19',
            price='$100.00',
            snapshot='2026-04-19',
            source_sheet_id=self._ORIG_SHEET_ID,
        )
        groups = generate_weekly_pdfs.group_source_rows([row_sub, row_orig])
        # WR_E (sub) MUST have new variant keys.
        sub_keys = [k for k in groups if '_WR_E' in k or k.endswith('_WR_E') or 'WR_E' in k]
        # The naming convention: ``{week}_{wr_key}_REDUCEDSUB`` etc.
        self.assertTrue(
            any('WR_E' in k and ('_AEPBILLABLE' in k or '_REDUCEDSUB' in k) for k in groups),
            f"Expected WR_E (sub) to produce new-variant keys; got: {list(groups.keys())}",
        )
        # WR_F (orig) MUST NOT produce any new variant keys.
        wr_f_new_variant_keys = [
            k for k in groups
            if 'WR_F' in k and ('_AEPBILLABLE' in k or '_REDUCEDSUB' in k)
        ]
        self.assertEqual(
            wr_f_new_variant_keys, [],
            f"WR_F (orig) must NOT produce new-variant keys (per-row gate); got: {wr_f_new_variant_keys}",
        )


class TestResolveRowPriceCanonicalColumnNames(unittest.TestCase):
    """Plan 01-03 Task 2 / Blocker 2: _resolve_row_price reads ONLY canonical column keys.

    Per round-3 checker Blocker 2, the helper MUST read the row dict
    using the canonical keys produced by ``_validate_single_sheet``'s
    synonyms layer (`'CU'`, `'Work Type'`, `'Quantity'`,
    `'Units Total Price'`). Reading non-canonical fallback keys would
    be a silent regression on any sheet whose source column titles
    differ — by the time the row reaches ``generate_excel``, only the
    canonical keys exist.
    """

    def setUp(self):
        # Snapshot the rates dict so tests can mutate.
        self._orig_rates = dict(generate_weekly_pdfs._SUBCONTRACTOR_RATES)
        generate_weekly_pdfs._SUBCONTRACTOR_RATES.clear()
        generate_weekly_pdfs._SUBCONTRACTOR_RATES['XYZ'] = {
            'cu_code': 'XYZ',
            'cu_wbs': '999',
            'compatible_unit_group': 'TestGroup',
            'reduced_install_price': 10.0,
            'reduced_remove_price': 5.0,
            'reduced_transfer_price': 2.5,
            'new_install_price': 20.0,
            'new_remove_price': 12.0,
            'new_transfer_price': 6.0,
        }

    def tearDown(self):
        generate_weekly_pdfs._SUBCONTRACTOR_RATES.clear()
        generate_weekly_pdfs._SUBCONTRACTOR_RATES.update(self._orig_rates)

    def test_helper_callable(self):
        """``_resolve_row_price`` exists at module scope (testable surface)."""
        self.assertTrue(hasattr(generate_weekly_pdfs, '_resolve_row_price'))
        self.assertTrue(callable(generate_weekly_pdfs._resolve_row_price))

    def test_canonical_keys_resolve_aep_billable_install(self):
        """Test 9 (Blocker 2 lock-in): canonical-keyed row produces new_install_price × qty."""
        from collections import Counter
        row = {
            'CU': 'XYZ',
            'Work Type': 'Install',
            'Quantity': 5,
            'Units Total Price': '$0.00',
        }
        missing = Counter()
        price = generate_weekly_pdfs._resolve_row_price(row, 'aep_billable', missing)
        self.assertAlmostEqual(price, 100.0)  # new_install_price 20.0 × 5
        self.assertEqual(missing, Counter())  # no missing CUs

    def test_canonical_keys_resolve_reduced_sub_remove(self):
        """Canonical-keyed row + reduced_sub variant + Removal → reduced_remove_price × qty."""
        from collections import Counter
        row = {
            'CU': 'XYZ',
            'Work Type': 'Removal',
            'Quantity': 4,
            'Units Total Price': '$0.00',
        }
        missing = Counter()
        price = generate_weekly_pdfs._resolve_row_price(row, 'reduced_sub', missing)
        self.assertAlmostEqual(price, 20.0)  # reduced_remove_price 5.0 × 4

    def test_canonical_keys_resolve_aep_billable_transfer(self):
        """AEP Billable + Transfer work type → new_transfer_price × qty."""
        from collections import Counter
        row = {
            'CU': 'XYZ',
            'Work Type': 'Transfer',
            'Quantity': 3,
            'Units Total Price': '$0.00',
        }
        missing = Counter()
        price = generate_weekly_pdfs._resolve_row_price(row, 'aep_billable_helper', missing)
        self.assertAlmostEqual(price, 18.0)  # new_transfer_price 6.0 × 3

    def test_missing_cu_falls_through_to_smartsheet(self):
        """Test 5 (D-16): CU absent from _SUBCONTRACTOR_RATES → retain SmartSheet price, no zero-out."""
        from collections import Counter
        row = {
            'CU': 'UNKNOWN_CU',
            'Work Type': 'Install',
            'Quantity': 5,
            'Units Total Price': '$77.50',
        }
        missing = Counter()
        price = generate_weekly_pdfs._resolve_row_price(row, 'aep_billable', missing)
        # Must keep SmartSheet pricing, not zero out, not raise.
        self.assertAlmostEqual(price, 77.5)

    def test_missing_cu_recorded_in_counter(self):
        """Test 6 (D-16): missing CU code accumulates in the per-call Counter."""
        from collections import Counter
        row = {
            'CU': 'ABSENT_CU',
            'Work Type': 'Install',
            'Quantity': 1,
            'Units Total Price': '$10.00',
        }
        missing = Counter()
        generate_weekly_pdfs._resolve_row_price(row, 'aep_billable', missing)
        self.assertIn('ABSENT_CU', missing)
        self.assertEqual(missing['ABSENT_CU'], 1)

    def test_primary_variant_unchanged_regardless_of_cu(self):
        """Test 7: primary/helper/vac_crew variants return SmartSheet price unchanged (D-14/D-15)."""
        from collections import Counter
        row = {
            'CU': 'XYZ',  # present in rates, but variant is primary so ignored
            'Work Type': 'Install',
            'Quantity': 5,
            'Units Total Price': '$42.00',
        }
        for variant in ('primary', 'helper', 'vac_crew'):
            missing = Counter()
            price = generate_weekly_pdfs._resolve_row_price(row, variant, missing)
            self.assertAlmostEqual(
                price, 42.0,
                msg=f"variant {variant!r} must keep SmartSheet pricing unchanged",
            )
            self.assertEqual(
                missing, Counter(),
                f"variant {variant!r} must not record missing CUs (CSV not consulted)",
            )

    def test_test_10_wrong_key_name_units_completed_falls_through(self):
        """Test 10 (Blocker 2 negative): wrong key 'Units Completed' is NOT read as quantity.

        Defensive: a future regression that "fixes" the helper to read
        'Units Completed' (a checkbox column, NOT a qty) must not pick
        up the value — instead it falls through to SmartSheet pricing.
        """
        from collections import Counter
        row = {
            'CU': 'XYZ',
            'Work Type': 'Install',
            'Units Completed': 5,  # WRONG key name (a checkbox column in source data)
            # No 'Quantity' canonical key — helper must treat qty as 0
            # and fall through.
            'Units Total Price': '$33.33',
        }
        missing = Counter()
        price = generate_weekly_pdfs._resolve_row_price(row, 'aep_billable', missing)
        # qty=0 → degenerate path → SmartSheet fallback
        self.assertAlmostEqual(price, 33.33)

    def test_helper_body_does_not_reference_forbidden_keys(self):
        """Negative invariant: executable body does NOT read non-canonical fallback keys.

        Per Blocker 2 acceptance criterion: the helper's EXECUTABLE
        body (i.e. code excluding the docstring) MUST NOT call
        ``row.get(...)`` against 'Billable Unit Code', 'Units Completed',
        'Qty', '# Units', 'Total Price' (bare), or 'Redlined Total
        Price'. Those names are the synonym surface, not the canonical
        surface, and only the canonical keys exist by the time the row
        reaches this helper. The docstring intentionally cites several
        of these names as the documented synonym set, so we strip the
        docstring before scanning.
        """
        import inspect
        func = generate_weekly_pdfs._resolve_row_price
        src = inspect.getsource(func)
        # Strip the function's docstring (the first triple-quoted block
        # after ``def``); the docstring mentions the synonym set on
        # purpose for future-reader context. The executable body must
        # not reference those tokens.
        doc = inspect.getdoc(func) or ''
        body = src
        if doc:
            # Remove every line that matches a docstring line; cheap
            # but precise enough — the docstring is the only place that
            # contains the long-form synonym citations.
            for line in doc.splitlines():
                if line.strip():
                    body = body.replace(line, '')
        # Check for the actual row.get(...) call pattern so a CODE
        # COMMENT mentioning a forbidden token (e.g.,
        # "Quantity ONLY — never 'Units Completed' (checkbox)")
        # doesn't trip the negative invariant.
        forbidden_keys = [
            'Billable Unit Code',
            'Units Completed',
            'Qty',
            '# Units',
            'Total Price',  # bare — canonical is 'Units Total Price'
            'Redlined Total Price',
        ]
        hits = []
        for key in forbidden_keys:
            if f"row.get('{key}')" in body or f'row.get("{key}")' in body:
                hits.append(key)
        self.assertEqual(
            hits, [],
            f"_resolve_row_price reads non-canonical keys: {hits} — Blocker 2 forbids; "
            f"add synonyms in _validate_single_sheet instead.",
        )


class TestSubcontractorVariantFilenameSuffixes(unittest.TestCase):
    """Plan 01-03 Task 2: generate_excel produces the 4 new variant filename suffixes."""

    def _make_group_row(self, variant, wr='99887766', week='2026-04-19',
                        snap='2026-04-19', helper_foreman='', cu='XYZ',
                        work_type='Install', quantity=2, price='$0.00'):
        return {
            'Work Request #': wr,
            'Weekly Reference Logged Date': week,
            'Snapshot Date': snap,
            'Units Completed?': True,
            'Units Total Price': price,
            'CU': cu,
            'Work Type': work_type,
            'Quantity': quantity,
            'Customer Name': 'TestCustomer',
            'Foreman': 'TestForeman',
            'Dept #': '500',
            'Job #': 'J-1',
            '__effective_user': 'TestForeman',
            '__current_foreman': helper_foreman or 'TestForeman',
            '__variant': variant,
            '__helper_foreman': helper_foreman,
            '__helper_dept': '123' if helper_foreman else '',
            '__helper_job': 'J-2' if helper_foreman else '',
            '__week_ending_date': __import__('datetime').datetime(2026, 4, 19),
        }

    def setUp(self):
        # Direct OUTPUT_FOLDER into a temp dir so tests don't pollute the repo.
        self._tmpdir = tempfile.TemporaryDirectory()
        self._orig_output_folder = generate_weekly_pdfs.OUTPUT_FOLDER
        generate_weekly_pdfs.OUTPUT_FOLDER = self._tmpdir.name
        # Seed rates so AEPBillable / ReducedSub price substitution
        # is exercised (the suffix tests are independent of the price
        # value, but instantiating workbook generation needs them).
        self._orig_rates = dict(generate_weekly_pdfs._SUBCONTRACTOR_RATES)
        generate_weekly_pdfs._SUBCONTRACTOR_RATES.clear()
        generate_weekly_pdfs._SUBCONTRACTOR_RATES['XYZ'] = {
            'cu_code': 'XYZ',
            'cu_wbs': '999',
            'compatible_unit_group': 'TestGroup',
            'reduced_install_price': 10.0,
            'reduced_remove_price': 5.0,
            'reduced_transfer_price': 2.5,
            'new_install_price': 20.0,
            'new_remove_price': 12.0,
            'new_transfer_price': 6.0,
        }

    def tearDown(self):
        generate_weekly_pdfs.OUTPUT_FOLDER = self._orig_output_folder
        self._tmpdir.cleanup()
        generate_weekly_pdfs._SUBCONTRACTOR_RATES.clear()
        generate_weekly_pdfs._SUBCONTRACTOR_RATES.update(self._orig_rates)

    def test_aep_billable_filename_contains_AEPBillable_suffix(self):
        """Test 1: variant='aep_billable' → filename contains _AEPBillable_."""
        import datetime as dt
        rows = [self._make_group_row('aep_billable')]
        result = generate_weekly_pdfs.generate_excel(
            '041926_99887766_AEPBILLABLE', rows, dt.datetime(2026, 4, 19),
            data_hash='deadbeefcafebab0',
        )
        # Return is now a 5-tuple
        excel_path, filename, wr_numbers = result[0], result[1], result[2]
        self.assertIn('_AEPBillable_', filename)
        self.assertNotIn('_Helper_', filename)
        self.assertNotIn('_ReducedSub_', filename)
        self.assertTrue(os.path.exists(excel_path), f"workbook not written: {excel_path}")

    def test_reduced_sub_filename_contains_ReducedSub_suffix(self):
        """variant='reduced_sub' → filename contains _ReducedSub_."""
        import datetime as dt
        rows = [self._make_group_row('reduced_sub')]
        result = generate_weekly_pdfs.generate_excel(
            '041926_99887766_REDUCEDSUB', rows, dt.datetime(2026, 4, 19),
            data_hash='deadbeefcafebab1',
        )
        filename = result[1]
        self.assertIn('_ReducedSub_', filename)
        self.assertNotIn('_AEPBillable_', filename)
        self.assertNotIn('_Helper_', filename)

    def test_aep_billable_helper_filename_includes_sanitized_helper_name(self):
        """Test 2: variant='aep_billable_helper' with helper 'Jane Smith' → filename includes _AEPBillable_Helper_Jane_Smith_."""
        import datetime as dt
        rows = [self._make_group_row('aep_billable_helper', helper_foreman='Jane Smith')]
        result = generate_weekly_pdfs.generate_excel(
            '041926_99887766_AEPBILLABLE_HELPER_Jane_Smith', rows,
            dt.datetime(2026, 4, 19), data_hash='deadbeefcafebab2',
        )
        filename = result[1]
        self.assertIn('_AEPBillable_Helper_Jane_Smith_', filename)

    def test_reduced_sub_helper_filename_includes_sanitized_helper_name(self):
        """variant='reduced_sub_helper' with helper 'Jane Smith' → filename includes _ReducedSub_Helper_Jane_Smith_."""
        import datetime as dt
        rows = [self._make_group_row('reduced_sub_helper', helper_foreman='Jane Smith')]
        result = generate_weekly_pdfs.generate_excel(
            '041926_99887766_REDUCEDSUB_HELPER_Jane_Smith', rows,
            dt.datetime(2026, 4, 19), data_hash='deadbeefcafebab3',
        )
        filename = result[1]
        self.assertIn('_ReducedSub_Helper_Jane_Smith_', filename)


class TestSubcontractorHelperVariantDeptJobDisplay(unittest.TestCase):
    """REPORT DETAILS Dept #/Job # for subcontractor helper-shadow variants.

    Regression for the operator-reported defect (2026-05-21): the
    ``reduced_sub_helper`` / ``aep_billable_helper`` variants fell through
    ``generate_excel``'s ``if variant == 'helper'`` exact-match gate to the
    ``else`` (primary) branch, so the REPORT DETAILS block showed the PRIMARY
    ``Dept #`` / ``Job #`` instead of the helper's ``__helper_dept`` /
    ``__helper_job``.

    The displayed Foreman is already correct for these variants because
    ``__current_foreman`` is set to the *attributed* helper (the file's
    partition key) at the ``keys_to_add`` site — NOT ``__helper_foreman``
    (the current "Foreman Helping?" value, which can diverge under Phase 1.1
    claim attribution). The fix MUST preserve that, so the foreman regression
    guard below pins ``display_foreman == current_foreman``.
    """

    def setUp(self):
        self._tmpdir = tempfile.TemporaryDirectory()
        self._orig_output_folder = generate_weekly_pdfs.OUTPUT_FOLDER
        generate_weekly_pdfs.OUTPUT_FOLDER = self._tmpdir.name
        self._orig_rates = dict(generate_weekly_pdfs._SUBCONTRACTOR_RATES)
        generate_weekly_pdfs._SUBCONTRACTOR_RATES.clear()
        generate_weekly_pdfs._SUBCONTRACTOR_RATES['XYZ'] = {
            'cu_code': 'XYZ',
            'cu_wbs': '999',
            'compatible_unit_group': 'TestGroup',
            'reduced_install_price': 10.0,
            'reduced_remove_price': 5.0,
            'reduced_transfer_price': 2.5,
            'new_install_price': 20.0,
            'new_remove_price': 12.0,
            'new_transfer_price': 6.0,
        }

    def tearDown(self):
        generate_weekly_pdfs.OUTPUT_FOLDER = self._orig_output_folder
        self._tmpdir.cleanup()
        generate_weekly_pdfs._SUBCONTRACTOR_RATES.clear()
        generate_weekly_pdfs._SUBCONTRACTOR_RATES.update(self._orig_rates)

    # Distinct sentinels: PRIMARY dept/job = 500 / J-1; HELPER = 123 / J-2.
    def _make_sub_helper_row(self, variant, helper_foreman='Jane Smith'):
        import datetime as dt
        return {
            'Work Request #': '99887766',
            'Weekly Reference Logged Date': '2026-04-19',
            'Snapshot Date': '2026-04-19',
            'Units Completed?': True,
            'Units Total Price': '$0.00',
            'CU': 'XYZ',
            'Work Type': 'Install',
            'Quantity': 2,
            'Customer Name': 'TestCustomer',
            'Foreman': 'PrimaryForeman',
            'Dept #': '500',
            'Job #': 'J-1',
            '__effective_user': 'PrimaryForeman',
            # For sub-helper rows the engine sets __current_foreman to the
            # ATTRIBUTED helper (the partition key), not the primary foreman.
            '__current_foreman': helper_foreman,
            '__variant': variant,
            '__helper_foreman': helper_foreman,
            '__helper_dept': '123',
            '__helper_job': 'J-2',
            '__week_ending_date': dt.datetime(2026, 4, 19),
        }

    def _read_detail(self, excel_path, label):
        """Return the REPORT DETAILS value (column G) for a given F-column label."""
        import openpyxl
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=6).value == label:  # column F = label
                return ws.cell(row=r, column=7).value      # column G = value
        return None

    def _generate(self, variant, group_key, data_hash, row=None):
        import datetime as dt
        rows = [row if row is not None else self._make_sub_helper_row(variant)]
        result = generate_weekly_pdfs.generate_excel(
            group_key, rows, dt.datetime(2026, 4, 19), data_hash=data_hash,
        )
        return result[0]  # excel_path

    def test_reduced_sub_helper_shows_helper_dept_and_job(self):
        path = self._generate(
            'reduced_sub_helper',
            '041926_99887766_REDUCEDSUB_HELPER_Jane_Smith',
            'deadbeefcafe0001',
        )
        self.assertEqual(
            self._read_detail(path, 'Dept #:'), '123',
            "reduced_sub_helper file must show __helper_dept (123), not primary Dept # (500)",
        )
        self.assertEqual(
            self._read_detail(path, 'Job #:'), 'J-2',
            "reduced_sub_helper file must show __helper_job (J-2), not primary Job # (J-1)",
        )

    def test_aep_billable_helper_shows_helper_dept_and_job(self):
        path = self._generate(
            'aep_billable_helper',
            '041926_99887766_AEPBILLABLE_HELPER_Jane_Smith',
            'deadbeefcafe0002',
        )
        self.assertEqual(
            self._read_detail(path, 'Dept #:'), '123',
            "aep_billable_helper file must show __helper_dept (123), not primary Dept # (500)",
        )
        self.assertEqual(
            self._read_detail(path, 'Job #:'), 'J-2',
            "aep_billable_helper file must show __helper_job (J-2), not primary Job # (J-1)",
        )

    def test_sub_helper_foreman_stays_attributed_helper(self):
        """Foreman must remain the attributed helper (current_foreman), NOT be
        switched to __helper_foreman. Here they coincide ('Jane Smith') but the
        fix must route through current_foreman so attribution divergence is
        respected."""
        path = self._generate(
            'reduced_sub_helper',
            '041926_99887766_REDUCEDSUB_HELPER_Jane_Smith',
            'deadbeefcafe0003',
        )
        self.assertEqual(self._read_detail(path, 'Foreman:'), 'Jane Smith')

    def test_sub_primary_variants_still_show_primary_dept_and_job(self):
        """Guard the else-branch behaviour we deliberately keep: reduced_sub /
        aep_billable (primary) files show the PRIMARY Dept # / Job #."""
        import datetime as dt
        for variant, key in (
            ('reduced_sub', '041926_99887766_REDUCEDSUB'),
            ('aep_billable', '041926_99887766_AEPBILLABLE'),
        ):
            with self.subTest(variant=variant):
                row = self._make_sub_helper_row(variant, helper_foreman='')
                row['__current_foreman'] = 'ClaimerForeman'
                row['__helper_foreman'] = ''
                row['__helper_dept'] = ''
                row['__helper_job'] = ''
                result = generate_weekly_pdfs.generate_excel(
                    key, [row], dt.datetime(2026, 4, 19),
                    data_hash='deadbeefcafe0004',
                )
                path = result[0]
                self.assertEqual(self._read_detail(path, 'Dept #:'), '500')
                self.assertEqual(self._read_detail(path, 'Job #:'), 'J-1')


class TestSubcontractorVariantPriceSubstitution(unittest.TestCase):
    """Plan 01-03 Task 2: generate_excel substitutes CSV-driven prices for the 4 new variants.

    Tests verify the workbook on disk contains the rate × qty values
    instead of the row's SmartSheet ``Units Total Price`` value.
    """

    def setUp(self):
        self._tmpdir = tempfile.TemporaryDirectory()
        self._orig_output_folder = generate_weekly_pdfs.OUTPUT_FOLDER
        generate_weekly_pdfs.OUTPUT_FOLDER = self._tmpdir.name
        self._orig_rates = dict(generate_weekly_pdfs._SUBCONTRACTOR_RATES)
        generate_weekly_pdfs._SUBCONTRACTOR_RATES.clear()
        generate_weekly_pdfs._SUBCONTRACTOR_RATES['ALB-6-AUR1'] = {
            'cu_code': 'ALB-6-AUR1',
            'cu_wbs': '100',
            'compatible_unit_group': 'TestGroup',
            'reduced_install_price': 45.95,
            'reduced_remove_price': 33.33,
            'reduced_transfer_price': 106.54,
            'new_install_price': 52.58,
            'new_remove_price': 38.14,
            'new_transfer_price': 121.93,
        }

    def tearDown(self):
        generate_weekly_pdfs.OUTPUT_FOLDER = self._orig_output_folder
        self._tmpdir.cleanup()
        generate_weekly_pdfs._SUBCONTRACTOR_RATES.clear()
        generate_weekly_pdfs._SUBCONTRACTOR_RATES.update(self._orig_rates)

    def _make_row(self, variant, cu='ALB-6-AUR1', work_type='Install',
                  qty=2, smartsheet_price='$999.00'):
        import datetime as dt
        return {
            'Work Request #': '99887766',
            'Weekly Reference Logged Date': '2026-04-19',
            'Snapshot Date': '2026-04-19',
            'Units Completed?': True,
            'Units Total Price': smartsheet_price,
            'CU': cu,
            'Work Type': work_type,
            'Quantity': qty,
            'Customer Name': 'TestCustomer',
            'Foreman': 'TestForeman',
            'Dept #': '500',
            'Job #': 'J-1',
            '__effective_user': 'TestForeman',
            '__current_foreman': 'TestForeman',
            '__variant': variant,
            '__helper_foreman': '',
            '__helper_dept': '',
            '__helper_job': '',
            '__week_ending_date': dt.datetime(2026, 4, 19),
        }

    def _read_pricing_cells(self, excel_path):
        """Return all numeric values found in column H of the daily-block rows.

        Column H is the ``Pricing`` column written by ``write_day_block``.
        We collect every numeric value > 0 (skipping headers/labels which
        are strings, and the TOTAL row which uses the same column but
        represents a sum).
        """
        import openpyxl
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        prices = []
        totals = []
        for row in ws.iter_rows(values_only=True):
            for ci, val in enumerate(row):
                if ci == 7 and isinstance(val, (int, float)) and val > 0:
                    # Column H (0-indexed 7) is Pricing. Distinguish
                    # row-level prices from the TOTAL row by checking
                    # if column A contains 'TOTAL'.
                    if row[0] == 'TOTAL':
                        totals.append(val)
                    else:
                        prices.append(val)
        return prices, totals

    def test_aep_billable_workbook_uses_new_install_price_times_qty(self):
        """Test 3 (D-08): aep_billable row → Pricing cell = new_install_price × Quantity."""
        import datetime as dt
        rows = [self._make_row('aep_billable', work_type='Install', qty=3)]
        result = generate_weekly_pdfs.generate_excel(
            '041926_99887766_AEPBILLABLE', rows, dt.datetime(2026, 4, 19),
            data_hash='deadbeef00000001',
        )
        excel_path = result[0]
        prices, totals = self._read_pricing_cells(excel_path)
        # new_install_price 52.58 × 3 = 157.74
        self.assertEqual(len(prices), 1)
        self.assertAlmostEqual(prices[0], 157.74, places=2)
        # Must NOT be the SmartSheet 999.0
        self.assertNotAlmostEqual(prices[0], 999.0, places=2)

    def test_reduced_sub_workbook_uses_reduced_install_price_times_qty(self):
        """Test 4 (SUB-02): reduced_sub row → Pricing cell = reduced_install_price × Quantity."""
        import datetime as dt
        rows = [self._make_row('reduced_sub', work_type='Install', qty=4)]
        result = generate_weekly_pdfs.generate_excel(
            '041926_99887766_REDUCEDSUB', rows, dt.datetime(2026, 4, 19),
            data_hash='deadbeef00000002',
        )
        excel_path = result[0]
        prices, _ = self._read_pricing_cells(excel_path)
        # reduced_install_price 45.95 × 4 = 183.80
        self.assertEqual(len(prices), 1)
        self.assertAlmostEqual(prices[0], 183.80, places=2)

    def test_missing_cu_aep_billable_keeps_smartsheet_price(self):
        """Test 5 (D-16): unknown CU + aep_billable → keep SmartSheet price (no zero-out, no raise)."""
        import datetime as dt
        rows = [self._make_row('aep_billable', cu='UNKNOWN_CU', smartsheet_price='$77.50', qty=5)]
        result = generate_weekly_pdfs.generate_excel(
            '041926_99887766_AEPBILLABLE', rows, dt.datetime(2026, 4, 19),
            data_hash='deadbeef00000003',
        )
        excel_path = result[0]
        prices, _ = self._read_pricing_cells(excel_path)
        # Must retain SmartSheet 77.50 (NOT zero out, NOT compute from missing rate)
        self.assertEqual(len(prices), 1)
        self.assertAlmostEqual(prices[0], 77.50, places=2)
        # Missing CU must surface in the return tuple's missing_cus Counter
        missing_cus = result[4]
        self.assertIn('UNKNOWN_CU', missing_cus)

    def test_primary_variant_keeps_smartsheet_price_unchanged(self):
        """Test 7: primary variant → SmartSheet pricing unchanged (D-14 invariant)."""
        import datetime as dt
        rows = [self._make_row('primary', cu='ALB-6-AUR1', smartsheet_price='$42.42', qty=1)]
        # NOTE: primary group_key has no _AEPBILLABLE/_REDUCEDSUB suffix
        result = generate_weekly_pdfs.generate_excel(
            '041926_99887766', rows, dt.datetime(2026, 4, 19),
            data_hash='deadbeef00000004',
        )
        excel_path = result[0]
        prices, _ = self._read_pricing_cells(excel_path)
        # Even though ALB-6-AUR1 is in rates, primary variant must keep SmartSheet price
        self.assertEqual(len(prices), 1)
        self.assertAlmostEqual(prices[0], 42.42, places=2)


class TestGenerateExcelReturnTupleShape(unittest.TestCase):
    """Plan 01-03 Task 2 / Blocker 4: generate_excel returns a 5-tuple."""

    def setUp(self):
        self._tmpdir = tempfile.TemporaryDirectory()
        self._orig_output_folder = generate_weekly_pdfs.OUTPUT_FOLDER
        generate_weekly_pdfs.OUTPUT_FOLDER = self._tmpdir.name

    def tearDown(self):
        generate_weekly_pdfs.OUTPUT_FOLDER = self._orig_output_folder
        self._tmpdir.cleanup()

    def test_return_is_5_tuple(self):
        """Return is a 5-tuple: (excel_path, filename, wr_numbers, customer_name, missing_cus)."""
        import datetime as dt
        from collections import Counter
        row = {
            'Work Request #': '12345678',
            'Weekly Reference Logged Date': '2026-04-19',
            'Snapshot Date': '2026-04-19',
            'Units Completed?': True,
            'Units Total Price': '$100.00',
            'CU': 'ABC',
            'Work Type': 'Install',
            'Quantity': 2,
            'Customer Name': 'Acme',
            'Foreman': 'F1',
            'Dept #': '500',
            'Job #': 'J',
            '__effective_user': 'F1',
            '__current_foreman': 'F1',
            '__variant': 'primary',
            '__week_ending_date': dt.datetime(2026, 4, 19),
        }
        result = generate_weekly_pdfs.generate_excel(
            '041926_12345678', [row], dt.datetime(2026, 4, 19),
            data_hash='deadbeef00000099',
        )
        self.assertEqual(len(result), 5,
                         f"generate_excel must return 5-tuple per Blocker 4; got {len(result)}-tuple")
        excel_path, filename, wr_numbers, customer_name, missing_cus = result
        self.assertTrue(os.path.exists(excel_path))
        self.assertEqual(customer_name, 'Acme')
        self.assertIsInstance(missing_cus, Counter)


class TestSubcontractorMissingCUWarning(unittest.TestCase):
    """Plan 01-03 Task 2 Change 3 (D-17): one WARNING per sheet at end-of-sheet processing.

    The WARNING text MUST contain the marker
    ``'Subcontractor rates CSV missing'`` (added to _PII_LOG_MARKERS in
    Plan 2) so the Sentry sanitizer drops it from before_send_log.
    """

    def test_warning_text_marker_present_in_module(self):
        """The WARNING template uses the stable marker the sanitizer recognises."""
        import inspect
        import pipeline.orchestrate  # W6: end-of-sheet WARNING lives in main()
        src = (inspect.getsource(generate_weekly_pdfs)
               + "\n" + inspect.getsource(pipeline.orchestrate))
        self.assertIn(
            'Subcontractor rates CSV missing', src,
            "Missing-CU WARNING template must include the marker "
            "'Subcontractor rates CSV missing' for _PII_LOG_MARKERS sanitization",
        )


class TestSubcontractorVariantKillSwitchAndScope(unittest.TestCase):
    """Plan 01-03 Task 3: kill-switch + ORIG-folder no-op regression coverage.

    Pins three invariants:

    1. ``SUBCONTRACTOR_RATE_VARIANTS_ENABLED=False`` short-circuits the
       new variant emission at the per-row gate in ``group_source_rows``
       (D-13 — operator emergency kill).
    2. A row whose ``__source_sheet_id`` is in
       ``_FOLDER_DISCOVERED_ORIG_IDS`` ONLY (not in SUB) emits NO new
       variant keys regardless of kill-switch state (SUB-06 — original-
       contract folders are unreachable through the new code path).
    3. A row on a sheet misconfigured into BOTH sets follows
       subcontractor-membership: the per-row gate fires on SUB
       membership (subcontractor exclusion stays primary in
       ``_fetch_and_process_sheet`` per Living Ledger 2026-04-24 11:30
       so the subcontractor flow runs end-to-end).
    """

    _SUB_SHEET_ID = 8162920222379908
    _ORIG_SHEET_ID = 7644752003786628

    def setUp(self):
        # Snapshot module state per the 2026-04-22 16:05 ledger rule
        # on test isolation; a sub-test that mutates folder ids or
        # the kill switch must not leak to unrelated suites.
        self._orig_enabled = generate_weekly_pdfs.SUBCONTRACTOR_RATE_VARIANTS_ENABLED
        self._orig_sub_ids = set(generate_weekly_pdfs._FOLDER_DISCOVERED_SUB_IDS)
        self._orig_orig_ids = set(generate_weekly_pdfs._FOLDER_DISCOVERED_ORIG_IDS)

    def tearDown(self):
        generate_weekly_pdfs.SUBCONTRACTOR_RATE_VARIANTS_ENABLED = self._orig_enabled
        generate_weekly_pdfs._FOLDER_DISCOVERED_SUB_IDS.clear()
        generate_weekly_pdfs._FOLDER_DISCOVERED_SUB_IDS.update(self._orig_sub_ids)
        generate_weekly_pdfs._FOLDER_DISCOVERED_ORIG_IDS.clear()
        generate_weekly_pdfs._FOLDER_DISCOVERED_ORIG_IDS.update(self._orig_orig_ids)

    def _make_sub_row(self, wr='WR_K', snapshot='2026-04-19',
                       source_sheet_id=None):
        if source_sheet_id is None:
            source_sheet_id = self._SUB_SHEET_ID
        return {
            'Work Request #': wr,
            'Weekly Reference Logged Date': '2026-04-19',
            'Snapshot Date': snapshot,
            'Units Completed?': True,
            'Units Total Price': '$100.00',
            '__effective_user': 'TestForeman',
            '__assignment_method': 'FOREMAN_COLUMN',
            '__is_helper_row': False,
            '__helper_foreman': '',
            '__helper_dept': '',
            '__helper_job': '',
            '__is_vac_crew': False,
            '__source_sheet_id': source_sheet_id,
        }

    def test_kill_switch_disables_new_variant_emission(self):
        """Test 1: SUBCONTRACTOR_RATE_VARIANTS_ENABLED=False → no new variant keys (D-13).

        Phase 1.1 D-22 contract change (CLAUDE.md Living Ledger entry
        for Phase 1.1 closure): pre-Phase-1.1 this test implicitly
        assumed the additive contract — subcontractor rows produced
        ``_AEPBILLABLE`` + ``_REDUCEDSUB`` IN ADDITION TO the legacy
        primary key. Phase 1.1 Bug B1 (Plan 01.1-02) inverts this for
        subcontractor rows only: they now produce ONLY the variant
        keys (partitioning, not additive — see RESEARCH.md §B1 for
        the full rationale; CONTEXT.md D-22 for the override
        authorization). See CLAUDE.md Living Ledger entry rule (b)
        for the design-intent override.

        The kill-switch assertion (variants suppressed when off) is
        STILL VALID and remains in place. The NEW assertion documents
        the post-Phase-1.1 invariant: even when the variant kill
        switch is off, the subcontractor non-helper row STILL does
        NOT emit the legacy primary key. Bug B1 partitioning is
        INDEPENDENT of the variant kill switch — it's a structural
        change to ``group_source_rows`` (see Plan 01.1-02 and the
        sibling ``test_partitioning_contract_for_subcontractor_non_helper_rows``
        method below for the standalone partitioning assertion).
        """
        generate_weekly_pdfs.SUBCONTRACTOR_RATE_VARIANTS_ENABLED = False
        generate_weekly_pdfs._FOLDER_DISCOVERED_SUB_IDS.clear()
        generate_weekly_pdfs._FOLDER_DISCOVERED_SUB_IDS.add(self._SUB_SHEET_ID)
        row = self._make_sub_row(snapshot='2026-04-19')
        groups = generate_weekly_pdfs.group_source_rows([row])
        keys = list(groups.keys())
        # PRESERVED Plan 01-03 contract — variant kill switch
        # suppresses _AEPBILLABLE / _REDUCEDSUB keys regardless of
        # Phase 1.1 Bug B1's partitioning structural change.
        self.assertFalse(
            any('_AEPBILLABLE' in k for k in keys),
            f"Kill switch must suppress _AEPBILLABLE keys; got: {keys}",
        )
        self.assertFalse(
            any('_REDUCEDSUB' in k for k in keys),
            f"Kill switch must suppress _REDUCEDSUB keys; got: {keys}",
        )
        # NEW Phase 1.1 D-22 assertion — Bug B1 partitioning is
        # independent of the variant kill switch. The subcontractor
        # non-helper row produces NO legacy primary key, even when
        # the variant kill switch is OFF. The expected primary key
        # shape that WOULD have been emitted pre-Phase-1.1 is
        # ``041926_WR_K`` (the legacy ``{week}_{wr}`` format); it
        # must be absent.
        self.assertFalse(
            any(k == '041926_WR_K' for k in keys),
            f"Bug B1 partitioning is independent of variant kill switch "
            f"— legacy primary key MUST NOT be emitted for "
            f"subcontractor rows; got: {keys}",
        )

    def test_partitioning_contract_for_subcontractor_non_helper_rows(self):
        """Phase 1.1 Bug B1 (D-04 / SUB-09): subcontractor non-helper
        rows produce ONLY variant keys, NEVER the legacy primary key.

        Asserts the partitioning contract directly — independent of
        the variant kill switch. With the kill switch ON, subcontractor
        rows produce ``_REDUCEDSUB`` (unconditional) AND ``_AEPBILLABLE``
        (post-cutoff). The legacy primary key ``{week}_{wr}`` is NEVER
        emitted for subcontractor rows.

        See CLAUDE.md Living Ledger entry for Phase 1.1 closure rule
        (b) for the design-intent override rationale (additive →
        partitioning, scope-limited to subcontractor rows). Primary
        / original-contract / vac_crew rows are UNCHANGED — covered
        by ``test_orig_folder_sheet_emits_no_new_variants`` below.
        """
        generate_weekly_pdfs.SUBCONTRACTOR_RATE_VARIANTS_ENABLED = True
        generate_weekly_pdfs._FOLDER_DISCOVERED_SUB_IDS.clear()
        generate_weekly_pdfs._FOLDER_DISCOVERED_SUB_IDS.add(self._SUB_SHEET_ID)
        row = self._make_sub_row(snapshot='2026-04-19')
        groups = generate_weekly_pdfs.group_source_rows([row])
        keys = list(groups.keys())
        # Partitioning: variant keys present
        self.assertTrue(
            any('_REDUCEDSUB' in k and '_HELPER' not in k for k in keys),
            f"Partitioning: _REDUCEDSUB key must be emitted; got: {keys}",
        )
        self.assertTrue(
            any('_AEPBILLABLE' in k and '_HELPER' not in k for k in keys),
            f"Partitioning: _AEPBILLABLE key must be emitted "
            f"(post-cutoff snapshot); got: {keys}",
        )
        # Partitioning: NO legacy primary key (the Phase 1.1 B1 invariant)
        self.assertFalse(
            any(k == '041926_WR_K' for k in keys),
            f"Bug B1 partitioning: legacy primary key MUST NOT be "
            f"emitted for subcontractor rows; got: {keys}",
        )

    def test_orig_folder_sheet_emits_no_new_variants(self):
        """Test 2: ``__source_sheet_id`` in ORIG only → no new variant keys (SUB-06)."""
        generate_weekly_pdfs.SUBCONTRACTOR_RATE_VARIANTS_ENABLED = True
        generate_weekly_pdfs._FOLDER_DISCOVERED_SUB_IDS.clear()
        # Don't add the orig sheet to SUB
        generate_weekly_pdfs._FOLDER_DISCOVERED_ORIG_IDS.clear()
        generate_weekly_pdfs._FOLDER_DISCOVERED_ORIG_IDS.add(self._ORIG_SHEET_ID)
        row = self._make_sub_row(
            wr='WR_ORIG',
            snapshot='2026-04-19',
            source_sheet_id=self._ORIG_SHEET_ID,
        )
        groups = generate_weekly_pdfs.group_source_rows([row])
        keys = list(groups.keys())
        self.assertFalse(
            any('_AEPBILLABLE' in k or '_REDUCEDSUB' in k for k in keys),
            f"ORIG-only sheet must emit no new variant keys; got: {keys}",
        )

    def test_dual_folder_membership_subcontractor_precedence(self):
        """Test 3: sheet in BOTH SUB and ORIG → per-row gate emits new variants.

        Per D-22 / 2026-04-24 11:30, the subcontractor-exclusion check
        in ``_fetch_and_process_sheet`` (line ~3194) ensures
        subcontractor flow runs when a sheet is misconfigured into
        both folder sets; the new variant emission follows the same
        rule because the per-row gate fires on SUB membership (the row
        only carries one sheet id and that id's membership in
        ``_FOLDER_DISCOVERED_SUB_IDS`` triggers emission).
        """
        generate_weekly_pdfs.SUBCONTRACTOR_RATE_VARIANTS_ENABLED = True
        generate_weekly_pdfs._FOLDER_DISCOVERED_SUB_IDS.clear()
        generate_weekly_pdfs._FOLDER_DISCOVERED_ORIG_IDS.clear()
        # Misconfigure: sheet id in BOTH sets.
        shared_sheet_id = 9999999999999
        generate_weekly_pdfs._FOLDER_DISCOVERED_SUB_IDS.add(shared_sheet_id)
        generate_weekly_pdfs._FOLDER_DISCOVERED_ORIG_IDS.add(shared_sheet_id)
        row = self._make_sub_row(
            wr='WR_DUAL',
            snapshot='2026-04-19',
            source_sheet_id=shared_sheet_id,
        )
        groups = generate_weekly_pdfs.group_source_rows([row])
        keys = list(groups.keys())
        # Subcontractor membership wins for emission — confirm new
        # variant keys ARE present.
        self.assertTrue(
            any('_REDUCEDSUB' in k for k in keys),
            f"Dual membership: subcontractor precedence must emit _REDUCEDSUB; got: {keys}",
        )
        self.assertTrue(
            any('_AEPBILLABLE' in k for k in keys),
            f"Dual membership: subcontractor precedence must emit _AEPBILLABLE (post-cutoff); got: {keys}",
        )

    def test_unparseable_snapshot_date_does_not_emit_aep_billable(self):
        """Test 4: unparseable Snapshot Date → _REDUCEDSUB only, no _AEPBILLABLE.

        Defensive guard: ``excel_serial_to_date('not-a-date')``
        returns ``None`` so the AEP-Billable cutoff check returns
        False safely. ReducedSub is unconditional and still emits.
        """
        generate_weekly_pdfs.SUBCONTRACTOR_RATE_VARIANTS_ENABLED = True
        generate_weekly_pdfs._FOLDER_DISCOVERED_SUB_IDS.clear()
        generate_weekly_pdfs._FOLDER_DISCOVERED_SUB_IDS.add(self._SUB_SHEET_ID)
        row = self._make_sub_row(
            wr='WR_BAD_SNAP',
            snapshot='not-a-date',  # unparseable
            source_sheet_id=self._SUB_SHEET_ID,
        )
        groups = generate_weekly_pdfs.group_source_rows([row])
        keys = list(groups.keys())
        self.assertTrue(
            any('_REDUCEDSUB' in k for k in keys),
            f"Unparseable snapshot: _REDUCEDSUB still unconditional; got: {keys}",
        )
        self.assertFalse(
            any('_AEPBILLABLE' in k for k in keys),
            f"Unparseable snapshot must suppress _AEPBILLABLE; got: {keys}",
        )


class TestSubcontractorB1PartitioningGate(unittest.TestCase):
    """Phase 1.1 Plan 01.1-02 Bug B1 (D-04 / SUB-09): partitioning gate.

    Pins the seven behaviours from the plan's ``<behavior>`` block:

    - Test 1: Non-helper subcontractor row (``is_subcontractor_row=True``,
      ``valid_helper_row=False``) under ``RES_GROUPING_MODE='both'``
      produces NO ``('primary', ...)`` group tuple — only the variant
      keys ``_REDUCEDSUB`` and ``_AEPBILLABLE`` (when post-cutoff). This
      is Bug B1's closure assertion.
    - Test 2: Non-helper NON-subcontractor row → legacy primary key
      ``{week}_{wr}`` IS emitted (legacy behaviour preserved).
    - Test 3: Helper subcontractor row → no legacy primary key
      (already partitioned by the existing ``not valid_helper_row``
      branch; Bug B1's gate must not regress this).
    - Test 4: Helper NON-subcontractor row → no legacy primary key
      AND the legacy "Helper row with both checkboxes" INFO log fires.
    - Test 5: ``is_subcontractor_row`` is hoisted (source-level grep
      guard — single ``_FOLDER_DISCOVERED_SUB_IDS`` membership
      expression inside ``group_source_rows``).
    - Test 6: ``RES_GROUPING_MODE='primary'`` is UNTOUCHED — both
      subcontractor and non-subcontractor non-helper rows produce a
      primary key in that mode (non-production configuration,
      preserves legacy behaviour per RESEARCH.md cascade-preservation
      rule).
    - Test 7: Subcontractor variant emission block continues to emit
      ``_REDUCEDSUB`` (and ``_AEPBILLABLE`` post-cutoff) for
      subcontractor rows — Bug B1 does not regress Phase 1's variant
      tagging; only suppresses the duplicate legacy primary.
    """

    _SUB_SHEET_ID = 8162920222379908
    _NON_SUB_SHEET_ID = 5723337641643908  # TARGET_SHEET_ID — known non-sub
    _ORIG_SHEET_ID = 7644752003786628

    def setUp(self):
        # Snapshot module state for test isolation (mirrors
        # TestSubcontractorVariantKillSwitchAndScope pattern).
        self._orig_enabled = generate_weekly_pdfs.SUBCONTRACTOR_RATE_VARIANTS_ENABLED
        self._orig_sub_ids = set(generate_weekly_pdfs._FOLDER_DISCOVERED_SUB_IDS)
        self._orig_orig_ids = set(generate_weekly_pdfs._FOLDER_DISCOVERED_ORIG_IDS)
        self._orig_mode = generate_weekly_pdfs.RES_GROUPING_MODE
        # Sub-project D (2026-05-25) pins this OFF: this class asserts the
        # legacy bare-primary behavior for NON-subcontractor rows, which D's
        # default-on production-primary partitioning would otherwise change
        # to _User_<claimer>. D's new behavior is covered by
        # tests/test_primary_claim_attribution.py::TestPrimaryEmission. Per
        # [2026-05-20 00:26] rule 2 (test-contract override), pin D off here
        # to keep this an isolated B/B1 guard.
        self._orig_primary_attr = generate_weekly_pdfs.PRIMARY_CLAIM_ATTRIBUTION_ENABLED
        generate_weekly_pdfs.PRIMARY_CLAIM_ATTRIBUTION_ENABLED = False
        # Default test config — production-like 'both' mode + variants on.
        generate_weekly_pdfs.RES_GROUPING_MODE = 'both'
        generate_weekly_pdfs.SUBCONTRACTOR_RATE_VARIANTS_ENABLED = True
        generate_weekly_pdfs._FOLDER_DISCOVERED_SUB_IDS.clear()
        generate_weekly_pdfs._FOLDER_DISCOVERED_SUB_IDS.add(self._SUB_SHEET_ID)
        generate_weekly_pdfs._FOLDER_DISCOVERED_ORIG_IDS.clear()

    def tearDown(self):
        generate_weekly_pdfs.PRIMARY_CLAIM_ATTRIBUTION_ENABLED = self._orig_primary_attr
        generate_weekly_pdfs.SUBCONTRACTOR_RATE_VARIANTS_ENABLED = self._orig_enabled
        generate_weekly_pdfs._FOLDER_DISCOVERED_SUB_IDS.clear()
        generate_weekly_pdfs._FOLDER_DISCOVERED_SUB_IDS.update(self._orig_sub_ids)
        generate_weekly_pdfs._FOLDER_DISCOVERED_ORIG_IDS.clear()
        generate_weekly_pdfs._FOLDER_DISCOVERED_ORIG_IDS.update(self._orig_orig_ids)
        generate_weekly_pdfs.RES_GROUPING_MODE = self._orig_mode

    def _make_row(
        self,
        wr='WR_PARTITION',
        source_sheet_id=None,
        is_helper_row=False,
        helper_foreman='',
        helper_dept='',
        snapshot='2026-04-19',
    ):
        if source_sheet_id is None:
            source_sheet_id = self._SUB_SHEET_ID
        return {
            'Work Request #': wr,
            'Weekly Reference Logged Date': '2026-04-19',
            'Snapshot Date': snapshot,
            'Units Completed?': True,
            'Units Total Price': '$100.00',
            '__effective_user': 'TestForeman',
            '__assignment_method': 'FOREMAN_COLUMN',
            '__is_helper_row': is_helper_row,
            '__helper_foreman': helper_foreman,
            '__helper_dept': helper_dept,
            '__helper_job': '',
            '__is_vac_crew': False,
            '__source_sheet_id': source_sheet_id,
        }

    # ---------- Test 1: Bug B1 closure assertion ----------
    def test_subcontractor_non_helper_row_emits_no_primary_key(self):
        """Test 1: non-helper subcontractor row → NO ('primary', ...) tuple.

        The legacy primary key ``{week}_{wr}`` MUST NOT appear in the
        emitted groups for a subcontractor non-helper row under
        production grouping mode 'both'. Only variant keys remain.

        Subproject B (2026-05-20) design-intent override per the
        [2026-05-20 00:26] Living Ledger rule 2: the variant keys are
        now PARTITIONED by the frozen primary claimer
        (``_REDUCEDSUB_USER_<claimer>`` / ``_AEPBILLABLE_USER_<claimer>``)
        instead of the bare ``_REDUCEDSUB`` / ``_AEPBILLABLE`` form. The
        no-legacy-primary closure assertion (the actual subject of Bug
        B1) is unchanged; only the variant-key shape assertion is
        updated to the partitioned form. A substring match tolerates the
        ``_USER_<claimer>`` suffix while still proving the variant
        emission fires.
        """
        row = self._make_row(
            wr='WR_SUB',
            source_sheet_id=self._SUB_SHEET_ID,
            is_helper_row=False,
        )
        groups = generate_weekly_pdfs.group_source_rows([row])
        keys = list(groups.keys())
        # Bug B1: no legacy primary
        self.assertNotIn(
            '041926_WR_SUB',
            keys,
            f"Bug B1: subcontractor non-helper row must NOT emit "
            f"legacy primary key; got: {keys}",
        )
        # Phase 1 variant emission is preserved, now partitioned by the
        # frozen primary claimer (Subproject B). effective_user is
        # 'TestForeman' and no resolve_claimer mock is active here, so
        # the real resolve_claimer falls back to use-current.
        self.assertTrue(
            any('041926_WR_SUB_REDUCEDSUB_USER_' in k for k in keys),
            f"Variant emission must still produce a partitioned "
            f"_REDUCEDSUB_USER_<claimer> key; got: {keys}",
        )
        self.assertTrue(
            any('041926_WR_SUB_AEPBILLABLE_USER_' in k for k in keys),
            f"Post-cutoff snapshot must still produce a partitioned "
            f"_AEPBILLABLE_USER_<claimer> key; got: {keys}",
        )

    # ---------- Test 2: legacy non-subcontractor flow preserved ----------
    def test_non_subcontractor_non_helper_row_emits_primary_key(self):
        """Test 2: non-helper non-sub row → legacy primary key preserved."""
        row = self._make_row(
            wr='WR_NONSUB',
            source_sheet_id=self._NON_SUB_SHEET_ID,
            is_helper_row=False,
        )
        groups = generate_weekly_pdfs.group_source_rows([row])
        keys = list(groups.keys())
        self.assertIn(
            '041926_WR_NONSUB',
            keys,
            f"Non-subcontractor row must continue to emit legacy "
            f"primary key; got: {keys}",
        )
        # And NO subcontractor variant keys
        self.assertFalse(
            any('_REDUCEDSUB' in k or '_AEPBILLABLE' in k for k in keys),
            f"Non-subcontractor row must NOT emit variant keys; got: {keys}",
        )

    # ---------- Test 3: helper subcontractor row still excluded ----------
    def test_helper_subcontractor_row_emits_no_primary_key(self):
        """Test 3: helper subcontractor row → no primary (existing behaviour)."""
        row = self._make_row(
            wr='WR_SUBHELP',
            source_sheet_id=self._SUB_SHEET_ID,
            is_helper_row=True,
            helper_foreman='Jane Helper',
            helper_dept='500',
        )
        groups = generate_weekly_pdfs.group_source_rows([row])
        keys = list(groups.keys())
        self.assertNotIn(
            '041926_WR_SUBHELP',
            keys,
            f"Helper subcontractor row must NOT emit legacy primary; "
            f"got: {keys}",
        )
        # Helper variants for sub rows still emit
        self.assertTrue(
            any('_REDUCEDSUB_HELPER_' in k for k in keys),
            f"Helper sub row must emit _REDUCEDSUB_HELPER_ key; "
            f"got: {keys}",
        )

    # ---------- Test 4: helper non-sub row still emits "Helper" log ----------
    def test_helper_non_subcontractor_row_emits_no_primary_and_logs(self):
        """Test 4: helper non-sub row → no primary, "Helper row" INFO log fires.

        The legacy ``elif valid_helper_row:`` branch in the cascade
        must still fire for helper non-subcontractor rows (Bug B1's
        gate change must not regress the "Helper row with both
        checkboxes" log).
        """
        row = self._make_row(
            wr='WR_NONSUBHELP',
            source_sheet_id=self._NON_SUB_SHEET_ID,
            is_helper_row=True,
            helper_foreman='Jane NonSubHelper',
            helper_dept='500',
        )
        with self.assertLogs(level='INFO') as cm:
            groups = generate_weekly_pdfs.group_source_rows([row])
        keys = list(groups.keys())
        self.assertNotIn(
            '041926_WR_NONSUBHELP',
            keys,
            f"Helper non-sub row must NOT emit legacy primary; "
            f"got: {keys}",
        )
        # The "Helper row with both checkboxes" INFO log must fire
        log_blob = '\n'.join(cm.output)
        self.assertIn(
            'Helper row with both checkboxes',
            log_blob,
            "Helper non-sub row must trigger the legacy 'Helper row "
            "with both checkboxes' INFO log (unchanged behaviour)",
        )

    # ---------- Test 5: source-level hoist guard ----------
    def test_source_level_grep_partitioning_gate_present(self):
        """Test 5: production source carries the partitioning gate.

        Defeats the "tests pass but production reverted" failure mode
        (mirror of TestHelperShadowVariantFileIdentifier's source-level
        grep test pattern).
        """
        import inspect
        import pathlib
        import pipeline.grouping  # W4: group_source_rows relocated here
        src = (
            pathlib.Path(
                inspect.getsourcefile(generate_weekly_pdfs)
            ).read_text(encoding='utf-8')
            + "\n"
            + pathlib.Path(
                inspect.getsourcefile(pipeline.grouping)
            ).read_text(encoding='utf-8')
        )
        self.assertIn(
            'if not is_subcontractor_row and not valid_helper_row:',
            src,
            "Bug B1 partitioning gate must be present in production",
        )
        self.assertIn(
            'elif is_subcontractor_row and not valid_helper_row:',
            src,
            "Bug B1 diagnostic branch must be present in production",
        )
        self.assertIn(
            'EXCLUDING from main Excel (subcontractor row)',
            src,
            "Bug B1 diagnostic DEBUG log body must be present",
        )
        # Hoist confirmation: the `is_subcontractor_row` assignment
        # must appear exactly ONCE inside group_source_rows (the
        # previously-duplicated site at L4244-4248 was removed).
        # We count occurrences of the multi-line assignment header.
        # The hoisted block is `is_subcontractor_row = (` and must
        # appear exactly once in the source.
        self.assertEqual(
            src.count('is_subcontractor_row = (\n'),
            1,
            "is_subcontractor_row must be assigned exactly once "
            "(hoisted; duplicated computation removed)",
        )

    # ---------- Test 6: RES_GROUPING_MODE='primary' untouched ----------
    def test_primary_mode_subcontractor_row_still_emits_primary(self):
        """Test 6: under RES_GROUPING_MODE='primary', sub rows still get primary key.

        Non-production mode; the partitioning gate is scoped to the
        ``('helper', 'both')`` branch only per the RESEARCH.md
        cascade-preservation rule.
        """
        generate_weekly_pdfs.RES_GROUPING_MODE = 'primary'
        row = self._make_row(
            wr='WR_SUB_PRIM',
            source_sheet_id=self._SUB_SHEET_ID,
            is_helper_row=False,
        )
        groups = generate_weekly_pdfs.group_source_rows([row])
        keys = list(groups.keys())
        self.assertIn(
            '041926_WR_SUB_PRIM',
            keys,
            f"RES_GROUPING_MODE='primary' must continue to emit "
            f"legacy primary key for ALL rows (sub + non-sub); "
            f"got: {keys}",
        )

    # ---------- Test 7: variant emission preserved ----------
    def test_phase1_variant_emission_block_untouched(self):
        """Test 7: subcontractor variant emission still produces variants.

        Bug B1 closes the duplicate-primary leak WITHOUT regressing the
        variant emission — _REDUCEDSUB (always) and _AEPBILLABLE
        (post-cutoff) continue to fire for subcontractor rows.

        Subproject B (2026-05-20) design-intent override per the
        [2026-05-20 00:26] Living Ledger rule 2: those variant keys are
        now PARTITIONED by the frozen primary claimer
        (``_REDUCEDSUB_USER_<claimer>`` / ``_AEPBILLABLE_USER_<claimer>``).
        The substring assertions below tolerate the new
        ``_USER_<claimer>`` suffix while still proving both variants
        emit for a post-cutoff subcontractor row.
        """
        row = self._make_row(
            wr='WR_VAR_INTACT',
            source_sheet_id=self._SUB_SHEET_ID,
            is_helper_row=False,
            snapshot='2026-04-19',  # post-cutoff
        )
        groups = generate_weekly_pdfs.group_source_rows([row])
        keys = list(groups.keys())
        self.assertTrue(
            any('041926_WR_VAR_INTACT_REDUCEDSUB_USER_' in k for k in keys),
            f"_REDUCEDSUB_USER_<claimer> still emitted for "
            f"subcontractor rows; got: {keys}",
        )
        self.assertTrue(
            any('041926_WR_VAR_INTACT_AEPBILLABLE_USER_' in k for k in keys),
            f"_AEPBILLABLE_USER_<claimer> still emitted for post-cutoff "
            f"subcontractor rows; got: {keys}",
        )


class TestSubcontractorPppSheetIdEmptyStringDisable(unittest.TestCase):
    """Phase 01 gap closure (REVIEW-WR-02): ``SUBCONTRACTOR_PPP_SHEET_ID=''``
    must resolve to ``0`` (disabled), matching the operator-facing
    documentation. Pre-fix the empty string silently fell back to
    the hardcoded default ``8162920222379908`` — asymmetric with
    ``'0'``, which already disabled correctly via the downstream
    truthy gate. The fix is a single-line special case at the
    SUBCONTRACTOR_PPP_SHEET_ID call site (not inside
    ``_coerce_sheet_id``, which is shared with ``TARGET_SHEET_ID``
    where default-fallback is correct).

    The reload pattern + Sentry-DSN suppression mirrors
    ``_safe_reload_gwp`` in ``test_performance_optimizations.py``
    per Living Ledger 2026-04-22 16:05.
    """

    def setUp(self):
        # Snapshot env so tearDown restores it cleanly.
        self._env_snapshot = {
            k: os.environ.get(k)
            for k in (
                'SUBCONTRACTOR_PPP_SHEET_ID',
                'SENTRY_DSN',
            )
        }
        # Defense-in-depth: silence Sentry during reload so a
        # developer's local DSN doesn't fire a real Sentry init.
        os.environ['SENTRY_DSN'] = ''

    def tearDown(self):
        for k, v in self._env_snapshot.items():
            if v is None:
                os.environ.pop(k, None)
            else:
                os.environ[k] = v
        # Restore module to its baseline by reloading once more under
        # the same Sentry-suppression bracket per the 2026-04-22 16:05
        # ledger rule.
        with mock.patch.dict(os.environ, {"SENTRY_DSN": ""}, clear=False):
            with mock.patch('sentry_sdk.init'):
                importlib.reload(generate_weekly_pdfs)

    def _reload_with_ppp(self, value):
        if value is None:
            os.environ.pop('SUBCONTRACTOR_PPP_SHEET_ID', None)
        else:
            os.environ['SUBCONTRACTOR_PPP_SHEET_ID'] = value
        # Silence Sentry init on reload.
        with mock.patch.dict(os.environ, {"SENTRY_DSN": ""}, clear=False):
            with mock.patch('sentry_sdk.init'):
                importlib.reload(generate_weekly_pdfs)

    def test_empty_string_disables_ppp(self):
        self._reload_with_ppp('')
        self.assertEqual(
            generate_weekly_pdfs.SUBCONTRACTOR_PPP_SHEET_ID, 0,
            "Per WR-02, SUBCONTRACTOR_PPP_SHEET_ID='' must resolve "
            "to 0 (disabled), not fall back to the hardcoded default."
        )

    def test_zero_string_disables_ppp(self):
        self._reload_with_ppp('0')
        self.assertEqual(
            generate_weekly_pdfs.SUBCONTRACTOR_PPP_SHEET_ID, 0,
            "Per pre-existing behavior, SUBCONTRACTOR_PPP_SHEET_ID='0' "
            "resolves to 0 — the empty-string fix must preserve this."
        )

    def test_unset_uses_hardcoded_default(self):
        self._reload_with_ppp(None)
        self.assertEqual(
            generate_weekly_pdfs.SUBCONTRACTOR_PPP_SHEET_ID,
            8162920222379908,
            "Per the default contract, an unset env var resolves to "
            "the hardcoded default.",
        )

    def test_invalid_value_falls_back_to_default(self):
        self._reload_with_ppp('not-an-int')
        # _coerce_sheet_id WARN-and-fallback behavior is preserved.
        # The special case in WR-02 ONLY fires on the literal '',
        # NOT on any other non-integer value.
        self.assertEqual(
            generate_weekly_pdfs.SUBCONTRACTOR_PPP_SHEET_ID,
            8162920222379908,
            "Per pre-existing _coerce_sheet_id behavior, a non-integer "
            "non-empty value falls back to the hardcoded default with "
            "a WARNING.",
        )

    def test_integer_string_passes_through(self):
        self._reload_with_ppp('1234567890')
        self.assertEqual(
            generate_weekly_pdfs.SUBCONTRACTOR_PPP_SHEET_ID, 1234567890,
        )


class TestHelperShadowSuffixDefensiveRaise(unittest.TestCase):
    """Phase 01 gap closure (REVIEW-WR-03): the two new
    helper-shadow filename-suffix branches in ``generate_excel``
    raise ``ValueError`` if ``__helper_foreman`` is empty, instead
    of silently producing a primary-looking filename. The
    legacy ``helper`` branch is explicitly OUT OF SCOPE per
    01-REVIEW.md and stays unchanged.

    These tests construct minimal groups with the variant
    tagged but ``__helper_foreman=''`` and call ``generate_excel``.
    The raise fires inside the variant-suffix builder, which runs
    BEFORE workbook construction, so the tests do not need a full
    openpyxl scaffold.
    """

    @staticmethod
    def _make_group(variant, helper_foreman=''):
        """Build a 1-row group eligible for the variant-suffix
        builder. Includes the minimum fields ``generate_excel``
        needs to reach the variant branch (wr_num / week_end_raw
        derivation depends on ``Work Request #`` and the group_key).
        """
        import datetime as _dt
        return [{
            'Work Request #': '91467680',
            'Weekly Reference Logged Date': '04/19/26',
            'Snapshot Date': '04/19/26',
            'Units Completed?': True,
            'Units Total Price': '$100.00',
            'CU': 'ABC123',
            'Work Type': 'Install',
            'Quantity': 1,
            'Customer Name': 'TestCustomer',
            'Foreman': 'TestForeman',
            'Dept #': '500',
            'Job #': 'JOB-99',
            '__effective_user': 'TestForeman',
            '__current_foreman': 'TestForeman',
            '__variant': variant,
            '__helper_foreman': helper_foreman,
            '__helper_dept': '500',
            '__helper_job': 'JOB-99',
            '__week_ending_date': _dt.datetime(2026, 4, 19),
        }]

    def setUp(self):
        # Direct OUTPUT_FOLDER into a temp dir so the (legacy-branch)
        # workbook test doesn't pollute the repo. The defensive-raise
        # tests don't reach workbook construction but the legacy test
        # does (it must NOT raise, so the function continues to
        # workbook write).
        self._tmpdir = tempfile.TemporaryDirectory()
        self._orig_output_folder = generate_weekly_pdfs.OUTPUT_FOLDER
        generate_weekly_pdfs.OUTPUT_FOLDER = self._tmpdir.name

    def tearDown(self):
        generate_weekly_pdfs.OUTPUT_FOLDER = self._orig_output_folder
        self._tmpdir.cleanup()

    def test_aep_billable_helper_empty_foreman_raises_value_error(self):
        import datetime as _dt
        group = self._make_group('aep_billable_helper', helper_foreman='')
        with self.assertRaises(ValueError) as ctx:
            generate_weekly_pdfs.generate_excel(
                '041926_91467680',
                group,
                _dt.datetime(2026, 4, 19),
                data_hash='deadbeefcafebab4',
            )
        # Message body MUST contain WR + variant name but NOT
        # foreman / dept / job (PII per CLAUDE.md / Living Ledger
        # 2026-04-20 12:00). __helper_foreman='' so there is no
        # foreman value to leak; the dept/job values from
        # _make_group MUST NOT appear in the message body.
        msg = str(ctx.exception)
        self.assertIn('aep_billable_helper', msg)
        self.assertIn('91467680', msg)
        self.assertNotIn('500', msg, f"dept # leaked into raise body: {msg!r}")
        self.assertNotIn('JOB-99', msg, f"job # leaked into raise body: {msg!r}")

    def test_reduced_sub_helper_empty_foreman_raises_value_error(self):
        import datetime as _dt
        group = self._make_group('reduced_sub_helper', helper_foreman='')
        with self.assertRaises(ValueError) as ctx:
            generate_weekly_pdfs.generate_excel(
                '041926_91467680',
                group,
                _dt.datetime(2026, 4, 19),
                data_hash='deadbeefcafebab5',
            )
        msg = str(ctx.exception)
        self.assertIn('reduced_sub_helper', msg)
        self.assertIn('91467680', msg)
        self.assertNotIn('500', msg, f"dept # leaked into raise body: {msg!r}")
        self.assertNotIn('JOB-99', msg, f"job # leaked into raise body: {msg!r}")

    def test_legacy_helper_branch_does_not_raise_on_empty_foreman(self):
        # 01-REVIEW.md WR-03 explicit scope restriction: the legacy
        # ``helper`` branch has the same silent-fallthrough shape but
        # is OUT OF SCOPE. A future tech-debt-cleanup plan can add
        # the defensive raise there with its own regression test.
        # Until then, this test guards against an accidental
        # broadening of the WR-03 fix that would regress the legacy
        # helper variant production path.
        import datetime as _dt
        group = self._make_group('helper', helper_foreman='')
        try:
            generate_weekly_pdfs.generate_excel(
                '041926_91467680',
                group,
                _dt.datetime(2026, 4, 19),
                data_hash='deadbeefcafebab6',
            )
        except ValueError as exc:
            self.fail(
                f"Legacy ``helper`` branch raised ValueError unexpectedly: "
                f"{exc!r}. WR-03 scope restriction: defensive raise is "
                f"added to NEW shadow variants only. If a future plan "
                f"intentionally extends the raise to the legacy branch, "
                f"remove this test and add the symmetric raise+test pair."
            )
        except Exception:
            # Other exceptions (e.g., openpyxl Workbook init failure
            # in TEST_MODE-off path, an unrelated bug in legacy helper
            # flow) are acceptable — they're not the WR-03 contract
            # surface. The WR-03 contract is specifically "no
            # ValueError raised here from the missing-foreman branch."
            pass

    def test_production_aep_billable_helper_branch_has_defensive_raise(self):
        # Source-level guard: confirm the raise landed. Belt-and-
        # suspenders complement to the behavioral tests above.
        import inspect
        import pathlib
        # W4: variant-suffix / generate_excel logic relocated to
        # pipeline/excel.py; group_source_rows to pipeline/grouping.py —
        # grep facade + both relocated modules so the guard follows the code.
        import pipeline.grouping
        import pipeline.excel
        src = (
            pathlib.Path(
                inspect.getsourcefile(generate_weekly_pdfs)
            ).read_text(encoding='utf-8')
            + "\n"
            + pathlib.Path(
                inspect.getsourcefile(pipeline.grouping)
            ).read_text(encoding='utf-8')
            + "\n"
            + pathlib.Path(
                inspect.getsourcefile(pipeline.excel)
            ).read_text(encoding='utf-8')
        )
        # The raise text should be unique enough to grep for.
        self.assertIn(
            'aep_billable_helper requires __helper_foreman', src,
        )
        self.assertIn(
            'reduced_sub_helper requires __helper_foreman', src,
        )


class TestSubcontractorVariantOpenpyxlCompliance(unittest.TestCase):
    """Plan 01-03 Task 2 Test 8: no raw merge_cells / no xlsxwriter / no oddFooter."""

    def test_no_xlsxwriter_import(self):
        """openpyxl-only contract preserved (.claude/rules/smartsheet-python-optimization.md)."""
        with open(generate_weekly_pdfs.__file__, 'r', encoding='utf-8') as f:
            src = f.read()
        self.assertNotIn('xlsxwriter', src)

    def test_no_oddFooter_right_text_assignment(self):
        """Known XML corruption vector must remain absent as an assignment (CLAUDE.md Critical Pitfalls).

        The existing module has a single NOTE comment mentioning the
        attribute name as a guardrail for future contributors — that
        is expected. Ban only the *assignment* pattern.
        """
        with open(generate_weekly_pdfs.__file__, 'r', encoding='utf-8') as f:
            src = f.read()
        # Look for assignment targets like ``.oddFooter.right.text = `` or
        # an actual attribute write. The comment line is the only
        # legitimate occurrence today; an assignment would be a regression.
        import re
        assignment_pattern = re.compile(r'\.oddFooter\.right\.text\s*=')
        hits = assignment_pattern.findall(src)
        self.assertEqual(hits, [], "oddFooter.right.text MUST NOT be assigned — XML corruption vector")


class TestPhase1IntegrationRegression(unittest.TestCase):
    """Locks ROADMAP success criterion 5 — the byte-identical
    guarantee for primary / helper / vac_crew Excel files under the
    introduction of ``_SUBCONTRACTOR_RATES_FINGERPRINT``.

    Any future change that accidentally widens the fingerprint mix-in
    to existing variants will fail Tests 1-3 here. Test 4 is the
    positive companion — it asserts that the fingerprint IS mixed in
    for the new ``aep_billable`` variant (and by symmetry, the other
    three new variants), so a regression that drops the mix-in
    silently for the variants that need it also fails loudly.

    Phase 01 Plan 06 Task 2. Split from prior planning text per
    round-3 checker Warning 11 — hash stability + filename round-trip
    are different invariants; this class owns hash stability only.
    """

    def setUp(self):
        # Pin module globals that calculate_data_hash() reads so
        # tests are robust against env-var overrides in developer
        # shells. Mirrors the pinning in
        # tests/test_vac_crew.py::TestVacCrewHashAggregation. The
        # new fourth pin (_SUBCONTRACTOR_RATES_FINGERPRINT) is the
        # variable this class actively mutates, so capturing and
        # restoring it is essential — leaking a mutated value into
        # later tests would destabilize them.
        self._saved_ext = generate_weekly_pdfs.EXTENDED_CHANGE_DETECTION
        self._saved_cutoff = generate_weekly_pdfs.RATE_CUTOFF_DATE
        self._saved_rates_fp = generate_weekly_pdfs._RATES_FINGERPRINT
        self._saved_sub_fp = (
            generate_weekly_pdfs._SUBCONTRACTOR_RATES_FINGERPRINT
        )
        generate_weekly_pdfs.EXTENDED_CHANGE_DETECTION = True
        generate_weekly_pdfs.RATE_CUTOFF_DATE = None
        generate_weekly_pdfs._RATES_FINGERPRINT = ''
        generate_weekly_pdfs._SUBCONTRACTOR_RATES_FINGERPRINT = ''

    def tearDown(self):
        generate_weekly_pdfs.EXTENDED_CHANGE_DETECTION = self._saved_ext
        generate_weekly_pdfs.RATE_CUTOFF_DATE = self._saved_cutoff
        generate_weekly_pdfs._RATES_FINGERPRINT = self._saved_rates_fp
        generate_weekly_pdfs._SUBCONTRACTOR_RATES_FINGERPRINT = (
            self._saved_sub_fp
        )

    def _primary_row(self, cu='CU-A', qty=1, price='$100.00',
                     pole='P-1'):
        return {
            'Work Request #': '91467680',
            'Snapshot Date': '2026-04-19',
            'CU': cu,
            'Quantity': qty,
            'Pole #': pole,
            'Work Type': 'Install',
            'Dept #': '500',
            'Units Total Price': price,
            'Units Completed?': True,
            '__variant': 'primary',
            'Foreman': 'AlicePrimary',
            '__effective_user': 'AlicePrimary',
        }

    def _helper_row(self, cu='CU-H1', qty=1, price='$200.00',
                    pole='P-H1', helper='BobHelper'):
        return {
            'Work Request #': '91467680',
            'Snapshot Date': '2026-04-19',
            'CU': cu,
            'Quantity': qty,
            'Pole #': pole,
            'Work Type': 'Install',
            'Dept #': '600',
            'Units Total Price': price,
            'Units Completed?': True,
            '__variant': 'helper',
            'Foreman': 'AlicePrimary',
            '__effective_user': 'AlicePrimary',
            '__current_foreman': helper,
            '__helper_foreman': helper,
            '__helper_dept': '600',
            '__helper_job': 'J1',
        }

    def _vac_crew_row(self, cu='CU-V1', qty=1, price='$300.00',
                      pole='P-V1', name='VacMember1'):
        return {
            'Work Request #': '91467680',
            'Snapshot Date': '2026-04-19',
            'CU': cu,
            'Quantity': qty,
            'Pole #': pole,
            'Work Type': 'Vacuum Switch',
            'Dept #': '700',
            'Units Total Price': price,
            'Units Completed?': True,
            '__variant': 'vac_crew',
            'Foreman': 'AlicePrimary',
            '__effective_user': 'AlicePrimary',
            '__current_foreman': name,
            '__vac_crew_name': name,
            '__vac_crew_dept': '700',
            '__vac_crew_job': 'VJ1',
        }

    def _aep_billable_row(self, cu='CU-S1', qty=1, price='$50.00',
                          pole='P-S1'):
        return {
            'Work Request #': '91467680',
            'Snapshot Date': '2026-04-19',
            'CU': cu,
            'Quantity': qty,
            'Pole #': pole,
            'Work Type': 'Install',
            'Dept #': '800',
            'Units Total Price': price,
            'Units Completed?': True,
            '__variant': 'aep_billable',
            'Foreman': 'SubForeman',
            '__effective_user': 'SubForeman',
        }

    # ── Test 1 ────────────────────────────────────────────────────
    def test_primary_hash_byte_identical_across_sub_fingerprint_mutations(self):
        """primary variant hash MUST NOT change when
        ``_SUBCONTRACTOR_RATES_FINGERPRINT`` is mutated — the
        new fingerprint is intentionally NOT mixed into primary
        groups so the existing primary Excel files stay byte-
        identical under Phase 1.
        """
        rows = [
            self._primary_row(cu='CU-A', qty=1, price='$100.00',
                              pole='P-1'),
            self._primary_row(cu='CU-B', qty=2, price='$200.00',
                              pole='P-2'),
            self._primary_row(cu='CU-C', qty=3, price='$300.00',
                              pole='P-3'),
        ]
        generate_weekly_pdfs._SUBCONTRACTOR_RATES_FINGERPRINT = 'A'
        h1 = generate_weekly_pdfs.calculate_data_hash(rows)
        generate_weekly_pdfs._SUBCONTRACTOR_RATES_FINGERPRINT = 'B'
        h2 = generate_weekly_pdfs.calculate_data_hash(rows)
        self.assertEqual(
            h1, h2,
            "primary variant hash changed when "
            "_SUBCONTRACTOR_RATES_FINGERPRINT mutated — regression: "
            "the fingerprint mix-in MUST be variant-gated to the "
            "four new subcontractor variants only. Widening it to "
            "primary breaks ROADMAP success criterion 5 (byte-"
            "identical existing variants)."
        )

    # ── Test 2 ────────────────────────────────────────────────────
    def test_helper_hash_byte_identical_across_sub_fingerprint_mutations(self):
        """helper variant hash MUST NOT change when
        ``_SUBCONTRACTOR_RATES_FINGERPRINT`` is mutated."""
        rows = [
            self._helper_row(cu='CU-H1', qty=1, price='$200.00',
                             pole='P-H1', helper='BobHelper'),
            self._helper_row(cu='CU-H2', qty=2, price='$400.00',
                             pole='P-H2', helper='BobHelper'),
        ]
        generate_weekly_pdfs._SUBCONTRACTOR_RATES_FINGERPRINT = 'A'
        h1 = generate_weekly_pdfs.calculate_data_hash(rows)
        generate_weekly_pdfs._SUBCONTRACTOR_RATES_FINGERPRINT = 'B'
        h2 = generate_weekly_pdfs.calculate_data_hash(rows)
        self.assertEqual(
            h1, h2,
            "helper variant hash changed when "
            "_SUBCONTRACTOR_RATES_FINGERPRINT mutated — regression: "
            "see Test 1 docstring."
        )

    # ── Test 3 ────────────────────────────────────────────────────
    def test_vac_crew_hash_byte_identical_across_sub_fingerprint_mutations(self):
        """vac_crew variant hash MUST NOT change when
        ``_SUBCONTRACTOR_RATES_FINGERPRINT`` is mutated. Per
        CLAUDE.md 2026-04-22 00:00, VAC crew per-row metadata
        (__vac_crew_name/dept/job) is captured at the row level,
        so this test exercises a multi-member group to confirm
        the variant gate operates on the meta_parts block and not
        on the per-row loop.
        """
        rows = [
            self._vac_crew_row(cu='CU-V1', qty=1, price='$300.00',
                               pole='P-V1', name='VacMember1'),
            self._vac_crew_row(cu='CU-V2', qty=2, price='$600.00',
                               pole='P-V2', name='VacMember2'),
            self._vac_crew_row(cu='CU-V3', qty=3, price='$900.00',
                               pole='P-V3', name='VacMember3'),
        ]
        generate_weekly_pdfs._SUBCONTRACTOR_RATES_FINGERPRINT = 'A'
        h1 = generate_weekly_pdfs.calculate_data_hash(rows)
        generate_weekly_pdfs._SUBCONTRACTOR_RATES_FINGERPRINT = 'B'
        h2 = generate_weekly_pdfs.calculate_data_hash(rows)
        self.assertEqual(
            h1, h2,
            "vac_crew variant hash changed when "
            "_SUBCONTRACTOR_RATES_FINGERPRINT mutated — regression: "
            "see Test 1 docstring."
        )

    # ── Test 4 ────────────────────────────────────────────────────
    def test_aep_billable_hash_changes_on_sub_fingerprint_mutation(self):
        """Positive regression — aep_billable variant hash MUST
        change when ``_SUBCONTRACTOR_RATES_FINGERPRINT`` mutates,
        so a CSV edit forces regeneration of the new-variant Excel
        files (D-20 — fingerprint is mixed in for the variants
        that actually consume the subcontractor rates CSV).

        Symmetry note: the production code mixes the fingerprint
        in for all four new variants
        (aep_billable / reduced_sub / aep_billable_helper /
        reduced_sub_helper). Testing one positive case is
        sufficient because the variant-gate is a single ``in
        (...)`` check; the other three variants share the same
        branch and would regress together.
        """
        rows = [
            self._aep_billable_row(cu='CU-S1', qty=1, price='$50.00',
                                   pole='P-S1'),
            self._aep_billable_row(cu='CU-S2', qty=2, price='$100.00',
                                   pole='P-S2'),
        ]
        generate_weekly_pdfs._SUBCONTRACTOR_RATES_FINGERPRINT = 'A'
        h1 = generate_weekly_pdfs.calculate_data_hash(rows)
        generate_weekly_pdfs._SUBCONTRACTOR_RATES_FINGERPRINT = 'B'
        h2 = generate_weekly_pdfs.calculate_data_hash(rows)
        self.assertNotEqual(
            h1, h2,
            "aep_billable variant hash did NOT change when "
            "_SUBCONTRACTOR_RATES_FINGERPRINT mutated — regression: "
            "the fingerprint mix-in MUST fire for the four new "
            "variants. Dropping it silently means CSV edits to "
            "data/subcontractor_rates.csv would NOT force "
            "regeneration of the _AEPBillable / _ReducedSub files."
        )

    # ── Test 5 ────────────────────────────────────────────────────
    def test_existing_variants_total_hash_count_unchanged(self):
        """Cardinality check: a deterministic set of 10 synthetic
        groups across the 3 existing variants
        (primary / helper / vac_crew) produces 10 DISTINCT hashes
        — i.e. no two synthetic groups collapsed into the same
        hash, which would indicate a variant-merge regression
        (e.g. variant tokens dropped from meta_parts so a primary
        group and a helper group with otherwise-identical content
        share a hash).

        ROADMAP success criterion 5 sub-invariant: the existing
        variant taxonomy (3 variants) remains discriminated by
        the hash. Phase 1 adds 4 new variants without affecting
        the 3 existing ones; this test guards against
        accidental fusion.
        """
        groups: list[list[dict]] = []
        # 4 primary groups, distinguished by Work Request # or CU
        # content (sufficient to drive 4 distinct hashes).
        for i in range(4):
            row = self._primary_row(
                cu=f'CU-P{i}',
                qty=i + 1,
                price=f'${100 * (i + 1)}.00',
                pole=f'P-P{i}',
            )
            # Vary WR so groups don't share Work Request #.
            row['Work Request #'] = f'9146{700 + i}'
            groups.append([row])
        # 3 helper groups (distinct helpers).
        for i, helper_name in enumerate(['BobHelper', 'CarolHelper',
                                         'DaveHelper']):
            row = self._helper_row(
                cu=f'CU-H{i}',
                qty=i + 1,
                price=f'${200 * (i + 1)}.00',
                pole=f'P-H{i}',
                helper=helper_name,
            )
            groups.append([row])
        # 3 vac_crew groups (distinct content).
        for i in range(3):
            row = self._vac_crew_row(
                cu=f'CU-V{i}',
                qty=i + 1,
                price=f'${300 * (i + 1)}.00',
                pole=f'P-V{i}',
                name=f'VacMember{i}',
            )
            groups.append([row])
        hashes = {
            generate_weekly_pdfs.calculate_data_hash(g)
            for g in groups
        }
        self.assertEqual(
            len(hashes), len(groups),
            f"Expected {len(groups)} distinct hashes across 10 "
            f"synthetic groups in the 3 existing variants, got "
            f"{len(hashes)}. Variant taxonomy may have collapsed: "
            f"groups={len(groups)} unique_hashes={len(hashes)}. "
            f"Hashes: {sorted(hashes)}"
        )


class TestHelperShadowVariantFileIdentifier(unittest.TestCase):
    """Phase 01 gap closure (REVIEW-CR-01): the main-loop variant
    branch, valid_wr_weeks cleanup-tuple builder, and current_keys
    hash-history-prune key must all derive ``file_identifier`` from
    ``__helper_foreman`` for the two helper-shadow variants
    (aep_billable_helper, reduced_sub_helper) so the round-trip
    with build_group_identity succeeds and
    _has_existing_week_attachment correctly identifies prior
    helper-shadow attachments.

    These tests cover the contract via a stand-in that mirrors the
    main-loop derivation body. Source-level grep tests below guard
    against the "tests pass but production reverted" failure mode
    (same defense pattern as TestExcludeWrsMatchesAllVariants /
    TestWrFilterMatchesAllVariants from Plan 01-07).
    """

    @staticmethod
    def _derive_main_loop_identifier(variant: str, first_row: dict):
        """Mirror of the main-loop variant branch (Site 1 in CR-01).

        Returns ``(identifier, file_identifier)`` -- the two outputs
        the production code computes per group. Must stay in sync
        with the production cascade in ``main()`` immediately
        above ``history_key = f"{wr_num}|...":``.
        """
        if variant in ('helper', 'aep_billable_helper', 'reduced_sub_helper'):
            helper_foreman = first_row.get('__helper_foreman', '')
            helper_dept = first_row.get('__helper_dept', '')
            helper_job = first_row.get('__helper_job', '')
            identifier = f"{helper_foreman}|{helper_dept}|{helper_job}"
            file_identifier = (
                generate_weekly_pdfs._RE_SANITIZE_HELPER_NAME.sub(
                    '_', helper_foreman
                )[:50] if helper_foreman else ''
            )
        elif variant == 'vac_crew':
            identifier = ''
            file_identifier = ''
        else:
            user_val = first_row.get('User')
            identifier = (
                generate_weekly_pdfs._RE_SANITIZE_IDENTIFIER.sub(
                    '_', user_val
                )[:50] if user_val else ''
            )
            file_identifier = identifier
        return identifier, file_identifier

    def test_aep_billable_helper_derives_from_helper_foreman(self):
        identifier, file_identifier = self._derive_main_loop_identifier(
            'aep_billable_helper',
            {
                '__helper_foreman': 'Jane Smith',
                '__helper_dept': '500',
                '__helper_job': 'JOB-99',
            },
        )
        self.assertEqual(identifier, 'Jane Smith|500|JOB-99')
        self.assertEqual(file_identifier, 'Jane_Smith')

    def test_reduced_sub_helper_derives_from_helper_foreman(self):
        identifier, file_identifier = self._derive_main_loop_identifier(
            'reduced_sub_helper',
            {
                '__helper_foreman': 'John Doe',
                '__helper_dept': '600',
                '__helper_job': 'JOB-12',
            },
        )
        self.assertEqual(identifier, 'John Doe|600|JOB-12')
        self.assertEqual(file_identifier, 'John_Doe')

    def test_aep_billable_non_helper_falls_through_to_empty_user(self):
        # No User field on shadow row -> '' identifier matches what
        # build_group_identity produces for ``_AEPBillable_<hash>.xlsx``
        # (identifier=''). Round-trip succeeds even though we take
        # the ``else`` branch.
        identifier, file_identifier = self._derive_main_loop_identifier(
            'aep_billable',
            {'__helper_foreman': '', 'User': ''},
        )
        self.assertEqual(identifier, '')
        self.assertEqual(file_identifier, '')

    def test_reduced_sub_non_helper_falls_through_to_empty_user(self):
        identifier, file_identifier = self._derive_main_loop_identifier(
            'reduced_sub',
            {'__helper_foreman': '', 'User': ''},
        )
        self.assertEqual(identifier, '')
        self.assertEqual(file_identifier, '')

    def test_aep_billable_helper_filename_round_trips(self):
        # The CR-01 bug shape: pre-fix the main loop produced
        # file_identifier='' but build_group_identity parses
        # ``_AEPBillable_Helper_Jane_Smith_<hash>.xlsx`` as
        # (wr, week, 'aep_billable_helper', 'Jane_Smith') -
        # a comparison
        # (parsed_ident or '') == (file_identifier or '')
        # always failed. With the fix, both sides produce
        # 'Jane_Smith' and the comparison succeeds.
        _, file_identifier = self._derive_main_loop_identifier(
            'aep_billable_helper',
            {
                '__helper_foreman': 'Jane Smith',
                '__helper_dept': '500',
                '__helper_job': 'JOB-99',
            },
        )
        ident = generate_weekly_pdfs.build_group_identity(
            'WR_91467680_WeekEnding_041926_123456_AEPBillable_Helper_Jane_Smith_ab12cd34ef.xlsx'
        )
        self.assertIsNotNone(ident)
        wr, week, parsed_variant, parsed_identifier = ident
        self.assertEqual(parsed_variant, 'aep_billable_helper')
        self.assertEqual(parsed_identifier, file_identifier)

    def test_reduced_sub_helper_filename_round_trips(self):
        _, file_identifier = self._derive_main_loop_identifier(
            'reduced_sub_helper',
            {
                '__helper_foreman': 'John Doe',
                '__helper_dept': '600',
                '__helper_job': 'JOB-12',
            },
        )
        ident = generate_weekly_pdfs.build_group_identity(
            'WR_91467680_WeekEnding_041926_123456_ReducedSub_Helper_John_Doe_ab12cd34ef.xlsx'
        )
        self.assertIsNotNone(ident)
        wr, week, parsed_variant, parsed_identifier = ident
        self.assertEqual(parsed_variant, 'reduced_sub_helper')
        self.assertEqual(parsed_identifier, file_identifier)

    def test_helper_shadow_with_empty_helper_foreman_returns_empty_file_id(self):
        # Defensive: upstream gate at _valid_helper_row in
        # group_source_rows prevents empty __helper_foreman from
        # reaching this code path in production, but the
        # derivation itself must not crash if the gate is ever
        # bypassed. WR-03 (plan 01-10) adds a defensive raise in
        # generate_excel's filename-suffix branch -- the
        # IDENTIFIER derivation here remains permissive.
        identifier, file_identifier = self._derive_main_loop_identifier(
            'aep_billable_helper',
            {'__helper_foreman': '', '__helper_dept': '500', '__helper_job': 'JOB-99'},
        )
        self.assertEqual(identifier, '|500|JOB-99')
        self.assertEqual(file_identifier, '')

    def test_production_main_loop_carries_shadow_variant_gate(self):
        # Source-level guard: confirm the production main loop
        # (Site 1) carries the three-variant gate. Defeats the
        # "test mirror passes but production reverted" failure mode.
        import inspect
        import pathlib
        import pipeline.orchestrate  # W6: main-loop Site 1 gate lives in main()
        src = (
            pathlib.Path(
                inspect.getsourcefile(generate_weekly_pdfs)
            ).read_text(encoding='utf-8')
            + "\n"
            + pathlib.Path(
                inspect.getsourcefile(pipeline.orchestrate)
            ).read_text(encoding='utf-8')
        )
        # Either single- or double-quote tuple syntax acceptable.
        gate_present = (
            "variant in ('helper', 'aep_billable_helper', 'reduced_sub_helper')"
            in src
            or 'variant in ("helper", "aep_billable_helper", "reduced_sub_helper")'
            in src
        )
        self.assertTrue(
            gate_present,
            "Main-loop variant branch must gate on the three-tuple "
            "'(helper, aep_billable_helper, reduced_sub_helper)' for "
            "CR-01 to be closed. See 01-08-PLAN.md.",
        )

    def test_production_valid_wr_weeks_and_current_keys_carry_shadow_variant_gate(self):
        # Source-level guard: Sites 2 and 3 must also carry
        # shadow-variant gates. We verify by counting occurrences
        # of the three-tuple gate; with Site 1's gate, expect
        # >= 3 total (Sites 1, 2, 3). The Site 3 cascade uses
        # ``_variant`` (underscore prefix) so the gate text is
        # slightly different -- count BOTH forms.
        import inspect
        import pathlib
        import pipeline.orchestrate  # W6: Sites 1/2/3 gates live in main()
        src = (
            pathlib.Path(
                inspect.getsourcefile(generate_weekly_pdfs)
            ).read_text(encoding='utf-8')
            + "\n"
            + pathlib.Path(
                inspect.getsourcefile(pipeline.orchestrate)
            ).read_text(encoding='utf-8')
        )
        count_v1 = (
            src.count("variant in ('helper', 'aep_billable_helper', 'reduced_sub_helper')")
            + src.count('variant in ("helper", "aep_billable_helper", "reduced_sub_helper")')
        )
        count_v2 = (
            src.count("_variant in ('helper', 'aep_billable_helper', 'reduced_sub_helper')")
            + src.count('_variant in ("helper", "aep_billable_helper", "reduced_sub_helper")')
        )
        # Sites 1 and 2 use ``variant``; Site 3 uses ``_variant``.
        # Total occurrences across both forms must be >= 3.
        # Note: ``variant in (...)`` count is >= 2 since it appears
        # in both Site 1 and Site 2 (and also in calculate_data_hash
        # which is unrelated but uses the same gate text).
        self.assertGreaterEqual(
            count_v1 + count_v2, 3,
            f"Expected the three-tuple gate to appear at Sites 1, 2, "
            f"and 3 (any combination of `variant` / `_variant`). Found "
            f"{count_v1} `variant in (...)` and {count_v2} `_variant in (...)`."
        )


class TestAepBillableCutoffEnvVarOverride(unittest.TestCase):
    """Phase 01 gap closure (REVIEW-IN-01): ``_AEP_BILLABLE_CUTOFF``
    is overridable via ``AEP_BILLABLE_CUTOFF`` env var with safe
    parse + fallback. Operator-facing workflow knob for contract
    re-negotiation or retroactive billing decisions.

    Reload pattern mirrors ``TestSubcontractorPppSheetIdEmptyStringDisable``
    — Sentry-DSN suppression + try/finally env restoration per
    Living Ledger 2026-04-22 16:05.
    """

    def setUp(self):
        self._env_snapshot = {
            k: os.environ.get(k)
            for k in ('AEP_BILLABLE_CUTOFF', 'SENTRY_DSN')
        }
        os.environ['SENTRY_DSN'] = ''

    def tearDown(self):
        for k, v in self._env_snapshot.items():
            if v is None:
                os.environ.pop(k, None)
            else:
                os.environ[k] = v
        with mock.patch.dict(os.environ, {"SENTRY_DSN": ""}, clear=False):
            with mock.patch('sentry_sdk.init'):
                importlib.reload(generate_weekly_pdfs)

    def _reload_with_cutoff(self, value):
        if value is None:
            os.environ.pop('AEP_BILLABLE_CUTOFF', None)
        else:
            os.environ['AEP_BILLABLE_CUTOFF'] = value
        with mock.patch.dict(os.environ, {"SENTRY_DSN": ""}, clear=False):
            with mock.patch('sentry_sdk.init'):
                importlib.reload(generate_weekly_pdfs)

    def test_env_unset_uses_hardcoded_default(self):
        import datetime
        self._reload_with_cutoff(None)
        self.assertEqual(
            generate_weekly_pdfs._AEP_BILLABLE_CUTOFF,
            datetime.date(2026, 4, 12),
            "Default cutoff must be byte-identical to the pre-fix "
            "constant (contract award date)."
        )

    def test_env_valid_forward_override(self):
        import datetime
        self._reload_with_cutoff('2026-05-01')
        self.assertEqual(
            generate_weekly_pdfs._AEP_BILLABLE_CUTOFF,
            datetime.date(2026, 5, 1),
        )

    def test_env_valid_backward_override(self):
        # Retroactive billing decision: operator moves cutoff
        # backward to capture earlier rows.
        import datetime
        self._reload_with_cutoff('2025-12-01')
        self.assertEqual(
            generate_weekly_pdfs._AEP_BILLABLE_CUTOFF,
            datetime.date(2025, 12, 1),
        )

    def test_env_invalid_format_falls_back_to_default(self):
        import datetime
        self._reload_with_cutoff('not-a-date')
        self.assertEqual(
            generate_weekly_pdfs._AEP_BILLABLE_CUTOFF,
            datetime.date(2026, 4, 12),
            "Invalid format must fail-safe to default, not crash "
            "import or silently disable _AEPBillable variant."
        )

    def test_env_empty_string_treated_as_unset(self):
        # Empty string is functionally "no override" — falls to
        # the default. (Different from SUBCONTRACTOR_PPP_SHEET_ID
        # in WR-02 where empty string is "disable"; cutoff has no
        # "disabled" state, so empty string means "use default.")
        import datetime
        self._reload_with_cutoff('')
        self.assertEqual(
            generate_weekly_pdfs._AEP_BILLABLE_CUTOFF,
            datetime.date(2026, 4, 12),
        )

    def test_production_source_carries_env_var_lookup(self):
        # Source-level guard: the env-var read landed.
        import inspect
        import pathlib
        src = pathlib.Path(
            inspect.getsourcefile(generate_weekly_pdfs)
        ).read_text(encoding='utf-8')
        self.assertIn("os.getenv('AEP_BILLABLE_CUTOFF'", src)
        self.assertIn('Invalid AEP_BILLABLE_CUTOFF format', src)


class TestResolveRowPriceQuantityCoercion(unittest.TestCase):
    """Phase 01 gap closure (REVIEW-IN-02): qty_raw coercion is explicit.
    Numeric outcome is byte-identical for every pre-existing input case;
    only readability improves. The float-zero passthrough case is the
    IN-02 regression guard against re-introducing the ``or 0`` collapse.

    Test design:
      - ``_resolve_row_price(row, variant='reduced_sub', missing_cus=Counter())``
        is the entrypoint (signature confirmed via grep of source).
      - ``_SUBCONTRACTOR_RATES`` is monkey-patched so ``rate=100.0/unit``
        for CU=ABC123 work_type=install → rate*qty path returns 100.0*qty.
      - ``Units Total Price=999.0`` is a SAFETY-FLOOR CANARY: if the
        function falls through (rate<=0 or qty<=0), the safety-floor
        path returns ``parse_price(row['Units Total Price'])`` = 999.0,
        making fallthrough observable.

      Outcomes by qty input:
        qty>0 → 100.0 * qty   (rate*qty path)
        qty<=0 / fallthrough → 999.0  (safety-floor path)
    """

    RATES_STUB = {
        'ABC123': {
            'reduced_install_price': 100.0,
            'reduced_remove_price': 0.0,
            'reduced_transfer_price': 0.0,
            'new_install_price': 100.0,
            'new_remove_price': 0.0,
            'new_transfer_price': 0.0,
        }
    }

    def _row(self, qty_value, *, include_qty_key=True, units_total_price=999.0):
        """Build a row that, if qty is correctly coerced AND > 0,
        prices via rate*qty = 100.0 * qty. ``units_total_price=999.0``
        is the safety-floor canary: if the function falls through to
        ``parse_price(row['Units Total Price'])`` (rate<=0 or qty<=0),
        we observe 999.0 and know the rate*qty path was NOT taken.
        """
        r = {
            'CU': 'ABC123',
            'Work Type': 'install',
            'Units Total Price': units_total_price,
        }
        if include_qty_key:
            r['Quantity'] = qty_value
        return r

    def _resolve(self, row):
        import collections
        return generate_weekly_pdfs._resolve_row_price(
            row, variant='reduced_sub', missing_cus=collections.Counter(),
        )

    def test_quantity_none_falls_through_to_units_total_price(self):
        # qty=None → qty=0.0 → rate*qty=0.0 ≤ 0 → safety floor fires →
        # returns parse_price(Units Total Price) = 999.0.
        with mock.patch.object(
            generate_weekly_pdfs, '_SUBCONTRACTOR_RATES', self.RATES_STUB,
        ):
            self.assertEqual(self._resolve(self._row(None)), 999.0)

    def test_quantity_empty_string_falls_through_to_units_total_price(self):
        with mock.patch.object(
            generate_weekly_pdfs, '_SUBCONTRACTOR_RATES', self.RATES_STUB,
        ):
            self.assertEqual(self._resolve(self._row('')), 999.0)

    def test_quantity_int_zero_falls_through_to_units_total_price(self):
        with mock.patch.object(
            generate_weekly_pdfs, '_SUBCONTRACTOR_RATES', self.RATES_STUB,
        ):
            self.assertEqual(self._resolve(self._row(0)), 999.0)

    def test_quantity_float_zero_passthrough_regression_guard(self):
        # THE IN-02 regression guard: legitimate Quantity=0.0 must not
        # produce a different price than int(0). Both must take the
        # safety-floor path because qty<=0.
        with mock.patch.object(
            generate_weekly_pdfs, '_SUBCONTRACTOR_RATES', self.RATES_STUB,
        ):
            self.assertEqual(self._resolve(self._row(0.0)), 999.0)

    def test_quantity_int_one_uses_rate_times_qty(self):
        with mock.patch.object(
            generate_weekly_pdfs, '_SUBCONTRACTOR_RATES', self.RATES_STUB,
        ):
            self.assertEqual(self._resolve(self._row(1)), 100.0)

    def test_quantity_float_uses_rate_times_qty(self):
        with mock.patch.object(
            generate_weekly_pdfs, '_SUBCONTRACTOR_RATES', self.RATES_STUB,
        ):
            self.assertEqual(self._resolve(self._row(1.5)), 150.0)

    def test_quantity_string_float_parses(self):
        with mock.patch.object(
            generate_weekly_pdfs, '_SUBCONTRACTOR_RATES', self.RATES_STUB,
        ):
            self.assertEqual(self._resolve(self._row('1.5')), 150.0)

    def test_quantity_invalid_string_falls_through_to_units_total_price(self):
        # Invalid string → float() raises → except clause → qty=0.0 →
        # rate*qty=0 ≤ 0 → safety floor → Units Total Price.
        with mock.patch.object(
            generate_weekly_pdfs, '_SUBCONTRACTOR_RATES', self.RATES_STUB,
        ):
            self.assertEqual(self._resolve(self._row('invalid')), 999.0)

    def test_missing_quantity_key_falls_through_to_units_total_price(self):
        # row.get('Quantity', 0) with no key → 0 → 0.0 → safety floor.
        with mock.patch.object(
            generate_weekly_pdfs, '_SUBCONTRACTOR_RATES', self.RATES_STUB,
        ):
            self.assertEqual(
                self._resolve(self._row(None, include_qty_key=False)),
                999.0,
            )

    def test_production_source_does_not_carry_or_zero_pattern(self):
        # Source-level guard: the ``or 0`` short-circuit is removed.
        # Phase 09 W2: ``_resolve_row_price`` was relocated to
        # ``pipeline/pricing.py``; inspect the function object's source
        # (which follows the facade re-export to the new module) instead
        # of the facade file so the guard tracks the code's real home.
        import inspect
        src = inspect.getsource(generate_weekly_pdfs._resolve_row_price)
        self.assertNotIn(
            "row.get('Quantity') or 0",
            src,
            "IN-02 regression: the ``or 0`` short-circuit pattern has "
            "been re-introduced. Use explicit ``row.get('Quantity', 0)`` "
            "+ ``if qty_raw not in (None, '')`` per 01-11-PLAN.md.",
        )
        # Confirm the explicit pattern is present.
        self.assertIn("qty_raw not in (None, '')", src)


class TestPhase1FilenameRoundTripCoverage(unittest.TestCase):
    """Covers the filename round-trip for all 7 Phase 1 variants.

    Split out from ``TestPhase1IntegrationRegression`` per round-3
    checker Warning 11 because this test exercises the
    parser/generator symmetry, not hash stability — different
    invariant, different class. ``build_group_identity`` is the
    only function exercised here; ``calculate_data_hash`` is not
    touched by these tests.

    The parser ↔ generator symmetry matters for:
      * Cleanup of stale variant Excel files
        (``cleanup_stale_excels`` calls ``build_group_identity``
        to identify which files share an identity with the
        kept set).
      * Hash-history bookkeeping (``history_key`` shape).
      * Target-row matching during upload.
    Any variant whose generated filename cannot be parsed back
    would silently regenerate every run and accumulate orphan
    attachments.

    Phase 01 Plan 06 Task 2 (split per Warning 11).
    """

    # Filename templates parallel the production generator. Each
    # tuple holds (variant_token, expected_variant_string,
    # expected_identifier). The ``variant_token`` segment is what
    # ``generate_excel`` appends after the timestamp; the test
    # constructs a realistic filename around it and asserts the
    # parser recovers the variant.
    _VARIANT_CASES = [
        # variant, filename_tail_after_timestamp, expected_identifier
        ('primary',             '',                                  None),
        ('helper',              'Helper_Jane_Smith',                 'Jane_Smith'),
        ('vac_crew',            'VacCrew',                           ''),
        ('aep_billable',        'AEPBillable',                       ''),
        ('reduced_sub',         'ReducedSub',                        ''),
        ('aep_billable_helper', 'AEPBillable_Helper_Jane_Smith',     'Jane_Smith'),
        ('reduced_sub_helper',  'ReducedSub_Helper_Jane_Smith',      'Jane_Smith'),
    ]

    def test_filename_round_trip_for_all_seven_variants(self):
        """For each of the 7 Phase 1 variants, build a realistic
        filename that mirrors ``generate_excel``'s output shape and
        confirm that ``build_group_identity`` recovers the exact
        ``variant`` string.

        The filename shape used:
          ``WR_91467680_WeekEnding_041926_123456[_<tail>]_<hash>.xlsx``

        Where ``<tail>`` is the variant marker (+ helper name for
        helper variants) and ``<hash>`` is a 16-char hex token —
        the same length the production ``calculate_data_hash``
        emits. We parametrize via subTest so a failure in one
        variant doesn't mask the others.
        """
        wr = '91467680'
        week = '041926'
        timestamp = '123456'
        data_hash = 'deadbeefcafebabe'  # 16-char hex, parser-stable
        for (variant, tail, expected_id) in self._VARIANT_CASES:
            with self.subTest(variant=variant):
                if tail:
                    fname = (
                        f'WR_{wr}_WeekEnding_{week}_{timestamp}'
                        f'_{tail}_{data_hash}.xlsx'
                    )
                else:
                    fname = (
                        f'WR_{wr}_WeekEnding_{week}_{timestamp}'
                        f'_{data_hash}.xlsx'
                    )
                result = generate_weekly_pdfs.build_group_identity(
                    fname
                )
                self.assertIsNotNone(
                    result,
                    f"build_group_identity returned None for "
                    f"variant={variant!r} filename={fname!r} — "
                    f"parser ↔ generator symmetry is broken; "
                    f"downstream cleanup_stale_excels would not "
                    f"identify this file and orphan attachments "
                    f"would accumulate."
                )
                parsed_wr, parsed_week, parsed_variant, parsed_id = (
                    result
                )
                self.assertEqual(
                    parsed_variant, variant,
                    f"variant mismatch for filename={fname!r}: "
                    f"expected {variant!r}, got "
                    f"{parsed_variant!r}. Round-trip failure "
                    f"breaks stale-file cleanup and target-row "
                    f"matching."
                )
                self.assertEqual(parsed_wr, wr)
                self.assertEqual(parsed_week, week)
                self.assertEqual(parsed_id, expected_id)


class TestPhase1GapClosureLedgerEntryPresent(unittest.TestCase):
    """Phase 01 gap closure (Living Ledger / autonomous-cloud-memory):
    a memory-bank/living-ledger.md Living Ledger entry timestamped 2026-05-15 MUST
    document the 13-finding gap-closure round and the 7 new rules
    encoded therein. Per CLAUDE.md's "AUTONOMOUS CLOUD MEMORY
    INJECTION (CRITICAL)" rule, any architectural standard /
    recurring fix / new operational rule introduced by a feature
    PR must be appended to the Living Ledger in the same PR.

    These source-level guards catch silent reversion of the
    documentation by a future PR.
    """

    @staticmethod
    def _read_ledger() -> str:
        # Living Ledger was relocated from CLAUDE.md to
        # memory-bank/living-ledger.md on 2026-05-28.
        import pathlib
        repo_root = pathlib.Path(
            generate_weekly_pdfs.__file__,
        ).parent
        return (repo_root / 'memory-bank' / 'living-ledger.md').read_text(encoding='utf-8')

    def test_timestamp_present(self):
        ledger = self._read_ledger()
        self.assertIn('[2026-05-15', ledger)

    def test_round_summary_phrase_present(self):
        ledger = self._read_ledger()
        self.assertIn(
            'Phase 01 (Subcontractor Rate Logic Modification) gap-closure',
            ledger,
        )

    def test_all_seven_new_rules_named(self):
        ledger = self._read_ledger()
        expected_rules = (
            'Three-site identity-consistency invariant',
            'Mirror-matcher invariant',
            'Explicit PII markers',
            'Defensive raise scope discipline',
            'Dual-target cleanup invocation pattern',
            'Env-var override safe-parse pattern',
            'Workflow pinning for new feature env vars',
        )
        for rule in expected_rules:
            with self.subTest(rule=rule):
                self.assertIn(
                    rule, ledger,
                    f"Living Ledger 2026-05-15 entry must name "
                    f"the rule {rule!r}. See plan 01-14.",
                )

    def test_entry_appears_after_2026_04_25_freeze_row_entry(self):
        ledger = self._read_ledger()
        prev_idx = ledger.find('[2026-04-25 14:00]')
        new_idx = ledger.find('[2026-05-15')
        self.assertGreater(prev_idx, 0, "Prior 2026-04-25 14:00 entry not found.")
        self.assertGreater(new_idx, 0, "New 2026-05-15 entry not found.")
        self.assertLess(
            prev_idx, new_idx,
            "New 2026-05-15 entry must appear AFTER the 2026-04-25 "
            "14:00 freeze_row entry (Living Ledger entries are "
            "chronologically appended).",
        )

    def test_entry_references_regression_test_classes(self):
        # The ledger entry should name the regression-test classes
        # added by this round so future maintainers can trace
        # rule → test directly.
        ledger = self._read_ledger()
        for cls_name in (
            'TestHelperShadowVariantFileIdentifier',
            'TestExcludeWrsMatchesAllVariants',
            'TestWrFilterMatchesAllVariants',
            'TestPppAttachmentPrefetchBudget',
            'TestPppCleanupUntrackedAttachments',
        ):
            with self.subTest(cls_name=cls_name):
                self.assertIn(
                    cls_name, ledger,
                    f"Living Ledger 2026-05-15 entry must reference "
                    f"{cls_name} so future readers can trace rule "
                    f"→ test."
                )


class TestResolveRowPriceAbbreviatedWorkType(unittest.TestCase):
    """Production hotfix 2026-05-16 — P0 data-integrity bug.

    Smartsheet operators commonly enter Work Type as the abbreviated
    forms ``Inst`` / ``Rem`` / ``Trans``, not the canonical full
    forms ``Install`` / ``Removal`` / ``Transfer``. The pre-fix
    matcher used ``'install' in work_type_raw`` — a substring check
    that succeeds for ``'install'`` (full) but FAILS for ``'inst'``
    because the search string ``'install'`` (7 chars) is not contained
    in the shorter ``'inst'`` (4 chars). Same direction error for
    ``'remov'`` vs ``'rem'`` and ``'transfer'`` vs ``'trans'``.

    When the matcher fell through to the ``else`` branch, the helper
    returned ``parse_price(row.get('Units Total Price'))`` — the
    safety-floor SmartSheet pricing. **The fallback returned the same
    value for BOTH AEP and ReducedSub variants** so the generated
    workbooks were byte-identical (verified via SHA256 on the
    2026-05-16 23:23 UTC GHA artifact: 8 of 8 AEP+ReducedSub file
    pairs had matching content hashes).

    The fix aligns with the existing ``recalculate_row_price`` pattern
    at L1655 (``'rem' in work_type_raw``) — uses the shortest
    unambiguous substring so both abbreviated and full forms match.
    """

    @classmethod
    def setUpClass(cls):
        cls._orig_rates = dict(generate_weekly_pdfs._SUBCONTRACTOR_RATES)
        # Different new_* vs reduced_* values so the test distinguishes
        # which rate column was consulted.
        generate_weekly_pdfs._SUBCONTRACTOR_RATES.clear()
        generate_weekly_pdfs._SUBCONTRACTOR_RATES['ABC'] = {
            'cu_code': 'ABC',
            'cu_wbs': '999',
            'compatible_unit_group': 'TestGroup',
            'reduced_install_price': 10.00,
            'reduced_remove_price': 5.00,
            'reduced_transfer_price': 2.50,
            'new_install_price': 20.00,
            'new_remove_price': 12.00,
            'new_transfer_price': 6.00,
        }

    @classmethod
    def tearDownClass(cls):
        generate_weekly_pdfs._SUBCONTRACTOR_RATES.clear()
        generate_weekly_pdfs._SUBCONTRACTOR_RATES.update(cls._orig_rates)

    def _resolve(self, work_type, variant, qty=5, units_total='$999.99'):
        """``Units Total Price=$999.99`` is the safety-floor canary —
        if it leaks into the return, the matcher fell through."""
        from collections import Counter
        row = {
            'CU': 'ABC',
            'Work Type': work_type,
            'Quantity': qty,
            'Units Total Price': units_total,
        }
        return generate_weekly_pdfs._resolve_row_price(
            row, variant, Counter()
        )

    # ── Abbreviated forms — the production-realistic case ──────

    def test_inst_abbrev_aep_billable_uses_new_install_price(self):
        """``Work Type='Inst'`` + AEP → new_install_price × qty.

        Pre-fix: 'install' in 'inst' is False → safety floor → 999.99.
        Post-fix: 20.00 × 5 = 100.00.
        """
        self.assertEqual(self._resolve('Inst', 'aep_billable'), 100.00)

    def test_inst_abbrev_reduced_sub_uses_reduced_install_price(self):
        self.assertEqual(self._resolve('Inst', 'reduced_sub'), 50.00)

    def test_rem_abbrev_aep_billable_uses_new_remove_price(self):
        self.assertEqual(self._resolve('Rem', 'aep_billable'), 60.00)

    def test_rem_abbrev_reduced_sub_uses_reduced_remove_price(self):
        self.assertEqual(self._resolve('Rem', 'reduced_sub'), 25.00)

    def test_trans_abbrev_aep_billable_uses_new_transfer_price(self):
        self.assertEqual(self._resolve('Trans', 'aep_billable'), 30.00)

    def test_trans_abbrev_reduced_sub_uses_reduced_transfer_price(self):
        self.assertEqual(self._resolve('Trans', 'reduced_sub'), 12.50)

    def test_xfr_abbrev_treated_as_transfer(self):
        """``recalculate_row_price`` already handles 'xfr' as transfer
        (L1657). The hotfix preserves that synonym."""
        self.assertEqual(self._resolve('Xfr', 'aep_billable'), 30.00)

    # ── Full forms — regression guard, MUST still work ─────────

    def test_install_full_form_unchanged(self):
        self.assertEqual(self._resolve('Install', 'aep_billable'), 100.00)

    def test_removal_full_form_unchanged(self):
        self.assertEqual(self._resolve('Removal', 'aep_billable'), 60.00)

    def test_transfer_full_form_unchanged(self):
        self.assertEqual(self._resolve('Transfer', 'aep_billable'), 30.00)

    # ── AEP and ReducedSub MUST diverge on the same abbreviated row ──

    def test_aep_and_reduced_diverge_on_abbreviated_inst(self):
        """The hotfix's key invariant: with abbreviated Work Type,
        the two variants must produce DIFFERENT prices (not the
        same safety-floor fallback).
        """
        aep = self._resolve('Inst', 'aep_billable')
        red = self._resolve('Inst', 'reduced_sub')
        self.assertNotEqual(
            aep, red,
            'AEP and ReducedSub MUST diverge for abbreviated Work '
            'Types — same-value response is the production bug.',
        )
        self.assertEqual(aep, 100.00)
        self.assertEqual(red, 50.00)

    def test_aep_and_reduced_diverge_on_abbreviated_rem(self):
        aep = self._resolve('Rem', 'aep_billable')
        red = self._resolve('Rem', 'reduced_sub')
        self.assertNotEqual(aep, red)
        self.assertEqual(aep, 60.00)
        self.assertEqual(red, 25.00)

    # ── Truly unknown work types still fall through ─────────────

    def test_unknown_work_type_falls_through_to_smartsheet(self):
        """Unknown work types (not inst/rem/trans/xfr prefix) MUST
        still trigger the safety floor — the fix must not over-broaden.
        """
        self.assertEqual(self._resolve('Maintenance', 'aep_billable'), 999.99)
        self.assertEqual(self._resolve('', 'aep_billable'), 999.99)
        self.assertEqual(self._resolve('Unknown', 'reduced_sub'), 999.99)

    def test_helper_shadow_variants_also_diverge(self):
        """Both helper-shadow variants (Plan 02) MUST also pick the
        correct rate column with abbreviated Work Type.
        """
        aep_h = self._resolve('Inst', 'aep_billable_helper')
        red_h = self._resolve('Inst', 'reduced_sub_helper')
        self.assertEqual(aep_h, 100.00)
        self.assertEqual(red_h, 50.00)
        self.assertNotEqual(aep_h, red_h)


class TestCleanupVariantWhitelist(unittest.TestCase):
    """Phase 1.1 Bug B2 (D-07 / D-08 / SUB-10): per-sheet variant
    whitelist for ``cleanup_untracked_sheet_attachments``.

    The PPP cleanup invocation passes ``variant_whitelist={'reduced_sub',
    'reduced_sub_helper'}`` so any other variant (primary, helper,
    vac_crew, aep_billable, aep_billable_helper) parsed from a filename
    on PPP is unconditionally deleted at cleanup time — regardless of
    ``valid_wr_weeks`` membership and regardless of
    ``KEEP_HISTORICAL_WEEKS``. TARGET cleanup passes ``None`` (default)
    and preserves byte-identical legacy behavior.

    These tests drive the function in-process with mocked Smartsheet
    SDK calls so the whitelist-gating, off-contract delete loop, and
    counter / log surfaces are exercised end-to-end.
    """

    def _make_attachment(self, name, attachment_id):
        """Build a MagicMock attachment with ``name`` + ``id`` attrs."""
        att = mock.MagicMock()
        att.name = name
        att.id = attachment_id
        return att

    def _make_sheet_with_attachments(self, attachments):
        """Build a mock sheet with a single row carrying the supplied
        attachments. Returns ``(sheet, client)`` — client's
        ``Attachments.delete_attachment`` records every call.
        """
        row = mock.MagicMock()
        row.id = 99999
        sheet = mock.MagicMock()
        sheet.rows = [row]
        client = mock.MagicMock()
        # Inject attachments via the per-row API path
        list_resp = mock.MagicMock()
        list_resp.data = attachments
        client.Attachments.list_row_attachments.return_value = list_resp
        return sheet, client

    def setUp(self):
        # The function returns early when test_mode=True; turn off the
        # global TEST_MODE for these tests so the cleanup logic runs.
        self._saved_test_mode = generate_weekly_pdfs.TEST_MODE
        self._saved_keep_hist = generate_weekly_pdfs.KEEP_HISTORICAL_WEEKS

    def tearDown(self):
        generate_weekly_pdfs.TEST_MODE = self._saved_test_mode
        generate_weekly_pdfs.KEEP_HISTORICAL_WEEKS = self._saved_keep_hist

    # ── Test 1: signature carries the new kwarg with default None ──

    def test_signature_carries_variant_whitelist_kwarg_default_none(self):
        import inspect
        sig = inspect.signature(
            generate_weekly_pdfs.cleanup_untracked_sheet_attachments
        )
        self.assertIn(
            'variant_whitelist', sig.parameters,
            'cleanup_untracked_sheet_attachments must accept '
            'variant_whitelist kwarg per SUB-10.'
        )
        self.assertIs(
            sig.parameters['variant_whitelist'].default, None,
            'variant_whitelist default must be None to preserve '
            'byte-identical legacy TARGET behavior (D-09).'
        )

    # ── Test 2: legacy behavior preserved when kwarg omitted (D-09) ──

    def test_variant_whitelist_none_preserves_legacy_behavior(self):
        """variant_whitelist=None: every variant accumulates into
        identity_groups; off_contract_attachments stays empty; no
        unconditional deletes. Only the legacy newest-keep logic
        runs (and with KEEP_HISTORICAL_WEEKS=True + empty
        valid_wr_weeks, the identity-prune branch short-circuits)."""
        atts = [
            self._make_attachment(
                'WR_111_WeekEnding_041926_120000_abc123.xlsx', 1
            ),  # primary
            self._make_attachment(
                'WR_222_WeekEnding_041926_120000_ReducedSub_def456.xlsx', 2
            ),  # reduced_sub
            self._make_attachment(
                'WR_333_WeekEnding_041926_120000_AEPBillable_'
                'Helper_Jane_ghi789.xlsx', 3
            ),  # aep_billable_helper
        ]
        sheet, client = self._make_sheet_with_attachments(atts)
        generate_weekly_pdfs.KEEP_HISTORICAL_WEEKS = True
        generate_weekly_pdfs.cleanup_untracked_sheet_attachments(
            client,
            target_sheet_id=42,
            valid_wr_weeks=set(),  # nothing in valid set
            test_mode=False,
            target_sheet=sheet,
        )
        # With variant_whitelist=None, no off-contract path fires.
        # KEEP_HISTORICAL_WEEKS skips identity-prune for these idents.
        # Expected: zero delete_attachment calls.
        self.assertEqual(
            client.Attachments.delete_attachment.call_count, 0,
            'variant_whitelist=None + empty valid_wr_weeks + '
            'KEEP_HISTORICAL_WEEKS=True must produce zero deletes '
            '(byte-identical legacy behavior).'
        )

    # ── Test 3: whitelist routes off-contract attachments to delete ──

    def test_whitelist_deletes_off_contract_attachments(self):
        """variant_whitelist={'reduced_sub','reduced_sub_helper'} on
        a sheet carrying multiple variants: reduced_sub / reduced_sub_helper
        accumulate into identity_groups; primary / aep_billable / helper
        are unconditionally deleted via delete_attachment."""
        atts = [
            # IN whitelist:
            self._make_attachment(
                'WR_111_WeekEnding_041926_120000_ReducedSub_aaa.xlsx', 1
            ),  # reduced_sub
            self._make_attachment(
                'WR_222_WeekEnding_041926_120000_ReducedSub_'
                'Helper_Alice_bbb.xlsx', 2
            ),  # reduced_sub_helper
            # OFF-CONTRACT (must be deleted):
            self._make_attachment(
                'WR_333_WeekEnding_041926_120000_ccc.xlsx', 3
            ),  # primary (no variant marker in tail)
            self._make_attachment(
                'WR_444_WeekEnding_041926_120000_AEPBillable_ddd.xlsx', 4
            ),  # aep_billable
            self._make_attachment(
                'WR_555_WeekEnding_041926_120000_Helper_Bob_eee.xlsx', 5
            ),  # helper
        ]
        sheet, client = self._make_sheet_with_attachments(atts)
        PPP_ID = 9999
        generate_weekly_pdfs.KEEP_HISTORICAL_WEEKS = True
        generate_weekly_pdfs.cleanup_untracked_sheet_attachments(
            client,
            target_sheet_id=PPP_ID,
            valid_wr_weeks=set(),
            test_mode=False,
            target_sheet=sheet,
            variant_whitelist={'reduced_sub', 'reduced_sub_helper'},
        )
        # 3 off-contract attachments deleted (ids 3, 4, 5):
        delete_calls = client.Attachments.delete_attachment.call_args_list
        self.assertEqual(
            len(delete_calls), 3,
            'Expected 3 off-contract deletes (primary, aep_billable, '
            f'helper). Got {len(delete_calls)} calls: {delete_calls}'
        )
        deleted_ids = sorted(call[0][1] for call in delete_calls)
        self.assertEqual(
            deleted_ids, [3, 4, 5],
            f'Expected deletes for ids [3, 4, 5]; got {deleted_ids}'
        )
        # All deletes targeted the supplied target_sheet_id (PPP_ID):
        for call in delete_calls:
            self.assertEqual(
                call[0][0], PPP_ID,
                'Off-contract delete must target the supplied '
                f'target_sheet_id (got {call[0][0]} vs {PPP_ID})'
            )

    # ── Test 4: whitelist short-circuits BEFORE KEEP_HISTORICAL_WEEKS ──

    def test_whitelist_fires_before_keep_historical_weeks(self):
        """Off-contract attachments are deleted regardless of
        valid_wr_weeks state — KEEP_HISTORICAL_WEEKS cannot legitimize
        an off-contract attachment because the whitelist check fires
        BEFORE the identity_groups accumulation."""
        # Off-contract primary on a WR that IS in valid_wr_weeks
        # (so the identity-prune path would normally preserve it).
        atts = [
            self._make_attachment(
                'WR_91467680_WeekEnding_041926_120000_abc.xlsx', 100
            ),  # primary (off-contract for PPP whitelist)
        ]
        sheet, client = self._make_sheet_with_attachments(atts)
        # Place the WR's primary tuple in valid_wr_weeks AND turn on
        # KEEP_HISTORICAL_WEEKS — these would normally save the
        # attachment under the legacy identity-prune logic.
        generate_weekly_pdfs.KEEP_HISTORICAL_WEEKS = True
        valid = {('91467680', '041926', 'primary', '')}
        generate_weekly_pdfs.cleanup_untracked_sheet_attachments(
            client,
            target_sheet_id=9999,
            valid_wr_weeks=valid,
            test_mode=False,
            target_sheet=sheet,
            variant_whitelist={'reduced_sub', 'reduced_sub_helper'},
        )
        # The off-contract primary MUST still be deleted — variant
        # set membership is the authoritative gate for this sheet.
        self.assertEqual(
            client.Attachments.delete_attachment.call_count, 1,
            'Off-contract attachment must be deleted even when its '
            'identity tuple is in valid_wr_weeks and '
            'KEEP_HISTORICAL_WEEKS=True. Variant-set membership is '
            'the authoritative gate.'
        )

    # ── Test 5: non-parsing filenames do not enter off_contract bucket ──

    def test_non_parsing_filename_does_not_route_to_off_contract(self):
        """Files that don't parse via build_group_identity (returns
        None) must be ignored entirely — they don't enter
        off_contract_attachments and must NOT be deleted."""
        atts = [
            self._make_attachment('some_other_file.txt', 7),
            self._make_attachment('WR_garbage.xlsx', 8),  # parses to None
            self._make_attachment('not_a_wr_file.pdf', 9),
        ]
        sheet, client = self._make_sheet_with_attachments(atts)
        generate_weekly_pdfs.cleanup_untracked_sheet_attachments(
            client,
            target_sheet_id=9999,
            valid_wr_weeks=set(),
            test_mode=False,
            target_sheet=sheet,
            variant_whitelist={'reduced_sub', 'reduced_sub_helper'},
        )
        self.assertEqual(
            client.Attachments.delete_attachment.call_count, 0,
            'Non-parsing filenames must NOT be deleted — the existing '
            'if ident: guard already filters them.'
        )

    # ── Test 6: delete_attachment exception → WARNING + continue ──

    def test_delete_attachment_exception_is_caught(self):
        """When delete_attachment raises, the WARNING fires and the
        loop continues to the next attachment (defensive try/except
        per the function's exception-isolation contract)."""
        atts = [
            self._make_attachment(
                'WR_111_WeekEnding_041926_120000_aaa.xlsx', 11
            ),  # primary off-contract → delete will raise
            self._make_attachment(
                'WR_222_WeekEnding_041926_120000_bbb.xlsx', 22
            ),  # primary off-contract → delete will succeed
        ]
        sheet, client = self._make_sheet_with_attachments(atts)

        # First delete raises; second succeeds:
        client.Attachments.delete_attachment.side_effect = [
            RuntimeError('Simulated SDK 429'),
            None,
        ]
        generate_weekly_pdfs.cleanup_untracked_sheet_attachments(
            client,
            target_sheet_id=9999,
            valid_wr_weeks=set(),
            test_mode=False,
            target_sheet=sheet,
            variant_whitelist={'reduced_sub', 'reduced_sub_helper'},
        )
        # Both calls attempted; exception did not halt the loop:
        self.assertEqual(
            client.Attachments.delete_attachment.call_count, 2,
            'Loop must continue after a delete_attachment exception.'
        )

    # ── Test 7: source-level invariants (signature + log + counter) ──

    def test_source_carries_removed_off_contract_counter(self):
        """Source-level grep guards against drift between tests and
        the production function. Mirrors the WR-01 source-level
        pattern in TestPppCleanupUntrackedAttachments."""
        import inspect, pathlib
        import pipeline.cleanup  # W5: cleanup_untracked_sheet_attachments relocated here
        src = pathlib.Path(
            inspect.getsourcefile(generate_weekly_pdfs)
        ).read_text(encoding='utf-8') + "\n" + pathlib.Path(
            inspect.getsourcefile(pipeline.cleanup)
        ).read_text(encoding='utf-8')
        self.assertIn(
            'removed_off_contract = 0', src,
            'Source must initialize removed_off_contract = 0 alongside '
            'removed_variants = 0.'
        )
        self.assertIn(
            'off_contract_attachments = []', src,
            'Source must initialize off_contract_attachments = [] '
            'inside the per-row loop body.'
        )
        self.assertIn(
            'variant_whitelist is not None', src,
            'Source must carry the whitelist gate clause.'
        )
        self.assertIn(
            'removed_off_contract={removed_off_contract}', src,
            'End-of-function summary log must include the new counter.'
        )

    def test_source_carries_new_pii_log_marker(self):
        """The new INFO log embeds attachment name (WR + week). The
        marker must live in _PII_LOG_MARKERS per the [2026-05-15
        12:00] rule 3 (explicit marker; no accidental substring
        containment with pre-existing markers)."""
        self.assertIn(
            'Removed off-contract variant on sheet',
            generate_weekly_pdfs._PII_LOG_MARKERS,
            '_PII_LOG_MARKERS must include the new marker for the '
            'off-contract delete INFO log body.'
        )

    def test_source_off_contract_delete_log_body_present(self):
        import inspect, pathlib
        import pipeline.cleanup  # W5: cleanup_untracked_sheet_attachments relocated here
        src = pathlib.Path(
            inspect.getsourcefile(generate_weekly_pdfs)
        ).read_text(encoding='utf-8') + "\n" + pathlib.Path(
            inspect.getsourcefile(pipeline.cleanup)
        ).read_text(encoding='utf-8')
        self.assertIn(
            'Removed off-contract variant on sheet', src,
            'Source must carry the INFO log body verbatim '
            '(matched against the new _PII_LOG_MARKERS entry).'
        )


class TestPppCleanupInvocationCarriesWhitelist(unittest.TestCase):
    """Phase 1.1 Bug B2 (D-07 / D-08 / SUB-10): the PPP invocation
    site at L7189-7203 must pass
    ``variant_whitelist={'reduced_sub','reduced_sub_helper'}`` as the
    trailing kwarg; the TARGET_SHEET_ID invocation must NOT pass the
    kwarg (preserves None default = byte-identical legacy)."""

    @staticmethod
    def _read_source() -> str:
        # Phase 09 W6: the PPP cleanup invocation lives in main() (relocated
        # to pipeline/orchestrate.py) — concatenate it (follow-the-code).
        import inspect, pathlib
        import pipeline.orchestrate
        return (
            pathlib.Path(
                inspect.getsourcefile(generate_weekly_pdfs)
            ).read_text(encoding='utf-8')
            + "\n"
            + pathlib.Path(
                inspect.getsourcefile(pipeline.orchestrate)
            ).read_text(encoding='utf-8')
        )

    def test_ppp_invocation_passes_literal_whitelist(self):
        src = self._read_source()
        self.assertIn(
            "variant_whitelist={'reduced_sub', 'reduced_sub_helper'}",
            src,
            'PPP cleanup invocation must pass the literal '
            "whitelist {'reduced_sub', 'reduced_sub_helper'} per D-08."
        )

    def test_only_ppp_invocation_carries_whitelist(self):
        """Exactly one occurrence of the literal whitelist (PPP call
        site only). TARGET must NOT carry it — preserves D-09."""
        src = self._read_source()
        self.assertEqual(
            src.count(
                "variant_whitelist={'reduced_sub', 'reduced_sub_helper'}"
            ),
            1,
            'Exactly one whitelist literal expected (PPP only); '
            'TARGET preserves None default per D-09.'
        )

    def test_target_cleanup_does_not_carry_whitelist(self):
        """The TARGET invocation (the line containing
        ``client, TARGET_SHEET_ID,`` in the cleanup call) must not
        carry a ``variant_whitelist=`` argument."""
        import re
        src = self._read_source()
        # Extract the TARGET cleanup invocation (multi-line or
        # single-line tolerant). Subproject B (Task 7) introduced one
        # level of nested parens inside the call body
        # (``sub_offcontract_variants=(_target_offcontract or None)``),
        # so the matcher must tolerate a single level of balanced
        # inner parens — ``(?:[^()]|\([^()]*\))*`` consumes either a
        # non-paren char or a balanced single-level group — and then
        # stop at the call's own closing paren.
        target_match = re.search(
            r"cleanup_untracked_sheet_attachments\s*\(\s*\n?\s*"
            r"client\s*,\s*\n?\s*TARGET_SHEET_ID(?:[^()]|\([^()]*\))*\)",
            src,
            re.MULTILINE,
        )
        self.assertIsNotNone(
            target_match,
            'Expected TARGET cleanup invocation in source.'
        )
        self.assertNotIn(
            'variant_whitelist', target_match.group(0),
            'TARGET cleanup invocation must NOT carry '
            'variant_whitelist kwarg per D-09 (preserves byte-'
            'identical legacy behavior).'
        )

    def test_ppp_invocation_inside_sentry_span(self):
        """The PPP invocation block carries
        ``op="smartsheet.cleanup_ppp"`` and the whitelist kwarg must
        live INSIDE that block (sequential lines, not above the
        ``with`` statement)."""
        src = self._read_source()
        # The Sentry op for PPP appears immediately above the
        # cleanup call (the whole block lives inside that with).
        # Heuristic: locate ``smartsheet.cleanup_ppp`` and assert
        # the whitelist literal appears within ~1500 chars after.
        idx = src.find('smartsheet.cleanup_ppp')
        self.assertGreater(idx, -1)
        window = src[idx:idx + 1500]
        self.assertIn(
            "variant_whitelist={'reduced_sub', 'reduced_sub_helper'}",
            window,
            'The PPP whitelist kwarg must live inside the '
            'op="smartsheet.cleanup_ppp" Sentry span (within '
            '~1500 chars after the op marker).'
        )


class TestSentryTelemetryHelpers(unittest.TestCase):
    """Unit tests for the three pure Sentry telemetry helpers.

    These helpers are intentionally pure (A/B) or side-effect-guarded (C)
    so that PII-safety is test-enforced rather than relying on review alone.
    Both add_attachment and sentry_sdk.logger bypass before_send_log, making
    these assertions the primary defence against PII leakage into Sentry.
    """

    # ------------------------------------------------------------------ #
    # Helper A: _build_run_kpis                                           #
    # ------------------------------------------------------------------ #

    def _kpi_defaults(self, **overrides):
        """Return a valid kwargs dict for _build_run_kpis."""
        base = dict(
            files_generated=10,
            groups_total=15,
            groups_skipped=2,
            groups_generated=10,
            groups_uploaded=9,
            groups_errored=1,
            duration_seconds=120.0,
            sheets_discovered=5,
            rows_fetched=550,
            api_calls=42,
        )
        base.update(overrides)
        return base

    def test_kpi_returns_expected_keys(self):
        """_build_run_kpis returns all required KPI keys."""
        result = generate_weekly_pdfs._build_run_kpis(**self._kpi_defaults())
        expected_keys = {
            "files_generated",
            "groups_total",
            "groups_skipped",
            "groups_generated",
            "groups_uploaded",
            "groups_errored",
            "duration_seconds",
            "sheets_discovered",
            "rows_fetched",
            "api_calls",
            "groups_per_minute",
        }
        self.assertEqual(set(result.keys()), expected_keys)

    def test_kpi_throughput_computed_correctly(self):
        """groups_per_minute is derived: groups_generated / (duration_seconds/60)."""
        result = generate_weekly_pdfs._build_run_kpis(**self._kpi_defaults(
            groups_generated=10,
            duration_seconds=120.0,
        ))
        self.assertAlmostEqual(result["groups_per_minute"], 5.0, places=2)

    def test_kpi_zero_duration_gives_zero_throughput(self):
        """Zero duration_seconds must not raise ZeroDivisionError; throughput is 0.0."""
        result = generate_weekly_pdfs._build_run_kpis(**self._kpi_defaults(
            duration_seconds=0.0,
        ))
        self.assertEqual(result["groups_per_minute"], 0.0)

    def test_kpi_all_values_are_numeric(self):
        """Every value in the KPI dict is int or float — guarantees no PII string leakage."""
        result = generate_weekly_pdfs._build_run_kpis(**self._kpi_defaults())
        for key, value in result.items():
            self.assertIsInstance(
                value, (int, float),
                msg=f"KPI key '{key}' has non-numeric value {value!r} — potential PII leakage",
            )

    def test_kpi_no_string_values(self):
        """No string values allowed in KPI dict (strings could carry PII)."""
        result = generate_weekly_pdfs._build_run_kpis(**self._kpi_defaults())
        string_values = {k: v for k, v in result.items() if isinstance(v, str)}
        self.assertEqual(string_values, {}, msg=f"Unexpected string values: {string_values}")

    # ------------------------------------------------------------------ #
    # Helper B: _build_run_context_snapshot                               #
    # ------------------------------------------------------------------ #

    def _snap_success_kwargs(self):
        return dict(
            success=True,
            duration_seconds=90.0,
            groups_attempted=12,
            groups_generated=10,
            groups_uploaded=9,
            groups_errored=1,
            error_type=None,
        )

    def _snap_failure_kwargs(self):
        return dict(
            success=False,
            duration_seconds=45.0,
            groups_attempted=5,
            groups_generated=2,
            groups_uploaded=1,
            groups_errored=3,
            error_type="RuntimeError",
        )

    def test_snapshot_success_shape(self):
        """Success snapshot contains expected keys and success=True."""
        result = generate_weekly_pdfs._build_run_context_snapshot(**self._snap_success_kwargs())
        self.assertIn("success", result)
        self.assertTrue(result["success"])
        self.assertIn("duration_seconds", result)
        self.assertIn("groups_attempted", result)
        self.assertIn("groups_generated", result)
        self.assertIn("groups_uploaded", result)
        self.assertIn("groups_errored", result)

    def test_snapshot_failure_shape(self):
        """Failure snapshot contains error_type; success=False."""
        result = generate_weekly_pdfs._build_run_context_snapshot(**self._snap_failure_kwargs())
        self.assertFalse(result["success"])
        self.assertEqual(result.get("error_type"), "RuntimeError")

    def test_snapshot_success_no_error_type(self):
        """Success snapshot has no error_type key or it is None."""
        result = generate_weekly_pdfs._build_run_context_snapshot(**self._snap_success_kwargs())
        # error_type may be absent or None on success path
        error_type_val = result.get("error_type")
        self.assertIsNone(error_type_val)

    def test_snapshot_values_are_safe_types(self):
        """All values are int, float, bool, None, or error_type string only."""
        result = generate_weekly_pdfs._build_run_context_snapshot(**self._snap_failure_kwargs())
        for key, value in result.items():
            self.assertIsInstance(
                value, (int, float, bool, str, type(None)),
                msg=f"Snapshot key '{key}' has unexpected type {type(value).__name__}",
            )

    def test_snapshot_no_wr_token(self):
        """Serialized snapshot JSON must not contain WR-like tokens."""
        import json
        import re
        result = generate_weekly_pdfs._build_run_context_snapshot(**self._snap_failure_kwargs())
        serialized = json.dumps(result)
        # Must not contain WR followed by digits (WR12345 style)
        wr_pattern = re.compile(r'(?i)\bWR\d+\b')
        self.assertIsNone(
            wr_pattern.search(serialized),
            msg=f"Snapshot JSON contains WR token: {serialized}",
        )

    def test_snapshot_no_dollar_sign(self):
        """Serialized snapshot must not contain dollar amounts."""
        import json
        result = generate_weekly_pdfs._build_run_context_snapshot(**self._snap_failure_kwargs())
        serialized = json.dumps(result)
        self.assertNotIn("$", serialized, msg="Snapshot JSON contains dollar sign — potential price PII")

    def test_snapshot_no_foreman_name(self):
        """Snapshot built from safe inputs must not carry any PII name strings."""
        import json
        result = generate_weekly_pdfs._build_run_context_snapshot(**self._snap_failure_kwargs())
        serialized = json.dumps(result)
        # The only string value allowed is error_type (the exception class name)
        # Ensure a sample PII name is not present
        self.assertNotIn("John Smith", serialized)
        self.assertNotIn("foreman", serialized.lower().replace("groups", ""))

    # ------------------------------------------------------------------ #
    # Helper C: _sentry_log_event                                         #
    # ------------------------------------------------------------------ #

    def test_log_event_noop_without_dsn(self):
        """_sentry_log_event must return immediately (no-op) when SENTRY_DSN is falsy."""
        original_dsn = generate_weekly_pdfs.SENTRY_DSN
        try:
            generate_weekly_pdfs.SENTRY_DSN = ""
            # Must not raise even if sentry_sdk.logger is present
            generate_weekly_pdfs._sentry_log_event("info", "test milestone", count=5)
        finally:
            generate_weekly_pdfs.SENTRY_DSN = original_dsn

    def test_log_event_noop_without_logger_attr(self):
        """_sentry_log_event must no-op when sentry_sdk has no 'logger' attribute."""
        import sentry_sdk as _sdk
        original_dsn = generate_weekly_pdfs.SENTRY_DSN
        original_logger = getattr(_sdk, "logger", None)
        try:
            # Simulate SDK older than 2.54 — no logger attribute
            generate_weekly_pdfs.SENTRY_DSN = "https://fake@sentry.io/123"
            if hasattr(_sdk, "logger"):
                delattr(_sdk, "logger")
            # Must not raise
            generate_weekly_pdfs._sentry_log_event("info", "test milestone", count=5)
        finally:
            generate_weekly_pdfs.SENTRY_DSN = original_dsn
            if original_logger is not None:
                _sdk.logger = original_logger
            elif hasattr(_sdk, "logger"):
                delattr(_sdk, "logger")

    def test_log_event_does_not_raise_on_bad_level(self):
        """_sentry_log_event must never raise even if level is invalid."""
        original_dsn = generate_weekly_pdfs.SENTRY_DSN
        try:
            generate_weekly_pdfs.SENTRY_DSN = ""
            # Should not raise regardless of level
            generate_weekly_pdfs._sentry_log_event("nonexistent_level", "message", x=1)
        finally:
            generate_weekly_pdfs.SENTRY_DSN = original_dsn

    def test_log_event_does_not_raise_when_logger_call_fails(self):
        """_sentry_log_event swallows internal errors and never propagates them."""
        import sentry_sdk as _sdk

        class _BrokenLogger:
            def info(self, *a, **kw):
                raise RuntimeError("broken logger")

        original_dsn = generate_weekly_pdfs.SENTRY_DSN
        original_logger = getattr(_sdk, "logger", None)
        try:
            generate_weekly_pdfs.SENTRY_DSN = "https://fake@sentry.io/123"
            _sdk.logger = _BrokenLogger()
            # Must not propagate the RuntimeError
            generate_weekly_pdfs._sentry_log_event("info", "milestone", count=1)
        finally:
            generate_weekly_pdfs.SENTRY_DSN = original_dsn
            if original_logger is not None:
                _sdk.logger = original_logger
            elif hasattr(_sdk, "logger"):
                delattr(_sdk, "logger")


if __name__ == '__main__':
    unittest.main()
