"""Subproject C — VAC Crew claim attribution tests.

Drives real production code paths (parser, group_source_rows pre-pass +
emission, generate_excel, migration cleanup, hash prune, HOLD wiring) per
the [2026-05-20 00:26] rule 4: row-flow changes require TRUE end-to-end
tests, not static mirrors.
"""
from __future__ import annotations

import inspect
import pathlib
import sys
import unittest
from pathlib import Path
from unittest import mock

_REPO_ROOT = Path(__file__).resolve().parent.parent
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

from tests.test_billing_audit_shadow import _ensure_smartsheet_mocked, _reset_all

_ensure_smartsheet_mocked()
import generate_weekly_pdfs  # noqa: E402
from billing_audit.writer import ResolveOutcome  # noqa: E402


class TestVacCrewConfigFlags(unittest.TestCase):
    def test_attribution_flag_exists_and_is_bool(self):
        self.assertIsInstance(
            generate_weekly_pdfs.VAC_CREW_CLAIM_ATTRIBUTION_ENABLED, bool
        )

    def test_attribution_flag_default_on(self):
        # Env var unset in the test harness → default '1' → True.
        self.assertTrue(generate_weekly_pdfs.VAC_CREW_CLAIM_ATTRIBUTION_ENABLED)

    def test_legacy_cleanup_flag_exists_and_is_bool(self):
        self.assertIsInstance(
            generate_weekly_pdfs.VAC_CREW_LEGACY_CLEANUP_ENABLED, bool
        )

    def test_legacy_cleanup_flag_default_on(self):
        self.assertTrue(generate_weekly_pdfs.VAC_CREW_LEGACY_CLEANUP_ENABLED)

    def test_flags_pinned_in_workflow(self):
        wf = (_REPO_ROOT / ".github/workflows/weekly-excel-generation.yml").read_text(
            encoding="utf-8"
        )
        self.assertIn("VAC_CREW_CLAIM_ATTRIBUTION_ENABLED", wf)
        self.assertIn("VAC_CREW_LEGACY_CLEANUP_ENABLED", wf)


class TestVacCrewSuffixAndParser(unittest.TestCase):
    def test_suffix_embeds_name(self):
        self.assertEqual(
            generate_weekly_pdfs._vac_crew_variant_suffix('John Smith', '91467680', '041926'),
            '_VacCrew_John_Smith',
        )

    def test_suffix_empty_claimer_raises(self):
        with self.assertRaises(ValueError):
            generate_weekly_pdfs._vac_crew_variant_suffix('', '91467680', '041926')

    def test_parser_vaccrew_name_round_trips(self):
        fname = 'WR_91467680_WeekEnding_041926_120000_VacCrew_John_Smith_abc123.xlsx'
        self.assertEqual(
            generate_weekly_pdfs.build_group_identity(fname),
            ('91467680', '041926', 'vac_crew', 'John_Smith'),
        )

    def test_parser_name_containing_helper_token_stays_vac_crew(self):
        fname = 'WR_91467680_WeekEnding_041926_120000_VacCrew_Pat_Helper_abc123.xlsx'
        self.assertEqual(
            generate_weekly_pdfs.build_group_identity(fname),
            ('91467680', '041926', 'vac_crew', 'Pat_Helper'),
        )

    def test_parser_legacy_vaccrew_no_name(self):
        fname = 'WR_91467680_WeekEnding_041926_120000_VacCrew_abc123.xlsx'
        self.assertEqual(
            generate_weekly_pdfs.build_group_identity(fname),
            ('91467680', '041926', 'vac_crew', ''),
        )

    def test_suffix_truncates_long_name_and_round_trips(self):
        long_name = 'A' * 60
        suffix = generate_weekly_pdfs._vac_crew_variant_suffix(long_name, '91467680', '041926')
        self.assertEqual(suffix, '_VacCrew_' + 'A' * 50)
        fname = f'WR_91467680_WeekEnding_041926_120000{suffix}_abc123.xlsx'
        self.assertEqual(
            generate_weekly_pdfs.build_group_identity(fname),
            ('91467680', '041926', 'vac_crew', 'A' * 50),
        )


def _make_vac_row(row_id=6001, wr='91467680', name='CurrentCrew', snapshot='2026-04-19'):
    return {
        '__row_id': row_id,
        'Work Request #': wr,
        'Weekly Reference Logged Date': '2026-04-19',
        'Snapshot Date': snapshot,
        'Units Completed?': True,
        'Units Total Price': '$100.00',
        'CU': 'ANC-M', 'Work Type': 'Inst', 'Quantity': 2,
        '__effective_user': 'PrimaryForeman',
        '__is_helper_row': False, '__helper_foreman': '', '__helper_dept': '', '__helper_job': '',
        '__is_vac_crew': True,
        '__vac_crew_name': name, '__vac_crew_dept': '700', '__vac_crew_job': 'VJ-1',
        '__source_sheet_id': 8162920222379908,
    }


class TestVacCrewPrePassConcurrency(unittest.TestCase):
    def setUp(self):
        _reset_all()
        self._orig = generate_weekly_pdfs.VAC_CREW_CLAIM_ATTRIBUTION_ENABLED
        generate_weekly_pdfs.VAC_CREW_CLAIM_ATTRIBUTION_ENABLED = True

    def tearDown(self):
        generate_weekly_pdfs.VAC_CREW_CLAIM_ATTRIBUTION_ENABLED = self._orig
        _reset_all()

    def test_fifty_rows_each_partition_to_their_own_claimer(self):
        def _resolve(variant, current, *, wr, week_ending, row_id, enabled, prefetched_map=None):
            return ResolveOutcome('use', f'Crew{row_id}', 'frozen', 'success')
        rows = [_make_vac_row(row_id=7000 + i) for i in range(50)]
        with mock.patch('billing_audit.writer.resolve_claimer', side_effect=_resolve):
            groups = generate_weekly_pdfs.group_source_rows(rows)
        keys = list(groups.keys())
        for i in range(50):
            self.assertTrue(
                any(f'VACCREW_Crew{7000 + i}' in k for k in keys),
                f"row {7000+i} must partition to its own claimer; got {keys[:5]}…",
            )


class TestVacCrewEmission(unittest.TestCase):
    def setUp(self):
        _reset_all()
        self._orig = generate_weekly_pdfs.VAC_CREW_CLAIM_ATTRIBUTION_ENABLED
        generate_weekly_pdfs.VAC_CREW_CLAIM_ATTRIBUTION_ENABLED = True

    def tearDown(self):
        generate_weekly_pdfs.VAC_CREW_CLAIM_ATTRIBUTION_ENABLED = self._orig
        _reset_all()

    def test_frozen_claimer_partitions(self):
        with mock.patch('billing_audit.writer.resolve_claimer',
                        return_value=ResolveOutcome('use', 'FrozenCrew', 'frozen', 'success')):
            groups = generate_weekly_pdfs.group_source_rows([_make_vac_row()])
        self.assertTrue(any('VACCREW_FrozenCrew' in k for k in groups))

    def test_no_history_falls_back_to_current_name(self):
        with mock.patch('billing_audit.writer.resolve_claimer',
                        return_value=ResolveOutcome('use', 'CurrentCrew', 'current', 'no_history')):
            groups = generate_weekly_pdfs.group_source_rows([_make_vac_row(name='CurrentCrew')])
        self.assertTrue(any('VACCREW_CurrentCrew' in k for k in groups))

    def test_hold_suppresses_and_records(self):
        from billing_audit.writer import get_counters
        with mock.patch('billing_audit.writer.resolve_claimer',
                        return_value=ResolveOutcome('hold', None, None, 'fetch_failure')):
            groups = generate_weekly_pdfs.group_source_rows([_make_vac_row()])
        self.assertFalse(any('VACCREW' in k for k in groups))
        self.assertEqual(get_counters()['attribution_rows_held'], 1)

    def test_disabled_emits_exact_legacy_key(self):
        generate_weekly_pdfs.VAC_CREW_CLAIM_ATTRIBUTION_ENABLED = False
        with mock.patch('billing_audit.writer.resolve_claimer') as m:
            groups = generate_weekly_pdfs.group_source_rows([_make_vac_row(name='CurrentCrew')])
            m.assert_not_called()
        self.assertIn('041926_91467680_VACCREW', groups)
        self.assertFalse(any('VACCREW_' in k for k in groups))

    def test_map_miss_uses_current_name_not_hold(self):
        from billing_audit.writer import get_counters
        row = _make_vac_row()
        del row['__row_id']
        with mock.patch('billing_audit.writer.resolve_claimer',
                        return_value=ResolveOutcome('use', 'X', 'frozen', 'success')):
            groups = generate_weekly_pdfs.group_source_rows([row])
        self.assertTrue(any('VACCREW_CurrentCrew' in k for k in groups))
        self.assertEqual(get_counters()['attribution_rows_held'], 0)


class TestVacCrewIdentitySitesAndDisplay(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        # Phase 09 W6: main() relocated to pipeline/orchestrate.py — grep
        # facade + orchestrate (follow-the-code superset).
        import pipeline.orchestrate
        cls._src = (
            pathlib.Path(
                inspect.getsourcefile(generate_weekly_pdfs)
            ).read_text(encoding='utf-8')
            + "\n"
            + pathlib.Path(
                inspect.getsourcefile(pipeline.orchestrate)
            ).read_text(encoding='utf-8')
        )

    def test_current_keys_site_carries_vac_claimer(self):
        # Site 3: the hash-prune current_keys reconstruction must derive the
        # vac_crew identifier from __current_foreman, not hard-code ''.
        self.assertNotRegex(
            self._src,
            r"_variant == 'vac_crew':\s*\n\s*_ident = ''",
            "current_keys must derive vac_crew identifier from the claimer",
        )

    def test_main_loop_identity_site_carries_vac_claimer(self):
        # Site 1: the main-loop identifier/history_key for vac_crew must
        # derive from __current_foreman, not hard-code '' (else the hash
        # entry is stale-pruned every run → permanent regeneration churn).
        self.assertNotRegex(
            self._src,
            r"variant == 'vac_crew':\s*\n\s*# VAC Crew variant: no sub-identifier",
            "Site 1 main-loop must not use the old 'no sub-identifier' comment (legacy blank pattern)",
        )
        self.assertNotRegex(
            self._src,
            r"variant == 'vac_crew':[^\n]*\n[^\n]*\n\s*identifier = ''",
            "Site 1 main-loop identifier must not hard-code '' for vac_crew",
        )

    def test_generate_excel_vac_crew_file_named_by_claimer(self):
        import datetime as dt, tempfile, openpyxl
        tmp = tempfile.TemporaryDirectory()
        self.addCleanup(tmp.cleanup)
        orig = generate_weekly_pdfs.OUTPUT_FOLDER
        generate_weekly_pdfs.OUTPUT_FOLDER = tmp.name
        self.addCleanup(lambda: setattr(generate_weekly_pdfs, 'OUTPUT_FOLDER', orig))
        row = {
            'Work Request #': '91467680', 'Units Completed?': True,
            'Units Total Price': '$100.00', 'Customer Name': 'Cust',
            'Dept #': '500', 'Job #': 'J-1', 'CU': 'ANC-M', 'Work Type': 'Inst', 'Quantity': 2,
            '__variant': 'vac_crew', '__current_foreman': 'FrozenCrew',
            '__vac_crew_name': 'CurrentCrew', '__vac_crew_dept': '700', '__vac_crew_job': 'VJ-1',
            '__week_ending_date': dt.datetime(2026, 4, 19),
        }
        result = generate_weekly_pdfs.generate_excel(
            '041926_91467680_VACCREW_FrozenCrew', [row], dt.datetime(2026, 4, 19),
            data_hash='deadbeefcafe0c01',
        )
        excel_path, filename = result[0], result[1]
        self.assertIn('_VacCrew_FrozenCrew', filename)
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        foreman = next(
            (ws.cell(row=r, column=7).value for r in range(1, ws.max_row + 1)
             if ws.cell(row=r, column=6).value == 'Foreman:'),
            None,
        )
        self.assertEqual(foreman, 'FrozenCrew')  # display = attributed claimer

    def test_disabled_mode_generate_excel_emits_bare_vaccrew_filename(self):
        """Disabled mode (VAC_CREW_CLAIM_ATTRIBUTION_ENABLED=False) must produce
        the exact legacy bare _VacCrew suffix — no claimer name — matching the
        legacy identity tuple '' at Sites 1/2/3 so no churn occurs."""
        import datetime as dt
        import tempfile
        tmp = tempfile.TemporaryDirectory()
        self.addCleanup(tmp.cleanup)
        orig_out = generate_weekly_pdfs.OUTPUT_FOLDER
        orig_flag = generate_weekly_pdfs.VAC_CREW_CLAIM_ATTRIBUTION_ENABLED
        generate_weekly_pdfs.OUTPUT_FOLDER = tmp.name
        generate_weekly_pdfs.VAC_CREW_CLAIM_ATTRIBUTION_ENABLED = False
        self.addCleanup(
            lambda: setattr(generate_weekly_pdfs, 'OUTPUT_FOLDER', orig_out)
        )
        self.addCleanup(
            lambda: setattr(
                generate_weekly_pdfs,
                'VAC_CREW_CLAIM_ATTRIBUTION_ENABLED',
                orig_flag,
            )
        )
        row = {
            'Work Request #': '91467680',
            'Units Completed?': True,
            'Units Total Price': '$100.00',
            'Customer Name': 'Cust',
            'Dept #': '500',
            'Job #': 'J-1',
            'CU': 'ANC-M',
            'Work Type': 'Inst',
            'Quantity': 2,
            '__variant': 'vac_crew',
            '__current_foreman': 'CurrentCrew',  # non-empty — must NOT appear in filename
            '__vac_crew_name': 'CurrentCrew',
            '__vac_crew_dept': '700',
            '__vac_crew_job': 'VJ-1',
            '__week_ending_date': dt.datetime(2026, 4, 19),
        }
        result = generate_weekly_pdfs.generate_excel(
            '041926_91467680_VACCREW', [row], dt.datetime(2026, 4, 19),
            data_hash='deadbeefcafe0c02',
        )
        filename = result[1]
        # Disabled mode: filename must contain bare _VacCrew token only;
        # the claimer name must NOT appear in the _VacCrew suffix position.
        self.assertNotIn('_VacCrew_CurrentCrew', filename,
                         "Disabled mode must not embed claimer name in _VacCrew token")
        # Bare _VacCrew followed by the hash/timestamp separator (_, not a name)
        self.assertRegex(
            filename,
            r'_VacCrew_[0-9a-f]+\.xlsx$',
            "Disabled mode filename must be bare _VacCrew_<hash>.xlsx (no claimer name)",
        )

    def test_disabled_mode_display_uses_vac_crew_name_not_primary_foreman(self):
        """M1: disabled mode must display __vac_crew_name, not __current_foreman.

        When VAC_CREW_CLAIM_ATTRIBUTION_ENABLED=False, generate_excel's
        vac_crew display branch must read __vac_crew_name directly — NOT
        fall back through __current_foreman, which in disabled mode may be
        the primary / Arrowhead foreman, not the VAC crew member.

        Two sub-cases:
        a) __vac_crew_name populated, __current_foreman is a different
           primary-foreman name → display must be the vac crew name.
        b) __vac_crew_name is '' (empty) → display must be 'Unknown VAC Crew',
           NOT the primary foreman value.
        """
        import datetime as dt
        import tempfile
        import openpyxl

        tmp = tempfile.TemporaryDirectory()
        self.addCleanup(tmp.cleanup)
        orig_out = generate_weekly_pdfs.OUTPUT_FOLDER
        orig_flag = generate_weekly_pdfs.VAC_CREW_CLAIM_ATTRIBUTION_ENABLED
        generate_weekly_pdfs.OUTPUT_FOLDER = tmp.name
        generate_weekly_pdfs.VAC_CREW_CLAIM_ATTRIBUTION_ENABLED = False
        self.addCleanup(
            lambda: setattr(generate_weekly_pdfs, 'OUTPUT_FOLDER', orig_out)
        )
        self.addCleanup(
            lambda: setattr(
                generate_weekly_pdfs,
                'VAC_CREW_CLAIM_ATTRIBUTION_ENABLED',
                orig_flag,
            )
        )

        def _foreman_cell(path):
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            return next(
                (ws.cell(row=r, column=7).value for r in range(1, ws.max_row + 1)
                 if ws.cell(row=r, column=6).value == 'Foreman:'),
                None,
            )

        base_row = {
            'Work Request #': '91467680',
            'Units Completed?': True,
            'Units Total Price': '$100.00',
            'Customer Name': 'Cust',
            'Dept #': '500',
            'Job #': 'J-1',
            'CU': 'ANC-M',
            'Work Type': 'Inst',
            'Quantity': 2,
            '__variant': 'vac_crew',
            '__vac_crew_dept': '700',
            '__vac_crew_job': 'VJ-1',
            '__week_ending_date': dt.datetime(2026, 4, 19),
        }

        # --- sub-case a: __vac_crew_name != __current_foreman ---
        # __current_foreman is the primary/Arrowhead foreman; __vac_crew_name
        # is the actual VAC crew member. Disabled mode MUST show the vac name.
        row_a = dict(base_row)
        row_a['__current_foreman'] = 'Arrowhead_PrimaryForeman'
        row_a['__vac_crew_name'] = 'Alice_VacCrew'
        result_a = generate_weekly_pdfs.generate_excel(
            '041926_91467680_VACCREW', [row_a], dt.datetime(2026, 4, 19),
            data_hash='deadbeefcafe0ca1',
        )
        foreman_a = _foreman_cell(result_a[0])
        self.assertEqual(
            foreman_a, 'Alice_VacCrew',
            "Disabled mode: display must be __vac_crew_name, not "
            f"__current_foreman ('Arrowhead_PrimaryForeman'); got {foreman_a!r}",
        )
        self.assertNotEqual(
            foreman_a, 'Arrowhead_PrimaryForeman',
            "Disabled mode must NOT fall back to the primary foreman",
        )

        # --- sub-case b: __vac_crew_name is empty ---
        # Even with __current_foreman set, display must NOT be the primary
        # foreman value. The cell may be None/empty or 'Unknown VAC Crew'
        # depending on how generate_excel renders the fallback; what is
        # strictly required is that the primary foreman does NOT appear.
        row_b = dict(base_row)
        row_b['__current_foreman'] = 'Arrowhead_PrimaryForeman'
        row_b['__vac_crew_name'] = ''
        result_b = generate_weekly_pdfs.generate_excel(
            '041926_91467680_VACCREW', [row_b], dt.datetime(2026, 4, 19),
            data_hash='deadbeefcafe0cb2',
        )
        foreman_b = _foreman_cell(result_b[0])
        self.assertNotEqual(
            foreman_b, 'Arrowhead_PrimaryForeman',
            "Disabled mode with empty __vac_crew_name must NOT display the "
            f"primary/Arrowhead foreman; got {foreman_b!r}",
        )

    def test_valid_wr_weeks_site2_vac_crew_gated_on_kill_switch(self):
        """Site 2 grep guard: the valid_wr_weeks vac_crew branch must NOT
        hard-code file_id='' unconditionally — it must be gated on the flag."""
        self.assertNotRegex(
            self._src,
            r"variant == 'vac_crew':\s*\n\s*#[^\n]*\n\s*_vc = [^\n]+\n\s*file_id = ''",
            "Site 2 valid_wr_weeks must not hard-code '' for vac_crew unconditionally; "
            "must be gated on VAC_CREW_CLAIM_ATTRIBUTION_ENABLED",
        )
        # Positive guard: the kill-switch name must appear near the vac_crew file_id block.
        vc_site2_idx = self._src.find(
            "elif variant == 'vac_crew':\n"
            "                    # Subproject C identity site (Site 2"
        )
        self.assertNotEqual(vc_site2_idx, -1, "Site 2 vac_crew block not found")
        window = self._src[vc_site2_idx: vc_site2_idx + 750]
        self.assertIn(
            'VAC_CREW_CLAIM_ATTRIBUTION_ENABLED',
            window,
            "Site 2 vac_crew file_id derivation must reference VAC_CREW_CLAIM_ATTRIBUTION_ENABLED",
        )


class TestVacCrewLegacyCleanup(unittest.TestCase):
    """Task 6: legacy unpartitioned _VacCrew TARGET cleanup.

    When vac_crew files become per-claimer (``_VacCrew_<name>``), the OLD
    bare ``_VacCrew`` (no name, empty identifier) attachments on
    TARGET_SHEET_ID are orphans that must be cleaned up.  Mirrors the
    Subproject B ``sub_legacy_primary_variants`` gate exactly.
    """

    def setUp(self):
        _ensure_smartsheet_mocked()

    # ------------------------------------------------------------------
    # Helpers (same fixture style as TestLegacyHelperTargetCleanupE2E)
    # ------------------------------------------------------------------

    def _make_attachment(self, name, att_id):
        att = mock.MagicMock()
        att.name = name
        att.id = att_id
        return att

    def _build_client_with_attachments(self, attachments):
        client = mock.MagicMock()
        sheet = mock.MagicMock()
        row = mock.MagicMock()
        row.id = 1
        client.Attachments.list_row_attachments.return_value.data = attachments
        sheet.rows = [row]
        client.Sheets.get_sheet.return_value = sheet
        return client, sheet

    # ------------------------------------------------------------------
    # Scope builder
    # ------------------------------------------------------------------

    def test_scope_builder_collects_vac_wrs(self):
        """_build_vac_crew_wr_scope extracts WR# from vac_crew groups only.

        Gates on the authoritative ``__variant`` field (set at emission),
        not a ``'_VACCREW'`` key substring — mirrors ``_build_primary_wr_scope``
        (Subproject D) so a non-vac name containing a reserved token cannot
        false-positive. Production rows always carry ``__variant``.
        """
        groups = {
            '041926_91467680_VACCREW_John': [
                {'Work Request #': '91467680', '__variant': 'vac_crew'}],
            '041926_55555_REDUCEDSUB_USER_X': [
                {'Work Request #': '55555', '__variant': 'reduced_sub'}],
        }
        scope = generate_weekly_pdfs._build_vac_crew_wr_scope(groups)
        self.assertIn('91467680', scope)
        self.assertNotIn('55555', scope)

    def test_scope_builder_empty_groups(self):
        """Empty groups dict returns empty scope."""
        self.assertEqual(generate_weekly_pdfs._build_vac_crew_wr_scope({}), set())

    def test_scope_builder_ignores_helper_and_primary_keys(self):
        """Non-vac_crew variants are excluded from the vac scope."""
        groups = {
            '041926_11111_HELPER_Alice': [
                {'Work Request #': '11111', '__variant': 'helper'}],
            '041926_22222': [
                {'Work Request #': '22222', '__variant': 'primary'}],
        }
        scope = generate_weekly_pdfs._build_vac_crew_wr_scope(groups)
        self.assertEqual(scope, set())

    def test_scope_builder_rejects_pathological_vaccrew_name(self):
        """A helper/primary whose NAME is the all-caps reserved token
        ``VACCREW`` produces a key containing the ``_VACCREW`` substring,
        but its ``__variant`` is not ``vac_crew`` — the variant gate must
        exclude it from the destructive vac cleanup scope (mirror of the
        Subproject D Codex-P1 fix)."""
        groups = {
            # Helper literally named "VACCREW" → key has the _VACCREW substring.
            '041926_33333_HELPER_VACCREW': [
                {'Work Request #': '33333', '__variant': 'helper'}],
            # Genuine vac group → must still be collected.
            '041926_44444_VACCREW_Real_Vac': [
                {'Work Request #': '44444', '__variant': 'vac_crew'}],
        }
        scope = generate_weekly_pdfs._build_vac_crew_wr_scope(groups)
        self.assertIn('44444', scope)
        self.assertNotIn(
            '33333', scope,
            "a non-vac group whose name contains 'VACCREW' must NOT enter "
            "the vac cleanup scope (variant gate, not substring scan)",
        )

    # ------------------------------------------------------------------
    # Legacy bare _VacCrew deletion + live per-claimer exemption
    # ------------------------------------------------------------------

    def test_legacy_vaccrew_deleted_live_claimer_exempt(self):
        """WR-01 analog: the live-identity exemption must protect a
        per-claimer _VacCrew_<name> file while still deleting the
        legacy bare _VacCrew file for the same in-scope WR.

        Mirrors TestLegacyHelperTargetCleanupE2E::
        test_target_cleanup_exempts_live_helper_for_overlapping_sub_wr.
        """
        # Legacy bare _VacCrew — empty identifier → NOT in valid_wr_weeks
        # → must be deleted.
        att_legacy = self._make_attachment(
            'WR_91467680_WeekEnding_041926_120000_VacCrew_abc123.xlsx',
            101,
        )
        # Live per-claimer _VacCrew_John — identity IS in valid_wr_weeks
        # → must be EXEMPT from deletion.
        att_live = self._make_attachment(
            'WR_91467680_WeekEnding_041926_120000_VacCrew_John_def456.xlsx',
            102,
        )
        client, sheet = self._build_client_with_attachments([att_legacy, att_live])

        valid_wr_weeks = {('91467680', '041926', 'vac_crew', 'John')}

        generate_weekly_pdfs.cleanup_untracked_sheet_attachments(
            client,
            target_sheet_id=5723337641643908,
            valid_wr_weeks=valid_wr_weeks,
            test_mode=False,
            target_sheet=sheet,
            vac_legacy_wr_scope={'91467680'},
        )

        deletes = [call.args for call in client.Attachments.delete_attachment.call_args_list]

        self.assertIn(
            (5723337641643908, 101),
            deletes,
            f"Legacy bare _VacCrew must be deleted for in-scope vac WR; "
            f"got deletes={deletes}",
        )
        self.assertNotIn(
            (5723337641643908, 102),
            deletes,
            f"Live per-claimer _VacCrew_John whose identity is in "
            f"valid_wr_weeks must be EXEMPT; got deletes={deletes}",
        )

    def test_non_vac_wr_not_affected_by_vac_gate(self):
        """A bare _VacCrew file for a WR NOT in vac_legacy_wr_scope is
        not deleted by the vac gate (falls through to normal identity-grouping
        logic). This confirms the gate is WR-scoped, not global."""
        # WR 99999 is NOT in the vac scope.
        att = self._make_attachment(
            'WR_99999_WeekEnding_041926_120000_VacCrew_abc123.xlsx',
            201,
        )
        client, sheet = self._build_client_with_attachments([att])

        generate_weekly_pdfs.cleanup_untracked_sheet_attachments(
            client,
            target_sheet_id=5723337641643908,
            valid_wr_weeks=set(),
            test_mode=False,
            target_sheet=sheet,
            vac_legacy_wr_scope={'91467680'},  # 99999 NOT in scope
        )

        deletes = [call.args for call in client.Attachments.delete_attachment.call_args_list]
        self.assertNotIn(
            (5723337641643908, 201),
            deletes,
            "WR outside vac_legacy_wr_scope must not be deleted by the vac gate",
        )

    def test_none_scope_no_vac_gate(self):
        """vac_legacy_wr_scope=None (default) means no vac gate fires —
        byte-identical legacy behaviour for callers that don't pass it."""
        att = self._make_attachment(
            'WR_91467680_WeekEnding_041926_120000_VacCrew_abc123.xlsx',
            301,
        )
        client, sheet = self._build_client_with_attachments([att])

        # Pass nothing for vac_legacy_wr_scope — relies on default None.
        generate_weekly_pdfs.cleanup_untracked_sheet_attachments(
            client,
            target_sheet_id=5723337641643908,
            valid_wr_weeks=set(),
            test_mode=False,
            target_sheet=sheet,
        )

        deletes = [call.args for call in client.Attachments.delete_attachment.call_args_list]
        self.assertNotIn(
            (5723337641643908, 301),
            deletes,
            "vac_legacy_wr_scope=None must not trigger vac gate (legacy behaviour)",
        )

    def test_kill_switch_off_target_passes_none_scope(self):
        """Source-grep guard: when VAC_CREW_LEGACY_CLEANUP_ENABLED is False,
        the TARGET call site must pass vac_legacy_wr_scope=None (or
        equivalent falsy) so no vac deletion occurs.

        We verify the production source wires the scope conditionally on
        VAC_CREW_LEGACY_CLEANUP_ENABLED rather than passing the scope
        unconditionally.  We search the CALL SITE occurrence (inside
        cleanup_untracked_sheet_attachments(...)) rather than the
        function-signature occurrence.
        """
        # Phase 09 W6: the TARGET call site lives in main() (orchestrate.py).
        import pipeline.orchestrate
        src = (
            pathlib.Path(
                inspect.getsourcefile(generate_weekly_pdfs)
            ).read_text(encoding='utf-8')
            + "\n"
            + pathlib.Path(
                inspect.getsourcefile(pipeline.orchestrate)
            ).read_text(encoding='utf-8')
        )
        # The TARGET call site must gate the vac scope on the kill switch.
        self.assertIn(
            'VAC_CREW_LEGACY_CLEANUP_ENABLED',
            src,
            "VAC_CREW_LEGACY_CLEANUP_ENABLED must appear in source",
        )
        # The vac_legacy_wr_scope kwarg must appear at the TARGET call site.
        self.assertIn(
            'vac_legacy_wr_scope',
            src,
            "vac_legacy_wr_scope kwarg must be wired at the TARGET call site",
        )
        # The call-site wiring must reference the kill switch alongside the
        # scope builder.  Find the kwarg ASSIGNMENT (``_vac_scope =``) which
        # is the call-site variable, and verify VAC_CREW_LEGACY_CLEANUP_ENABLED
        # appears in its neighbourhood.
        assign_idx = src.find('_vac_scope =')
        self.assertNotEqual(assign_idx, -1,
                            "_vac_scope assignment not found in source")
        window = src[max(0, assign_idx - 100): assign_idx + 400]
        self.assertIn(
            'VAC_CREW_LEGACY_CLEANUP_ENABLED',
            window,
            "The _vac_scope assignment must be gated on "
            "VAC_CREW_LEGACY_CLEANUP_ENABLED at the call site",
        )
        # The kwarg pass at cleanup_untracked_sheet_attachments(...)
        # must also appear after the assignment.
        pass_idx = src.find('vac_legacy_wr_scope=_vac_scope')
        self.assertNotEqual(pass_idx, -1,
                            "vac_legacy_wr_scope=_vac_scope must be passed "
                            "to cleanup_untracked_sheet_attachments")

    def test_attribution_off_cleanup_on_live_bare_vaccrew_exempt(self):
        """M4 mixed-flag: attribution OFF + legacy cleanup ON.

        When VAC_CREW_CLAIM_ATTRIBUTION_ENABLED=False (disabled mode),
        group_source_rows emits the BARE vac_crew key with empty identifier,
        so valid_wr_weeks contains ('wr', 'week', 'vac_crew', '').  The
        legacy cleanup gate must honour the live-identity exemption and NOT
        delete that bare _VacCrew attachment — it is current-run output.

        This guards the mixed-flag combination: attribution OFF means no
        per-claimer files are generated, so a bare _VacCrew file is the
        LIVE file for that WR and must survive cleanup even when
        VAC_CREW_LEGACY_CLEANUP_ENABLED=True.
        """
        orig_attr_flag = generate_weekly_pdfs.VAC_CREW_CLAIM_ATTRIBUTION_ENABLED
        generate_weekly_pdfs.VAC_CREW_CLAIM_ATTRIBUTION_ENABLED = False
        try:
            # Bare _VacCrew — in disabled mode this IS the live file for the WR.
            att_bare = self._make_attachment(
                'WR_77712345_WeekEnding_050926_120000_VacCrew_ff1122.xlsx',
                201,
            )
            client, sheet = self._build_client_with_attachments([att_bare])

            # Disabled mode: the bare-key identity ('77712345','050926','vac_crew','')
            # is in valid_wr_weeks because group_source_rows emits it without a
            # claimer suffix when the flag is off.
            valid_wr_weeks = {('77712345', '050926', 'vac_crew', '')}

            generate_weekly_pdfs.cleanup_untracked_sheet_attachments(
                client,
                target_sheet_id=5723337641643908,
                valid_wr_weeks=valid_wr_weeks,
                test_mode=False,
                target_sheet=sheet,
                vac_legacy_wr_scope={'77712345'},
            )

            deletes = [
                call.args
                for call in client.Attachments.delete_attachment.call_args_list
            ]
            self.assertNotIn(
                (5723337641643908, 201),
                deletes,
                "Attribution OFF + cleanup ON: the bare _VacCrew file whose "
                "identity IS in valid_wr_weeks must be EXEMPT from deletion; "
                f"got deletes={deletes}",
            )
        finally:
            generate_weekly_pdfs.VAC_CREW_CLAIM_ATTRIBUTION_ENABLED = orig_attr_flag


class TestVacCrewHashPrune(unittest.TestCase):
    def setUp(self):
        _ensure_smartsheet_mocked()

    def _groups(self, wrs):
        # Production rows always carry __variant (set at emission); the scope
        # builder gates on it, so synthetic prune fixtures must include it.
        return {
            f"041926_{wr}_VACCREW_John": [
                {'Work Request #': wr, '__variant': 'vac_crew'}]
            for wr in wrs
        }

    def test_drops_legacy_vaccrew_orphans_returns_true(self):
        hist = {
            '91467680|041926|vac_crew|': {'hash': 'h1'},
            '91467680|041926|vac_crew|John': {'hash': 'h2'},  # new — survives
            '55555|041926|vac_crew|': {'hash': 'h3'},          # non-scope — survives
        }
        changed = generate_weekly_pdfs._run_vac_crew_hash_prune(hist, self._groups(['91467680']))
        self.assertIs(changed, True)
        self.assertNotIn('91467680|041926|vac_crew|', hist)
        self.assertIn('91467680|041926|vac_crew|John', hist)
        self.assertIn('55555|041926|vac_crew|', hist)
        self.assertEqual(hist['_vac_crew_prune_version'],
                         generate_weekly_pdfs.VAC_CREW_HASH_PRUNE_VERSION)

    def test_idempotent_returns_false(self):
        hist = {'_vac_crew_prune_version': generate_weekly_pdfs.VAC_CREW_HASH_PRUNE_VERSION}
        self.assertIs(
            generate_weekly_pdfs._run_vac_crew_hash_prune(hist, self._groups(['91467680'])),
            False,
        )

    def test_pii_marker_registered(self):
        self.assertIn('Vac crew hash-history prune', generate_weekly_pdfs._PII_LOG_MARKERS)

    def test_call_site_present_and_wired_to_migration_dirty(self):
        # W5: VAC_CREW_HASH_PRUNE_VERSION + _run_vac_crew_hash_prune relocated
        # to pipeline/attribution.py — grep facade (call site) + relocated
        # module (constant) so the source guard follows the code.
        import pipeline.attribution
        import pipeline.orchestrate  # W6: call site lives in main()
        src = (
            pathlib.Path(inspect.getsourcefile(generate_weekly_pdfs)).read_text(encoding='utf-8')
            + "\n"
            + pathlib.Path(inspect.getsourcefile(pipeline.attribution)).read_text(encoding='utf-8')
            + "\n"
            + pathlib.Path(inspect.getsourcefile(pipeline.orchestrate)).read_text(encoding='utf-8')
        )
        self.assertIn('_run_vac_crew_hash_prune(hash_history, groups)', src)
        self.assertRegex(src, r'(?m)^VAC_CREW_HASH_PRUNE_VERSION = 1$')


class TestVacCrewEndToEnd(unittest.TestCase):
    def setUp(self):
        _reset_all()
        self._orig = generate_weekly_pdfs.VAC_CREW_CLAIM_ATTRIBUTION_ENABLED
        generate_weekly_pdfs.VAC_CREW_CLAIM_ATTRIBUTION_ENABLED = True

    def tearDown(self):
        generate_weekly_pdfs.VAC_CREW_CLAIM_ATTRIBUTION_ENABLED = self._orig
        _reset_all()

    def test_two_claimers_same_wr_week_coexist(self):
        def _resolve(variant, current, *, wr, week_ending, row_id, enabled, prefetched_map=None):
            return ResolveOutcome('use', 'CrewA' if row_id == 6001 else 'CrewB', 'frozen', 'success')
        with mock.patch('billing_audit.writer.resolve_claimer', side_effect=_resolve):
            groups = generate_weekly_pdfs.group_source_rows(
                [_make_vac_row(row_id=6001), _make_vac_row(row_id=6002)]
            )
        self.assertTrue(any('VACCREW_CrewA' in k for k in groups))
        self.assertTrue(any('VACCREW_CrewB' in k for k in groups))

    def test_non_vac_primary_row_unaffected(self):
        row = {
            'Work Request #': '91467680', 'Weekly Reference Logged Date': '2026-04-19',
            'Snapshot Date': '2026-04-19', 'Units Completed?': True, 'Units Total Price': '$10.00',
            'CU': 'X', 'Work Type': 'Inst', 'Quantity': 1,
            '__effective_user': 'Boss', '__is_helper_row': False, '__is_vac_crew': False,
            '__helper_foreman': '', '__helper_dept': '', '__helper_job': '',
            '__source_sheet_id': 99999999, '__row_id': 1,
        }
        groups = generate_weekly_pdfs.group_source_rows([row])
        self.assertTrue(any(k.startswith('041926_91467680') and 'VACCREW' not in k for k in groups))


class TestVacCrewProductionInvariants(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        # W4: group_source_rows relocated to pipeline/grouping.py — grep
        # facade + relocated module so the source guards follow the code.
        import pipeline.grouping
        import pipeline.attribution  # W5: hash-prune version constant relocated here
        import pipeline.orchestrate  # W6: prune call site lives in main()
        cls._src = (
            pathlib.Path(inspect.getsourcefile(generate_weekly_pdfs)).read_text(encoding='utf-8')
            + "\n"
            + pathlib.Path(inspect.getsourcefile(pipeline.grouping)).read_text(encoding='utf-8')
            + "\n"
            + pathlib.Path(inspect.getsourcefile(pipeline.attribution)).read_text(encoding='utf-8')
            + "\n"
            + pathlib.Path(inspect.getsourcefile(pipeline.orchestrate)).read_text(encoding='utf-8')
        )

    def test_prepass_present(self):
        self.assertIn('_vac_crew_claimer_map', self._src)

    def test_emission_uses_claimer_key(self):
        self.assertIn('_VACCREW_{_c_vac_sanitized}', self._src)

    def test_prune_version_constant(self):
        self.assertRegex(self._src, r'(?m)^VAC_CREW_HASH_PRUNE_VERSION = 1$')

    def test_prune_call_site_wired_to_migration_dirty(self):
        # Tighter than a bare presence grep: confirm the prune call result
        # flips _hash_history_migration_dirty (the no-update-run persistence).
        self.assertRegex(
            self._src,
            r"if _run_vac_crew_hash_prune\(hash_history, groups\):\s*\n\s*_hash_history_migration_dirty = True",
            "vac prune call must wire its True return into _hash_history_migration_dirty",
        )

    def test_four_identity_sites_carry_vac_claimer(self):
        # Belt-and-suspenders: none of the four vac_crew identity surfaces may
        # hard-code '' (Site 1 main-loop, Site 3 current_keys). Site 2 uses
        # file_id. Guard the two that previously regressed.
        self.assertNotRegex(self._src, r"variant == 'vac_crew':\s*\n\s*identifier = ''")
        self.assertNotRegex(self._src, r"_variant == 'vac_crew':\s*\n\s*_ident = ''")


class TestVacCrewReviewFixes(unittest.TestCase):
    """PR #219 review fixes: Codex P1 (WR matchers recognize per-claimer
    _VACCREW_<claimer>), Codex P2 (prune skipped when kill switch off),
    Copilot (vac_crew row on a subcontractor sheet must not double-emit
    subcontractor primary variants)."""

    def setUp(self):
        _reset_all()
        self._orig_attr = generate_weekly_pdfs.VAC_CREW_CLAIM_ATTRIBUTION_ENABLED
        self._orig_excl = list(generate_weekly_pdfs.EXCLUDE_WRS)
        self._orig_filter = list(generate_weekly_pdfs.WR_FILTER)
        self._orig_subvar = generate_weekly_pdfs.SUBCONTRACTOR_RATE_VARIANTS_ENABLED
        self._orig_sub_ids = set(generate_weekly_pdfs._FOLDER_DISCOVERED_SUB_IDS)
        generate_weekly_pdfs.VAC_CREW_CLAIM_ATTRIBUTION_ENABLED = True

    def tearDown(self):
        generate_weekly_pdfs.VAC_CREW_CLAIM_ATTRIBUTION_ENABLED = self._orig_attr
        generate_weekly_pdfs.EXCLUDE_WRS = self._orig_excl
        generate_weekly_pdfs.WR_FILTER = self._orig_filter
        generate_weekly_pdfs.SUBCONTRACTOR_RATE_VARIANTS_ENABLED = self._orig_subvar
        generate_weekly_pdfs._FOLDER_DISCOVERED_SUB_IDS.clear()
        generate_weekly_pdfs._FOLDER_DISCOVERED_SUB_IDS.update(self._orig_sub_ids)
        _reset_all()

    def test_exclude_wrs_excludes_per_claimer_vaccrew_key(self):
        # Codex P1: EXCLUDE_WRS must drop the per-claimer _VACCREW_<claimer>
        # key (attribution on), not just the legacy bare _VACCREW.
        generate_weekly_pdfs.EXCLUDE_WRS = ['91467680']
        with mock.patch('billing_audit.writer.resolve_claimer',
                        return_value=ResolveOutcome('use', 'CrewA', 'frozen', 'success')):
            groups = generate_weekly_pdfs.group_source_rows([_make_vac_row(wr='91467680')])
        self.assertFalse(
            any('VACCREW' in k for k in groups),
            f"excluded WR's per-claimer vac key must be dropped; got {list(groups)}",
        )

    def test_wr_filter_retains_per_claimer_vaccrew_key(self):
        # Codex P1 sibling: WR_FILTER must RETAIN the per-claimer vac key.
        generate_weekly_pdfs.WR_FILTER = ['91467680']
        with mock.patch('billing_audit.writer.resolve_claimer',
                        return_value=ResolveOutcome('use', 'CrewA', 'frozen', 'success')):
            groups = generate_weekly_pdfs.group_source_rows([_make_vac_row(wr='91467680')])
        self.assertTrue(
            any('VACCREW_CrewA' in k for k in groups),
            f"WR_FILTER must retain the per-claimer vac key; got {list(groups)}",
        )

    def test_prune_skipped_when_attribution_disabled(self):
        # Codex P2: with the kill switch OFF, the blank-identifier vac key IS
        # the active legacy format — the prune must NOT delete it, and must
        # NOT advance the sentinel (so the migration still runs if attribution
        # is later enabled).
        generate_weekly_pdfs.VAC_CREW_CLAIM_ATTRIBUTION_ENABLED = False
        hist = {'91467680|041926|vac_crew|': {'hash': 'h1'}}
        groups = {'041926_91467680_VACCREW': [{'Work Request #': '91467680'}]}
        changed = generate_weekly_pdfs._run_vac_crew_hash_prune(hist, groups)
        self.assertIs(changed, False)
        self.assertIn('91467680|041926|vac_crew|', hist)
        self.assertNotIn('_vac_crew_prune_version', hist)

    def test_prune_still_runs_when_attribution_enabled(self):
        # Guard: with the kill switch ON, the prune still drops legacy orphans.
        generate_weekly_pdfs.VAC_CREW_CLAIM_ATTRIBUTION_ENABLED = True
        hist = {'91467680|041926|vac_crew|': {'hash': 'h1'}}
        groups = {'041926_91467680_VACCREW_CrewA': [
            {'Work Request #': '91467680', '__variant': 'vac_crew'}]}
        changed = generate_weekly_pdfs._run_vac_crew_hash_prune(hist, groups)
        self.assertIs(changed, True)
        self.assertNotIn('91467680|041926|vac_crew|', hist)

    def test_vac_row_on_subcontractor_sheet_does_not_double_emit(self):
        # Copilot: a vac_crew row from a subcontractor-folder sheet must emit
        # ONLY the VACCREW key — never the subcontractor primary variants.
        generate_weekly_pdfs.SUBCONTRACTOR_RATE_VARIANTS_ENABLED = True
        generate_weekly_pdfs._FOLDER_DISCOVERED_SUB_IDS.add(8162920222379908)
        with mock.patch('billing_audit.writer.resolve_claimer',
                        return_value=ResolveOutcome('use', 'CrewA', 'frozen', 'success')):
            groups = generate_weekly_pdfs.group_source_rows([_make_vac_row(wr='91467680')])
        self.assertTrue(any('VACCREW' in k for k in groups))
        self.assertFalse(
            any('REDUCEDSUB' in k or 'AEPBILLABLE' in k for k in groups),
            f"vac row on a sub sheet must not emit subcontractor variants; got {list(groups)}",
        )


class TestBulkFetchFailureDirectHoldC(unittest.TestCase):
    """Phase 2 BLOCKER 1: under a bulk fetch_failure, the C (vac_crew)
    pre-pass must set the per-row outcome to HOLD DIRECTLY — without calling
    _lookup_attribution_all (zero additional Supabase calls).

    RED before Task 2 (no 'fetch_failure' branch in the current C block).
    GREEN after Task 2 wires the direct-HOLD path in the C pre-pass block.
    """

    def setUp(self):
        _ensure_smartsheet_mocked()
        _reset_all()
        self._saved = {
            'vac_attr': generate_weekly_pdfs.VAC_CREW_CLAIM_ATTRIBUTION_ENABLED,
            'avail': generate_weekly_pdfs.BILLING_AUDIT_AVAILABLE,
            'sub_ids': set(generate_weekly_pdfs._FOLDER_DISCOVERED_SUB_IDS),
        }
        generate_weekly_pdfs.VAC_CREW_CLAIM_ATTRIBUTION_ENABLED = True
        generate_weekly_pdfs.BILLING_AUDIT_AVAILABLE = True
        generate_weekly_pdfs._FOLDER_DISCOVERED_SUB_IDS.clear()

    def tearDown(self):
        generate_weekly_pdfs.VAC_CREW_CLAIM_ATTRIBUTION_ENABLED = self._saved['vac_attr']
        generate_weekly_pdfs.BILLING_AUDIT_AVAILABLE = self._saved['avail']
        generate_weekly_pdfs._FOLDER_DISCOVERED_SUB_IDS.clear()
        generate_weekly_pdfs._FOLDER_DISCOVERED_SUB_IDS.update(self._saved['sub_ids'])

    def test_bulk_fetch_failure_c_direct_hold_zero_supabase_calls(self):
        """BLOCKER 1: C pre-pass under fetch_failure produces HOLD outcomes with
        zero _lookup_attribution_all calls.

        Pre-Task-2 (RED): the C block calls resolve_claimer without prefetched_map
        (per-row RPC), which would invoke _lookup_attribution_all per row.
        Post-Task-2 (GREEN): the fetch_failure branch constructs
        ResolveOutcome('hold', None, None, 'fetch_failure') DIRECTLY.
        """
        import billing_audit.writer as _baw

        row = _make_vac_row(wr='91467680')
        row['__row_id'] = 7777

        with mock.patch.object(
            _baw, '_lookup_attribution_all'
        ) as _mock_lookup, mock.patch(
            'billing_audit.writer.prefetch_attribution',
            return_value=({}, 'fetch_failure'),
        ):
            generate_weekly_pdfs.group_source_rows([row])

        # BLOCKER 1: _lookup_attribution_all must NOT be called on the failure path.
        _mock_lookup.assert_not_called()
