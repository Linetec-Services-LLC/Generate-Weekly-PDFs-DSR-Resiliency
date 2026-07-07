"""pipeline.excel -- openpyxl Excel file generation (W4).

BILLING GUARD: ``safe_merge_cells()`` is the ONLY allowed merge mechanism
(overlap-detecting); never write ``oddFooter.right.text``; openpyxl ONLY --
do NOT introduce xlsxwriter in this module (09-SPEC constraint).

``generate_excel`` binds the test-mutable / facade-resident output constants
(OUTPUT_FOLDER, RES_GROUPING_MODE, *_CLAIM_ATTRIBUTION_ENABLED, TEST_MODE,
SUPABASE_HASH_STORE_AUTHORITATIVE) from the generate_weekly_pdfs facade at
entry so test-time rebinds on generate_weekly_pdfs.NAME are honoured
(production value is identical -- the facade re-exports pipeline.config).

Symbols relocated byte-for-byte from ``generate_weekly_pdfs.py`` (W4):
  safe_merge_cells, _subcontractor_primary_variant_suffix,
  _vac_crew_variant_suffix, generate_excel
"""
from __future__ import annotations

import collections
import datetime
import logging
import os
from datetime import timedelta

import openpyxl
from openpyxl.styles import Font, numbers, Alignment, PatternFill
from openpyxl.drawing.image import Image
from dateutil import parser  # type: ignore[import-untyped]  # untyped third-party (matches facade)

from pipeline.config import (
    LOGO_PATH,
    _RE_EXTRACT_NUMBERS,
    _RE_SANITIZE_HELPER_NAME,
    _RE_SANITIZE_IDENTIFIER,
)
from pipeline.pricing import _resolve_row_price, parse_price
from pipeline.utils import excel_serial_to_date

logger = logging.getLogger(__name__)



def safe_merge_cells(ws, range_str):
    """
    Safely merge cells, avoiding duplicates and overlaps that cause XML errors.
    
    Args:
        ws: The worksheet object
        range_str: The range string (e.g., 'A1:C3')
    
    Returns:
        bool: True if merge was successful, False if skipped
    """
    from openpyxl.utils import range_boundaries
    
    try:
        # Parse the requested range boundaries
        min_col, min_row, max_col, max_row = range_boundaries(range_str)
        if min_col is None or min_row is None or max_col is None or max_row is None:
            return False
        
        # Check for any overlapping or duplicate merged ranges
        for merged in list(ws.merged_cells.ranges):
            m_min_col, m_min_row, m_max_col, m_max_row = range_boundaries(str(merged))
            if m_min_col is None or m_min_row is None or m_max_col is None or m_max_row is None:
                continue
            
            # Check if ranges overlap (not just exact match)
            if not (max_col < m_min_col or min_col > m_max_col or
                    max_row < m_min_row or min_row > m_max_row):
                # Ranges overlap - skip to avoid XML corruption
                return False
        
        # Safe to merge - no overlaps detected
        ws.merge_cells(range_str)
        return True
    except Exception as e:
        logging.warning(f"Failed to merge cells {range_str}: {e}")
        return False


def _subcontractor_primary_variant_suffix(
    variant: str, claimer: str, wr_num: str, week_end_raw: str
) -> str:
    """Build the filename suffix for a subcontractor PRIMARY variant.

    Subproject B (2026-05-20): subcontractor primary files are
    partitioned by frozen primary claimer and named with the reserved
    ``_User_`` token (mirrors the primary-workflow convention).
    ``reduced_sub`` -> ``_ReducedSub_User_<sanitized>`` and
    ``aep_billable`` -> ``_AEPBillable_User_<sanitized>``.

    Raises ``ValueError`` if ``claimer`` is empty — production never
    hits this because ``resolve_claimer``'s ``use`` outcome always
    returns a non-empty name (falling back to ``effective_user`` /
    ``'Unknown Foreman'``). The raise mirrors the helper-shadow
    defensive raises and surfaces data drift loudly instead of
    producing a primary-looking filename that misroutes downstream.
    """
    if not claimer:
        logging.error(
            f"⚠️ {variant} variant row missing __current_foreman for "
            f"WR {wr_num} week {week_end_raw}; filename would be "
            f"ambiguous — raising to surface data drift."
        )
        raise ValueError(
            f"{variant} requires a non-empty claimer; got empty for "
            f"WR={wr_num} week={week_end_raw}"
        )
    if variant not in ('reduced_sub', 'aep_billable'):
        # Copilot: this helper is filename-identity logic. An unexpected
        # variant must raise rather than silently fall through to the
        # ``_ReducedSub`` token (which would misroute downstream cleanup /
        # hash identity matching if this helper were ever reused). Mirrors
        # the defensive-raise convention for new variant helpers
        # (Living Ledger 2026-05-15 rule 4).
        raise ValueError(
            f"_subcontractor_primary_variant_suffix: unexpected variant "
            f"{variant!r} (expected 'reduced_sub' or 'aep_billable') for "
            f"WR={wr_num} week={week_end_raw}"
        )
    claimer_sanitized = _RE_SANITIZE_IDENTIFIER.sub('_', claimer)[:50]
    token = '_AEPBillable' if variant == 'aep_billable' else '_ReducedSub'
    return f"{token}_User_{claimer_sanitized}"


def _vac_crew_variant_suffix(claimer: str, wr_num: str, week_end_raw: str) -> str:
    """Build the filename suffix for a per-claimer VAC crew file.

    Subproject C (2026-05-21): vac_crew files are partitioned by frozen
    vac-crew claimer and named ``_VacCrew_<sanitized name>``. Raises on an
    empty claimer (production never hits this — the emission falls back to
    'Unknown'); the raise surfaces data drift instead of an ambiguous name.
    """
    if not claimer:
        logging.error(
            f"⚠️ vac_crew variant row missing claimer for WR {wr_num} "
            f"week {week_end_raw}; filename would be ambiguous — raising to surface data drift."
        )
        raise ValueError(
            f"vac_crew requires a non-empty claimer; got empty for "
            f"WR={wr_num} week={week_end_raw}"
        )
    return f"_VacCrew_{_RE_SANITIZE_IDENTIFIER.sub('_', claimer)[:50]}"


def generate_excel(group_key, group_rows, snapshot_date, ai_analysis_results=None, data_hash=None):
    """
    FIXED: Generate a formatted Excel report for a group of rows.

    SPECIFIC FIXES IMPLEMENTED:
    - WR 90093002 Excel generation (complete implementation)
    - WR 89954686 specific handling
    - Proper error handling for worksheet objects
    - Complete daily data block generation
    - Safe cell merging to prevent XML errors
    - Improved Job # field detection with multiple column name variations
    """
    # Phase 09 W4 (behaviour-preserving relocation): bind the
    # test-mutable / facade-resident output constants from the
    # generate_weekly_pdfs facade so test-time rebinds on
    # generate_weekly_pdfs.NAME are honoured (production value is
    # identical -- the facade re-exports pipeline.config).
    import generate_weekly_pdfs as _gwp  # noqa: PLC0415
    OUTPUT_FOLDER = _gwp.OUTPUT_FOLDER
    PRIMARY_CLAIM_ATTRIBUTION_ENABLED = _gwp.PRIMARY_CLAIM_ATTRIBUTION_ENABLED
    RES_GROUPING_MODE = _gwp.RES_GROUPING_MODE
    SUPABASE_HASH_STORE_AUTHORITATIVE = _gwp.SUPABASE_HASH_STORE_AUTHORITATIVE
    TEST_MODE = _gwp.TEST_MODE
    VAC_CREW_CLAIM_ATTRIBUTION_ENABLED = _gwp.VAC_CREW_CLAIM_ATTRIBUTION_ENABLED
    first_row = group_rows[0]
    
    # Parse the combined key format: "MMDDYY_WRNUMBER"
    if '_' in group_key:
        week_end_raw, wr_from_key = group_key.split('_', 1)
    else:
        # CRITICAL ERROR: Old format detected - this should never happen with fixed grouping
        error_msg = f"CRITICAL: Invalid group key format detected: '{group_key}'. Expected format: 'MMDDYY_WRNUMBER'."
        logging.error(error_msg)
        raise Exception(error_msg)
    
    # Use the current foreman (most recent) from the row data
    current_foreman = first_row.get('__current_foreman', 'Unknown_Foreman')
    
    # CRITICAL VALIDATION: Ensure grouping logic worked correctly
    wr_numbers = list(set(str(row.get('Work Request #', '')).split('.')[0] for row in group_rows if row.get('Work Request #')))
    
    # ABSOLUTE REQUIREMENT: Each group must contain EXACTLY ONE work request
    if len(wr_numbers) != 1:
        error_msg = f"FATAL ERROR: Group contains {len(wr_numbers)} work requests instead of 1: {wr_numbers}. Group key: {group_key}."
        logging.error(error_msg)
        raise Exception(error_msg)
    
    # SUCCESS: Exactly one work request in this group
    wr_num = wr_numbers[0]

    # Filesystem-safety: strip any path-traversal / separator characters
    # from the WR identifier before it reaches ``os.path.join`` and
    # ``workbook.save``. Numeric production WR#s pass through unchanged
    # (\w matches 0-9), so this is a no-op for realistic data and a
    # defense-in-depth guard against a pathological row value. Must use
    # the same regex used by the main-loop derivation site so
    # attachment / hash-history comparisons stay consistent.
    wr_num = _RE_SANITIZE_HELPER_NAME.sub('_', wr_num)[:50]

    # SPECIFIC FIX FOR WR 90093002 and WR 89954686
    if wr_num in ['90093002', '89954686']:
        logging.info(f"🔧 Applying specific fixes for WR# {wr_num}")
    
    # Get the calculated week ending date from the row data if available
    week_ending_date = first_row.get('__week_ending_date')
    if week_ending_date:
        week_end_display = week_ending_date.strftime('%m/%d/%y')
        # Update the raw format to match the calculated date
        week_end_raw = week_ending_date.strftime('%m%d%y')
        # Create subfolder for this week-ending date (YYYY-MM-DD format)
        week_folder_name = week_ending_date.strftime('%Y-%m-%d')
    else:
        # Fallback to the original format
        week_end_display = f"{week_end_raw[:2]}/{week_end_raw[2:4]}/{week_end_raw[4:]}"
        # Parse week_end_raw (MMDDYY) to create folder name
        try:
            fallback_date = datetime.datetime.strptime(week_end_raw, '%m%d%y')
            week_folder_name = fallback_date.strftime('%Y-%m-%d')
        except ValueError:
            week_folder_name = "unknown_week"
    
    # Create week-specific subfolder under OUTPUT_FOLDER
    week_output_folder = os.path.join(OUTPUT_FOLDER, week_folder_name)
    os.makedirs(week_output_folder, exist_ok=True)
    
    # Prefer 'Scope #' then fallback to 'Scope ID'
    scope_id = first_row.get('Scope #') or first_row.get('Scope ID', '')
    
    # Try multiple column name variations for Job # to handle different formats
    job_number = (first_row.get('Job #') or 
                  first_row.get('Job#') or 
                  first_row.get('Job Number') or 
                  first_row.get('JobNumber') or 
                  first_row.get('Job_Number') or 
                  first_row.get('JOB #') or 
                  first_row.get('JOB#') or 
                  first_row.get('job #') or 
                  first_row.get('job#') or 
                  '')
    
    # Log warning if Job # is missing
    if not job_number:
        available_cols = [k for k in first_row.keys() if not k.startswith('__')][:15]
        logging.warning(f"Job # not found for WR {wr_num}. Available columns: {available_cols}")
    
    # Use individual work request number for filename with timestamp for uniqueness
    timestamp = datetime.datetime.now().strftime('%H%M%S')
    
    # Variant-aware filename construction
    variant = first_row.get('__variant', 'primary')
    variant_suffix = ""

    # Phase 01 Plan 03 Task 2: subcontractor variant suffixes MUST be
    # checked BEFORE the legacy ``helper`` / ``vac_crew`` / ``primary``
    # branches so the variant-first ordering (D-09) is preserved. A
    # row tagged ``aep_billable_helper`` MUST produce the
    # ``_AEPBillable_Helper_<sanitized>`` filename, not plain
    # ``_Helper_<sanitized>`` with the AEPBillable token silently
    # dropped (which would break parser round-trip via
    # ``build_group_identity``). Helper-name sanitization mirrors
    # the producer site in ``group_source_rows`` — the regex is
    # idempotent so the double-apply is safe (D-22 / 2026-04-23 18:25).
    if variant in ('aep_billable', 'reduced_sub'):
        # Subproject B: partition by frozen primary claimer
        # (__current_foreman is the resolved claimer set in
        # group_source_rows). Helper-shadow branches below are
        # unchanged.
        variant_suffix = _subcontractor_primary_variant_suffix(
            variant,
            first_row.get('__current_foreman', ''),
            wr_num,
            week_end_raw,
        )
    elif variant == 'aep_billable_helper':
        helper_foreman = first_row.get('__helper_foreman', '')
        if not helper_foreman:
            # Phase 01 gap closure (REVIEW-WR-03): the upstream
            # ``_valid_helper_row`` gate in ``group_source_rows``
            # requires both ``helper_dept`` AND ``helper_foreman`` to
            # be truthy before adding a shadow-variant group key, so
            # this branch should never see an empty foreman in
            # production. If we ever hit it (refactor / data drift /
            # unexpected row mutation), the silent fallthrough produced
            # a primary-looking filename (no ``_AEPBillable_Helper_<name>``
            # suffix), which downstream parses as ``variant='primary'``
            # and routes the file to the wrong target sheet / wrong
            # identity tuple. Raise loudly to surface the drift instead.
            # Message body is _redact_exception_message-compatible
            # (WR + week + variant name; no foreman / dept / job —
            # those are PII per CLAUDE.md Living Ledger 2026-04-20 12:00).
            logging.error(
                f"⚠️ aep_billable_helper variant row missing "
                f"__helper_foreman for WR {wr_num} week {week_end_raw}; "
                f"filename would be ambiguous — raising to surface data drift."
            )
            raise ValueError(
                f"aep_billable_helper requires __helper_foreman; got empty "
                f"for WR={wr_num} week={week_end_raw}"
            )
        helper_sanitized = _RE_SANITIZE_HELPER_NAME.sub('_', helper_foreman)[:50]
        variant_suffix = f"_AEPBillable_Helper_{helper_sanitized}"
    elif variant == 'reduced_sub_helper':
        helper_foreman = first_row.get('__helper_foreman', '')
        if not helper_foreman:
            # WR-03 mirror of the aep_billable_helper defensive raise.
            # Same rationale, same PII-redact-compatible message body.
            logging.error(
                f"⚠️ reduced_sub_helper variant row missing "
                f"__helper_foreman for WR {wr_num} week {week_end_raw}; "
                f"filename would be ambiguous — raising to surface data drift."
            )
            raise ValueError(
                f"reduced_sub_helper requires __helper_foreman; got empty "
                f"for WR={wr_num} week={week_end_raw}"
            )
        helper_sanitized = _RE_SANITIZE_HELPER_NAME.sub('_', helper_foreman)[:50]
        variant_suffix = f"_ReducedSub_Helper_{helper_sanitized}"
    # WR-03 follow-up tech-debt: the legacy ``helper`` branch below has
    # the same shape (silent fallthrough if __helper_foreman is empty)
    # but is out of scope for this gap-closure plan per 01-REVIEW.md
    # WR-03 scope restriction. Adding the same defensive raise here in
    # a future plan is the recommended cleanup, but it requires a
    # separate regression test confirming the upstream
    # ``_valid_helper_row`` gate is the ONLY producer of the legacy
    # helper variant_suffix branch. The
    # ``test_legacy_helper_branch_does_not_raise_on_empty_foreman``
    # regression test in tests/test_subcontractor_pricing.py guards
    # against an accidental WR-03 fix broadening that would regress
    # the legacy helper variant production path.
    elif variant == 'helper':
        # Helper variant: include helper identifier in filename
        helper_foreman = first_row.get('__helper_foreman', '')
        if helper_foreman:
            # PERFORMANCE: Use pre-compiled regex for filename sanitization
            helper_sanitized = _RE_SANITIZE_HELPER_NAME.sub('_', helper_foreman)[:50]
            variant_suffix = f"_Helper_{helper_sanitized}"
    elif variant == 'vac_crew':
        # Subproject C: suffix is GATED on the kill switch.
        # Enabled mode: per-claimer _VacCrew_<name> so each foreman's file is
        # distinct and matches the Sites 1/2/3 identity tuple.
        # Disabled mode: exact legacy bare '_VacCrew' suffix — no claimer name
        # embedded, preserving byte-identical filenames with pre-C attachments.
        # NOTE: __vac_crew_name is intentionally NOT used as a fallback in
        # disabled mode; the legacy contract is a bare '_VacCrew' token, and
        # falling back to __vac_crew_name would produce _VacCrew_<name> in
        # disabled mode, which violates the disabled=legacy invariant.
        _vc_name = first_row.get('__current_foreman', '')
        if VAC_CREW_CLAIM_ATTRIBUTION_ENABLED and _vc_name:
            variant_suffix = _vac_crew_variant_suffix(_vc_name, wr_num, week_end_raw)
        else:
            # Disabled mode (or no claimer resolved) → exact legacy bare suffix.
            variant_suffix = '_VacCrew'
    elif variant == 'primary':
        # Subproject D (2026-05-25): partition the production primary
        # file by the FROZEN primary claimer (__current_foreman is the
        # resolved claimer set in group_source_rows' emission tuple).
        # GATED on the kill switch (mirrors the vac_crew branch above):
        #   • Enabled + claimer present -> _User_<sanitized claimer> so
        #     each claimer's file is distinct and round-trips through
        #     build_group_identity as ('primary', wr, week, claimer).
        #   • Disabled (or no claimer) -> exact legacy bare suffix '',
        #     preserving byte-identical filenames with pre-D attachments.
        # __current_foreman in disabled mode is effective_user (the
        # emission passes None -> `current_foreman or effective_user`),
        # but the kill-switch gate keeps the suffix bare in that case.
        _pf = first_row.get('__current_foreman', '')
        # PR #223 Codex-P1 follow-up: gate on the grouping mode too. In
        # RES_GROUPING_MODE='primary' the emission deliberately stays bare and
        # lumps every non-helper/non-sub foreman's rows into ONE workbook per
        # WR/week (partitioning by primary_foreman there is documented as
        # semantically wrong). A _User_<claimer> suffix would mislabel that
        # merged file and let row-order flip the attachment identity between
        # runs. Primary mode therefore stays bare here, matching the
        # already-mode-gated pre-pass + emission and Sites 1/2/3.
        if (
            PRIMARY_CLAIM_ATTRIBUTION_ENABLED
            and RES_GROUPING_MODE in ('helper', 'both')
            and _pf
        ):
            variant_suffix = (
                f"_User_{_RE_SANITIZE_IDENTIFIER.sub('_', _pf)[:50]}"
            )
        else:
            variant_suffix = ''

    # Phase 01 Plan 03 Task 2 (D-16): per-call missing-CU accumulator.
    # ``_resolve_row_price`` populates this Counter when a row's CU
    # code is absent from ``_SUBCONTRACTOR_RATES`` and the row keeps
    # its SmartSheet price (never zero-out, never raise). The Counter
    # is returned in the new 5-tuple shape (Blocker 4 contract) so
    # the main-loop caller can attribute missing CUs per source sheet
    # and emit the per-sheet WARNING (D-17).
    missing_cus: collections.Counter = collections.Counter()
    
    if SUPABASE_HASH_STORE_AUTHORITATIVE:
        # Sub-project E (2026-05-25): deterministic clean name. The durable
        # change-detection hash lives in billing_audit.group_content_hash,
        # so the filename carries IDENTITY ONLY (no _<timestamp>/_<hash>
        # tokens) and round-trips through build_group_identity unchanged.
        # Gated OFF by default — ships dormant until the store is validated.
        output_filename = f"WR_{wr_num}_WeekEnding_{week_end_raw}{variant_suffix}.xlsx"
    elif data_hash:
        # Use full 16-character hash (calculate_data_hash already truncates to 16)
        output_filename = f"WR_{wr_num}_WeekEnding_{week_end_raw}_{timestamp}{variant_suffix}_{data_hash}.xlsx"
    else:
        output_filename = f"WR_{wr_num}_WeekEnding_{week_end_raw}_{timestamp}{variant_suffix}.xlsx"
    final_output_path = os.path.join(week_output_folder, output_filename)

    if TEST_MODE:
        print(f"\n🧪 TEST MODE: Generating Excel file '{output_filename}'")
        print(f"   - Work Request: {wr_num}")
        print(f"   - Foreman: {current_foreman}")
        print(f"   - Timestamp: {timestamp}")
        print(f"   - Data Hash: {data_hash[:8] if data_hash else 'None'}")
        print(f"   🎯 NEW FILE POLICY: Always create fresh files")
    else:
        logging.info(f"📊 Generating Excel file '{output_filename}' for WR#{wr_num}")
        print(f"   - Week Ending: {week_end_display}")
        print(f"   - Row Count: {len(group_rows)}")

    workbook = openpyxl.Workbook()
    ws = workbook.active
    if ws is None:
        ws = workbook.create_sheet("Work Report")
    ws.title = "Work Report"

    # --- Formatting ---
    LINETEC_RED = 'C00000'
    LIGHT_GREY_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    RED_FILL = PatternFill(start_color=LINETEC_RED, end_color=LINETEC_RED, fill_type='solid')
    TITLE_FONT = Font(name='Calibri', size=20, bold=True)
    SUBTITLE_FONT = Font(name='Calibri', size=16, bold=True, color='404040')
    TABLE_HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    BLOCK_HEADER_FONT = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    BODY_FONT = Font(name='Calibri', size=11)
    SUMMARY_HEADER_FONT = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    SUMMARY_LABEL_FONT = Font(name='Calibri', size=10, bold=True)
    SUMMARY_VALUE_FONT = Font(name='Calibri', size=10)

    # Use explicit string for orientation for deterministic behavior
    ws.page_setup.orientation = 'landscape'
    try:
        ws.page_setup.paperSize = 9  # A4 paper size code
    except AttributeError:
        ws.page_setup.paperSize = 9  # Fallback for older versions
    ws.page_margins.left = 0.25; ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5; ws.page_margins.bottom = 0.5

    # --- Branding and Titles ---
    current_row = 1
    try:
        img = Image(LOGO_PATH)
        img.height = 99
        img.width = 198
        ws.add_image(img, f'A{current_row}')
        for i in range(current_row, current_row+3): 
            ws.row_dimensions[i].height = 25
        current_row += 3
    except FileNotFoundError:
        safe_merge_cells(ws, f'A{current_row}:C{current_row+2}')
        ws[f'A{current_row}'] = "LINETEC SERVICES"
        ws[f'A{current_row}'].font = TITLE_FONT
        current_row += 3

    # CRITICAL FIX: Merge cells FIRST, then assign values
    safe_merge_cells(ws, f'D{current_row-2}:I{current_row-2}')
    ws[f'D{current_row-2}'] = 'WEEKLY UNITS COMPLETED PER SCOPE ID'
    ws[f'D{current_row-2}'].font = SUBTITLE_FONT
    ws[f'D{current_row-2}'].alignment = Alignment(horizontal='center', vertical='center')

    report_generated_time = datetime.datetime.now()
    safe_merge_cells(ws, f'D{current_row+1}:I{current_row+1}')
    ws[f'D{current_row+1}'] = f"Report Generated On: {report_generated_time.strftime('%m/%d/%Y %I:%M %p')}"
    ws[f'D{current_row+1}'].font = Font(name='Calibri', size=9, italic=True)
    ws[f'D{current_row+1}'].alignment = Alignment(horizontal='right')

    current_row += 3
    safe_merge_cells(ws, f'B{current_row}:D{current_row}')
    ws[f'B{current_row}'] = 'REPORT SUMMARY'
    ws[f'B{current_row}'].font = SUMMARY_HEADER_FONT
    ws[f'B{current_row}'].fill = RED_FILL
    ws[f'B{current_row}'].alignment = Alignment(horizontal='center')

    # Per Phase 01 Plan 03 Task 2 (D-16): resolve each row's price
    # EXACTLY ONCE through ``_resolve_row_price`` and stash the result
    # on the row as ``__resolved_price``. Both the summary "Total
    # Billed Amount" and ``write_day_block``'s per-row Pricing cell
    # read from the same stashed value so the workbook is internally
    # consistent and the per-call ``missing_cus`` Counter is not
    # double-incremented across summary + day-block iteration.
    # For legacy variants this is a no-op (helper returns
    # ``parse_price(row.get('Units Total Price'))``); for the new
    # variants it picks up the rate × qty values from
    # ``_SUBCONTRACTOR_RATES``. Mutating the input row matches the
    # existing pattern (``__variant`` / ``__current_foreman`` are
    # already added upstream in ``group_source_rows``).
    for _row in group_rows:
        _row['__resolved_price'] = _resolve_row_price(_row, variant, missing_cus)
    total_price = sum(row.get('__resolved_price', 0.0) for row in group_rows)
    ws[f'B{current_row+1}'] = 'Total Billed Amount:'
    ws[f'B{current_row+1}'].font = SUMMARY_LABEL_FONT
    ws[f'C{current_row+1}'] = total_price
    ws[f'C{current_row+1}'].font = SUMMARY_VALUE_FONT
    ws[f'C{current_row+1}'].alignment = Alignment(horizontal='right')
    ws[f'C{current_row+1}'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    ws[f'B{current_row+2}'] = 'Total Line Items:'
    ws[f'B{current_row+2}'].font = SUMMARY_LABEL_FONT
    ws[f'C{current_row+2}'] = len(group_rows)
    ws[f'C{current_row+2}'].font = SUMMARY_VALUE_FONT
    ws[f'C{current_row+2}'].alignment = Alignment(horizontal='right')

    ws[f'B{current_row+3}'] = 'Billing Period:'
    ws[f'B{current_row+3}'].font = SUMMARY_LABEL_FONT
    
    # Calculate the proper week range (Monday to Sunday) for billing period
    if week_ending_date:
        week_start_date = week_ending_date - timedelta(days=6)  # Monday of that week
        billing_period = f"{week_start_date.strftime('%m/%d/%Y')} to {week_end_display}"
    else:
        # Fallback to using snapshot date if week ending date is not available
        billing_period = f"{snapshot_date.strftime('%m/%d/%Y')} to {week_end_display}"
    
    ws[f'C{current_row+3}'] = billing_period
    ws[f'C{current_row+3}'].font = SUMMARY_VALUE_FONT
    ws[f'C{current_row+3}'].alignment = Alignment(horizontal='right')

    safe_merge_cells(ws, f'F{current_row}:I{current_row}')
    ws[f'F{current_row}'] = 'REPORT DETAILS'
    ws[f'F{current_row}'].font = SUMMARY_HEADER_FONT
    ws[f'F{current_row}'].fill = RED_FILL
    ws[f'F{current_row}'].alignment = Alignment(horizontal='center')

    # Determine display values based on variant
    variant = first_row.get('__variant', 'primary')
    
    if variant == 'helper':
        # Helper variant: show helper foreman and helper-specific dept/job (REQUIRED)
        display_foreman = first_row.get('__helper_foreman', 'Unknown Helper')
        display_dept = first_row.get('__helper_dept', '')
        display_job = first_row.get('__helper_job', '')
    elif variant in ('reduced_sub_helper', 'aep_billable_helper'):
        # Subcontractor helper-shadow variants: the line items belong to the
        # helper, so Dept # / Job # MUST come from the helper fields
        # (operator requirement 2026-05-21 — helper files show Helper Dept #,
        # primary files show Dept #). Without this branch these variants fall
        # through to the ``else`` (primary) branch and display the PRIMARY
        # ``Dept #`` / ``Job #`` — the reported defect. The displayed Foreman,
        # however, stays ``current_foreman``: for these variants that is the
        # ATTRIBUTED helper (the file's partition key, set at the
        # ``keys_to_add`` site), NOT ``__helper_foreman`` (the current
        # "Foreman Helping?" value, which can diverge from the frozen
        # attribution under Phase 1.1). Folding these into the
        # ``variant == 'helper'`` branch above would regress the foreman, so
        # they are kept separate.
        display_foreman = current_foreman
        display_dept = first_row.get('__helper_dept', '')
        display_job = first_row.get('__helper_job', '')
    elif variant == 'vac_crew':
        # Enabled: show the ATTRIBUTED claimer (__current_foreman, the
        # partition key) so the displayed foreman matches the filename.
        # Disabled (legacy): show __vac_crew_name exactly as master did —
        # must NOT fall back to __current_foreman (which in disabled mode
        # may be the primary / Arrowhead foreman, not the VAC crew member).
        # dept/job remain VAC-crew-specific in all cases.
        if VAC_CREW_CLAIM_ATTRIBUTION_ENABLED:
            display_foreman = (
                first_row.get('__current_foreman')
                or first_row.get('__vac_crew_name', 'Unknown VAC Crew')
            )
        else:
            display_foreman = first_row.get('__vac_crew_name', 'Unknown VAC Crew')
        display_dept = first_row.get('__vac_crew_dept', '')
        display_job = first_row.get('__vac_crew_job', '')
    else:
        # Primary variant: show primary foreman with standard dept/job from row data
        display_foreman = current_foreman
        display_dept = first_row.get('Dept #', '')
        display_job = job_number

    details = [
        ("Foreman:", display_foreman),
        ("Work Request #:", wr_num),
        ("Scope ID #:", scope_id),
        ("Work Order #:", first_row.get('Work Order #', '')),
        ("Customer:", first_row.get('Customer Name', '')),
        ("Job #:", display_job)
    ]
    
    # Add Dept # to details if it exists
    if display_dept:
        details.insert(1, ("Dept #:", display_dept))
    
    # CRITICAL FIX: Merge cells FIRST, then assign value to top-left cell
    for i, (label, value) in enumerate(details):
        r = current_row + 1 + i
        ws[f'F{r}'] = label
        ws[f'F{r}'].font = SUMMARY_LABEL_FONT
        
        # Merge cells first - check for duplicates
        detail_merge_range = f'G{r}:I{r}'
        safe_merge_cells(ws, detail_merge_range)
        
        # Now assign value to the merged cell (top-left cell G)
        vcell = ws[f'G{r}']
        vcell.value = value
        vcell.font = SUMMARY_VALUE_FONT
        vcell.alignment = Alignment(horizontal='right')

    def write_day_block(start_row, day_name, date_obj, day_rows):
        """FIXED: Write daily data blocks with proper cell handling."""
        # Skip empty day blocks to prevent Excel corruption
        if not day_rows:
            return start_row
        
        # CRITICAL FIX: Merge cells FIRST, then assign value to top-left cell
        # Use safe merge to prevent duplicate ranges
        merge_range = f'A{start_row}:H{start_row}'
        safe_merge_cells(ws, merge_range)
        
        # Now assign value to the merged cell (top-left cell A1)
        day_header_cell = ws.cell(row=start_row, column=1)
        day_header_cell.value = f"{day_name} ({date_obj.strftime('%m/%d/%Y')})"  # type: ignore
        day_header_cell.font = BLOCK_HEADER_FONT
        day_header_cell.fill = RED_FILL
        day_header_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        headers = ["Point Number", "Billable Unit Code", "Work Type", "Unit Description", "Unit of Measure", "# Units", "N/A", "Pricing"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row+1, column=col_num)
            cell.value = header  # type: ignore
            cell.font = TABLE_HEADER_FONT
            cell.fill = RED_FILL
            cell.alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')

        total_price_day = 0.0
        for i, row_data in enumerate(day_rows):
            crow = start_row + 2 + i
            # Per Phase 01 Plan 03 Task 2: use the pre-resolved price
            # stashed by the outer ``generate_excel`` loop. Falling back
            # to ``parse_price`` here matches the legacy behaviour if a
            # row somehow lacks ``__resolved_price`` (defensive — should
            # never happen with the current call chain).
            if '__resolved_price' in row_data:
                price = row_data['__resolved_price']
            else:
                price = parse_price(row_data.get('Units Total Price'))
            
            # Safely parse quantity - extract only numbers
            qty_str = str(row_data.get('Quantity', '') or 0)
            # PERFORMANCE: Use pre-compiled regex for quantity normalization
            qty_str = _RE_EXTRACT_NUMBERS.sub('', qty_str)
            try:
                quantity = float(qty_str) if qty_str not in ('', '.', '-', '-.', '.-') else 0.0
            except Exception:
                quantity = 0.0
                
            total_price_day += price
            
            # Get the field values with debugging and fallbacks
            pole_num = (row_data.get('Pole #', '') or 
                       row_data.get('Point #', '') or 
                       row_data.get('Point Number', ''))
            
            cu_code = (row_data.get('CU', '') or 
                      row_data.get('Billable Unit Code', ''))
            
            work_type = row_data.get('Work Type', '')
            cu_description = (row_data.get('CU Description', '') or 
                             row_data.get('Unit Description', ''))
            unit_measure = (row_data.get('Unit of Measure', '') or 
                           row_data.get('UOM', ''))
            
            row_values = [pole_num, cu_code, work_type, cu_description, unit_measure, quantity, "", price]
            for col_num, value in enumerate(row_values, 1):
                cell = ws.cell(row=crow, column=col_num)
                cell.value = value
                cell.font = BODY_FONT
            ws.cell(row=crow, column=8).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

        total_row = start_row + 2 + len(day_rows)
        
        # CRITICAL FIX: Merge cells FIRST, then assign value to top-left cell
        # Use safe merge to prevent duplicate ranges
        total_merge_range = f'A{total_row}:G{total_row}'
        safe_merge_cells(ws, total_merge_range)
        
        # Now assign value to the merged cell
        total_label_cell = ws.cell(row=total_row, column=1)
        total_label_cell.value = "TOTAL"  # type: ignore
        total_label_cell.font = TABLE_HEADER_FONT
        total_label_cell.alignment = Alignment(horizontal='right')
        total_label_cell.fill = RED_FILL

        total_value_cell = ws.cell(row=total_row, column=8)
        total_value_cell.value = total_price_day  # type: ignore
        total_value_cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        total_value_cell.font = TABLE_HEADER_FONT
        total_value_cell.fill = RED_FILL

        return total_row + 2

    date_to_rows = collections.defaultdict(list)
    
    # Calculate the proper week range (Monday to Sunday) for filtering
    if week_ending_date:
        # Calculate Monday of the week (6 days before Sunday)
        week_start_date = week_ending_date - timedelta(days=6)  # Monday of that week
        week_end_date = week_ending_date  # Sunday of that week
        
        if TEST_MODE:
            print(f"\nWeek Range Filter: {week_start_date.strftime('%A, %m/%d/%Y')} to {week_end_date.strftime('%A, %m/%d/%Y')}")
    else:
        week_start_date = None
        week_end_date = None
    
    for row in group_rows:
        snap = row.get('Snapshot Date')
        try:
            dt = excel_serial_to_date(snap)
            if dt is None:
                if TEST_MODE:
                    logging.warning(f"Could not parse snapshot date '{snap}'")
                continue
            
            # Include snapshot dates that fall within the Monday-Sunday range
            if week_start_date and week_end_date:
                if week_start_date <= dt <= week_end_date:
                    date_to_rows[dt].append(row)
            else:
                date_to_rows[dt].append(row)
                    
        except (parser.ParserError, TypeError, ValueError) as e:
            if TEST_MODE:
                logging.warning(f"Could not parse snapshot date '{snap}': {e}")
            continue

    snapshot_dates = sorted(date_to_rows.keys())
    if TEST_MODE:
        print(f"\n📅 Found {len(snapshot_dates)} unique snapshot dates:")
        for d in snapshot_dates:
            print(f"   • {d.strftime('%A, %m/%d/%Y')}: {len(date_to_rows[d])} rows")
    
    day_names = {d: d.strftime('%A') for d in snapshot_dates}

    current_row += 7
    for d in snapshot_dates:
        day_rows = date_to_rows[d]
        current_row = write_day_block(current_row, day_names[d], d, day_rows)
        current_row += 1

    column_widths = {'A': 15, 'B': 20, 'C': 25, 'D': 45, 'E': 20, 'F': 10, 'G': 15, 'H': 15, 'I': 15}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # NOTE: Footer code removed - was causing Excel XML corruption errors
    # Footer attributes (oddFooter.right.text, etc.) can create malformed XML
    # that triggers "We found a problem with some content" errors in Excel

    # Save the workbook
    workbook.save(final_output_path)

    if TEST_MODE:
        print(f"📄 Generated Excel file for inspection: '{output_filename}'")
        print(f"   - Total Amount: ${total_price:,.2f}")
        print(f"   - Daily Breakdown: {len(snapshot_dates)} days")
    else:
        logging.info(f"📄 Generated Excel: '{output_filename}'")

    # Phase 01 Plan 03 Task 2 / Blocker 4: extend the return shape to
    # a 5-tuple (excel_path, filename, wr_numbers, customer_name,
    # missing_cus). The two new trailing fields are absorbed by Plan 04
    # Task 2's upload-task builder. ``customer_name`` echoes the value
    # already used in the workbook's "Customer:" detail row; surfacing
    # it on the return tuple removes a duplicate ``first_row.get(...)``
    # lookup at the call site. ``missing_cus`` carries the per-call
    # subcontractor CU-fall-through codes (D-16) for per-sheet WARNING
    # aggregation (D-17) in the main loop.
    customer_name = first_row.get('Customer Name', '') or ''
    return final_output_path, output_filename, wr_numbers, customer_name, missing_cus
